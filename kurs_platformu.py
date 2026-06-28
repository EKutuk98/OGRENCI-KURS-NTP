"""
╔══════════════════════════════════════════════════════════════════════════════╗
║   🎓 ONLINE KURS PLATFORMU — TEK DOSYA                                     ║
║   Admin + Öğrenci | PyQt5 | SQLite3 | Dark Luxury Tema                     ║
║   Çalıştır: python kurs_platformu.py                                        ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""
# Dosya adı 'platform.py' yerine 'kurs_platformu.py' olmalı —
# Python'un kendi 'platform' modülüyle çakışmaması için.
import sys as _sys, os as _os
# Mevcut dizini path'in ÖNÜNE koyma — stdlib 'platform' modülüyle çakışmasını önle
_this_dir = _os.path.dirname(_os.path.abspath(__file__))
while _this_dir in _sys.path:
    _sys.path.remove(_this_dir)

import sys, sqlite3, hashlib, os, io
from contextlib import contextmanager
from datetime import datetime

try:
    import matplotlib
    matplotlib.use("Qt5Agg")
    from matplotlib.figure import Figure as MplFigure
    import matplotlib.pyplot as plt
    import matplotlib.patches as mpatches
    from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
    MATPLOTLIB_OK = True
except ImportError:
    MATPLOTLIB_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QLineEdit, QDialog, QTabWidget, QFrame,
    QMessageBox, QScrollArea, QGridLayout, QTableWidget, QTableWidgetItem,
    QHeaderView, QComboBox, QTextEdit, QSpinBox, QCheckBox, QGroupBox,
    QAbstractItemView, QSplitter, QFormLayout, QDoubleSpinBox,
    QProgressBar, QListWidget, QListWidgetItem, QStackedWidget,
    QRadioButton, QButtonGroup, QFileDialog, QAction, QMenu,
    QShortcut, QSizePolicy, QToolButton, QInputDialog, QDesktopWidget,
    QAbstractButton
)
from PyQt5.QtCore import (
    Qt, QTimer, QPropertyAnimation, QEasingCurve, QRect,
    pyqtSignal, QSize, QPoint
)
from PyQt5.QtGui import (
    QFont, QColor, QPalette, QBrush, QPainter, QPixmap,
    QLinearGradient, QKeySequence, QPen, QIcon
)

COLORS = {
    # Admin paneli renkleri
    "bg_dark":"#0f0f1a","bg_card":"#1a1a2e","bg_surface":"#2d2d3a",
    "primary":"#0077b6","primary_h":"#005a91","secondary":"#00b4d8",
    "success":"#4CAF50","success_h":"#388E3C","warning":"#ff9f1c",
    "danger":"#e63946","danger_h":"#c62828",
    "gold":"#ffd700","silver":"#c0c0c0","bronze":"#cd7f32",
    "xp_color":"#7c3aed","level_color":"#2ec4b6",
    "text_main":"#f4f4f5","text_sec":"#a1a1aa","border":"#3f3f46",
    "tier_free":"#7c3aed","tier_pro2":"#059669","tier_pro3":"#d97706",
    # Öğrenci portalı renkleri
    "accent":"#f5a623","accent_h":"#e09018",
    "green":"#4CAF50","green_h":"#388E3C",
    "blue":"#2196F3","blue_h":"#1565C0",
    "red":"#e63946","red_h":"#c62828",
    "purple":"#9c27b0","purple_h":"#6a0080",
    "teal":"#00BCD4","xp":"#7c3aed",
    "text_dim":"#999999","bg_main":"#0f0f1a","bg_hover":"#3d3d4a",
}

# ── TEMA PALETLERİ ──────────────────────────────────────────────────────────────
TEMALAR = {
    "Okyanus Karanlığı": {
        "bg_dark":"#0f0f1a","bg_card":"#1a1a2e","bg_surface":"#2d2d3a",
        "primary":"#0077b6","secondary":"#00b4d8","accent":"#2ec4b6",
        "xp_color":"#7c3aed","level_color":"#2ec4b6","gold":"#ffd700",
    },
    "Zümrüt Gece": {
        "bg_dark":"#0a1628","bg_card":"#0f2137","bg_surface":"#1a3248",
        "primary":"#00c896","secondary":"#00e5b0","accent":"#7fffd4",
        "xp_color":"#9c27b0","level_color":"#00c896","gold":"#ffd700",
    },
    "Crimson Dark": {
        "bg_dark":"#1a0a0a","bg_card":"#2d1111","bg_surface":"#3d1a1a",
        "primary":"#e63946","secondary":"#ff6b6b","accent":"#ff9f1c",
        "xp_color":"#7c3aed","level_color":"#ff9f1c","gold":"#ffd700",
    },
    "Mor Gece": {
        "bg_dark":"#0d0a1a","bg_card":"#1a1230","bg_surface":"#2a1f45",
        "primary":"#7c3aed","secondary":"#a855f7","accent":"#ec4899",
        "xp_color":"#0077b6","level_color":"#a855f7","gold":"#ffd700",
    },
}

TIER_META = {
    "free": {"label":"FREE",   "renk":"#7c3aed","icon":"🆓","fiyat":0},
    "pro2": {"label":"PRO",    "renk":"#059669","icon":"⭐","fiyat":499},
    "pro3": {"label":"PREMIUM","renk":"#d97706","icon":"👑","fiyat":999},
}

ENHANCEMENTS = {
    "chat":       {"ad":"Canlı Chat",      "icon":"💬","fiyat":50, "tier":"free"},
    "qa_forum":   {"ad":"QA Forum",         "icon":"❓","fiyat":30, "tier":"free"},
    "kaynaklar":  {"ad":"Ekstra Kaynaklar", "icon":"📚","fiyat":20, "tier":"free"},
    "assignment": {"ad":"Proje Ödevi",      "icon":"📝","fiyat":80, "tier":"pro2"},
    "feedback":   {"ad":"Birebir Feedback", "icon":"🎯","fiyat":60, "tier":"pro2"},
    "analytics":  {"ad":"Gelişim Analitik", "icon":"📊","fiyat":40, "tier":"pro2"},
    "mentoring":  {"ad":"1-on-1 Mentoring", "icon":"🧑‍🏫","fiyat":200,"tier":"pro3"},
    "certificate":{"ad":"Sertifika",        "icon":"🏆","fiyat":100,"tier":"pro3"},
}

BADGES = {
    "ilk_adim":       {"ad":"İlk Adım",    "icon":"👣","aciklama":"İlk kursa kayıt ol",       "renk":"#7c3aed","xp":10},
    "uc_kurs":        {"ad":"Üçlü",        "icon":"📚","aciklama":"3 farklı kursa kayıt ol",   "renk":"#059669","xp":25},
    "bes_kurs":       {"ad":"Öğrenci",     "icon":"🎒","aciklama":"5 farklı kursa kayıt ol",   "renk":"#d97706","xp":50},
    "ilk_tamamlama":  {"ad":"Mezun",       "icon":"🎓","aciklama":"İlk kursu tamamla",         "renk":"#2ec4b6","xp":30},
    "uc_tamamlama":   {"ad":"Azimli",      "icon":"💪","aciklama":"3 kurs tamamla",            "renk":"#0077b6","xp":75},
    "tam_tamamlama":  {"ad":"Akademisyen", "icon":"🏛","aciklama":"5 kurs tamamla",            "renk":"#ffd700","xp":150},
    "xp_100":         {"ad":"XP Toplayıcı","icon":"⚡","aciklama":"100 XP kazan",              "renk":"#ff9f1c","xp":0},
    "xp_500":         {"ad":"XP Ustası",   "icon":"🌟","aciklama":"500 XP kazan",              "renk":"#ffd700","xp":0},
    "xp_1000":        {"ad":"XP Efsanesi", "icon":"💎","aciklama":"1000 XP kazan",             "renk":"#00b4d8","xp":0},
    "level_5":        {"ad":"Tecrübeli",   "icon":"🔥","aciklama":"Level 5'e ulaş",           "renk":"#e63946","xp":0},
    "level_10":       {"ad":"Uzman",       "icon":"🦅","aciklama":"Level 10'a ulaş",          "renk":"#d97706","xp":0},
    "ilk_enhancement":{"ad":"Yükseltici", "icon":"🛒","aciklama":"İlk enhancement satın al",  "renk":"#059669","xp":20},
    "pro_uye":        {"ad":"Pro Üye",     "icon":"💼","aciklama":"Pro pakete geç",            "renk":"#7c3aed","xp":50},
    "devamlilik":     {"ad":"Sadık Öğrenci","icon":"📅","aciklama":"7 gün üst üste aktif ol", "renk":"#2ec4b6","xp":100},
}

XP_TABLOSU = {
    "video_izle":10,"quiz_tamamla":25,"proje_gonder":50,
    "kurs_tamamla":100,"enhancement_al":15,"profil_doldur":20,
    "ilk_giris":5,"badge_kazanimi":0,
}


# ═══ ÖĞRENCİ PORTALI EK İMPORTLAR ═══



# ════════════════════════════════════════
# ADMİN PANELİ SINIFLAR (prefix: Admin*)
# ════════════════════════════════════════
def level_hesapla(xp):
    if xp <= 0: return 1
    lv = 1
    while xp >= lv*100: xp -= lv*100; lv += 1
    return lv

def level_progress(xp):
    lv = level_hesapla(xp)
    harcanan = sum(i*100 for i in range(1,lv))
    return min(100, int(((xp-harcanan)/(lv*100))*100))

def sonraki_level_xp(xp): return level_hesapla(xp)*100

def msg_info(p,b,m):  QMessageBox.information(p,b,m)
def msg_warn(p,b,m):  QMessageBox.warning(p,b,m)
def msg_error(p,b,m): QMessageBox.critical(p,b,m)
def msg_question(p,b,s): return QMessageBox.question(p,b,s)==QMessageBox.Yes
def sha256(t): return hashlib.sha256(t.encode()).hexdigest()

def tablo_stili():
    return (f"QTableWidget{{background-color:{COLORS['bg_card']};color:{COLORS['text_main']};"
            f"gridline-color:{COLORS['border']};border:none;font-size:12px;}}"
            f"QTableWidget::item{{padding:6px;}}"
            f"QTableWidget::item:selected{{background-color:{COLORS['primary']};color:white;}}"
            f"QTableWidget::item:alternate{{background-color:{COLORS['bg_surface']};}}"
            f"QHeaderView::section{{background:qlineargradient(x1:0,y1:0,x2:0,y2:1,"
            f"stop:0 {COLORS['primary']},stop:1 {COLORS['primary_h']});"
            f"color:white;font-weight:bold;padding:8px;border:none;font-size:12px;}}"
            f"QScrollBar:vertical{{background:{COLORS['bg_surface']};width:8px;border-radius:4px;}}"
            f"QScrollBar::handle:vertical{{background:{COLORS['primary']};border-radius:4px;}}")

class AdminDB:
    def __init__(self,db_name="online_kurs_master.db"):
        self.db_name=db_name; self._create_tables(); self._load_sample_data()

    @contextmanager
    def get_connection(self):
        conn=sqlite3.connect(self.db_name); conn.row_factory=sqlite3.Row
        conn.execute("PRAGMA foreign_keys=ON")
        try: yield conn; conn.commit()
        except Exception as e: conn.rollback(); raise e
        finally: conn.close()

    def _create_tables(self):
        with self.get_connection() as conn:
            c=conn.cursor()
            c.execute("""CREATE TABLE IF NOT EXISTS sistem_kullanicilar(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                kullanici_adi TEXT UNIQUE NOT NULL,sifre_hash TEXT NOT NULL,
                rol TEXT DEFAULT 'ogretmen',ad TEXT,soyad TEXT,email TEXT,
                durum TEXT DEFAULT 'aktif',olusturulma_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
            c.execute("""CREATE TABLE IF NOT EXISTS egitmenler(
                id INTEGER PRIMARY KEY AUTOINCREMENT,ad TEXT NOT NULL,uzmanlik TEXT,
                email TEXT,biyografi TEXT,rating REAL DEFAULT 5.0,
                toplam_ogrenci INTEGER DEFAULT 0,durum TEXT DEFAULT 'aktif')""")
            c.execute("""CREATE TABLE IF NOT EXISTS kurslar(
                id INTEGER PRIMARY KEY AUTOINCREMENT,kurs_kodu TEXT UNIQUE NOT NULL,
                ad TEXT NOT NULL,aciklama TEXT,egitmen_id INTEGER,
                tier TEXT DEFAULT 'free',fiyat REAL DEFAULT 0,
                kontenjan INTEGER DEFAULT 50,sure_saat INTEGER DEFAULT 20,
                kategori TEXT DEFAULT 'Genel',
                olusturulma_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(egitmen_id) REFERENCES egitmenler(id))""")
            c.execute("""CREATE TABLE IF NOT EXISTS ogrenciler(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                kullanici_adi TEXT UNIQUE NOT NULL,sifre_hash TEXT NOT NULL,
                ad TEXT,soyad TEXT,email TEXT,paket TEXT DEFAULT 'free',
                xp INTEGER DEFAULT 0,level INTEGER DEFAULT 1,
                kayit_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                son_aktif TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                durum TEXT DEFAULT 'aktif')""")
            c.execute("""CREATE TABLE IF NOT EXISTS kayitlar(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ogrenci_id INTEGER NOT NULL,kurs_id INTEGER NOT NULL,
                kayit_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                tamamlanma_tarihi TIMESTAMP,ilerleme INTEGER DEFAULT 0,
                durum TEXT DEFAULT 'devam_ediyor',xp_kazanilan INTEGER DEFAULT 0,
                FOREIGN KEY(ogrenci_id) REFERENCES ogrenciler(id),
                FOREIGN KEY(kurs_id) REFERENCES kurslar(id),
                UNIQUE(ogrenci_id,kurs_id))""")
            c.execute("""CREATE TABLE IF NOT EXISTS enhancement_kayitlar(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ogrenci_id INTEGER NOT NULL,kurs_id INTEGER NOT NULL,
                enhancement_kodu TEXT NOT NULL,
                satin_alma_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,fiyat REAL,
                FOREIGN KEY(ogrenci_id) REFERENCES ogrenciler(id),
                FOREIGN KEY(kurs_id) REFERENCES kurslar(id))""")
            c.execute("""CREATE TABLE IF NOT EXISTS odemeler(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ogrenci_id INTEGER NOT NULL,kurs_id INTEGER,
                enhancement_kodu TEXT,miktar REAL NOT NULL,
                odeme_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,aciklama TEXT,
                FOREIGN KEY(ogrenci_id) REFERENCES ogrenciler(id))""")
            c.execute("""CREATE TABLE IF NOT EXISTS xp_hareketleri(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ogrenci_id INTEGER NOT NULL,islem_turu TEXT NOT NULL,
                xp_miktari INTEGER NOT NULL,aciklama TEXT,
                tarih TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(ogrenci_id) REFERENCES ogrenciler(id))""")
            c.execute("""CREATE TABLE IF NOT EXISTS badge_kayitlar(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ogrenci_id INTEGER NOT NULL,badge_kodu TEXT NOT NULL,
                kazanilma_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(ogrenci_id) REFERENCES ogrenciler(id),
                UNIQUE(ogrenci_id,badge_kodu))""")
            c.execute("""CREATE TABLE IF NOT EXISTS aktivite_logu(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ogrenci_id INTEGER NOT NULL,islem_turu TEXT NOT NULL,
                tarih DATE DEFAULT (date('now')),
                FOREIGN KEY(ogrenci_id) REFERENCES ogrenciler(id))""")
            # TIER 4: Mentor ilişkileri
            c.execute("""CREATE TABLE IF NOT EXISTS mentor_iliskileri(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ogrenci_id INTEGER NOT NULL,
                egitmen_id INTEGER NOT NULL,
                durum TEXT DEFAULT 'beklemede',
                baslangic_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                bitis_tarihi TIMESTAMP,
                notlar TEXT,
                FOREIGN KEY(ogrenci_id) REFERENCES ogrenciler(id),
                FOREIGN KEY(egitmen_id) REFERENCES egitmenler(id),
                UNIQUE(ogrenci_id,egitmen_id))""")
            # TIER 4: Sertifika talepleri
            c.execute("""CREATE TABLE IF NOT EXISTS sertifikalar(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ogrenci_id INTEGER NOT NULL,
                kurs_id INTEGER NOT NULL,
                durum TEXT DEFAULT 'beklemede',
                talep_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                onay_tarihi TIMESTAMP,
                onaylayan_egitmen_id INTEGER,
                sertifika_no TEXT UNIQUE,
                FOREIGN KEY(ogrenci_id) REFERENCES ogrenciler(id),
                FOREIGN KEY(kurs_id) REFERENCES kurslar(id))""")
            # TIER 4: Raporlar logu
            c.execute("""CREATE TABLE IF NOT EXISTS rapor_log(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                rapor_turu TEXT NOT NULL,
                olusturulma_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                dosya_yolu TEXT)""")
            # Migration: eski DB'lerde eksik olabilecek kolonları ekle
            for tablo, kolon, tip in [
                ("sertifikalar",      "onaylayan_egitmen_id", "INTEGER"),
                ("sertifikalar",      "sertifika_no",         "TEXT"),
                ("mentor_iliskileri", "bitis_tarihi",         "TIMESTAMP"),
                ("mentor_iliskileri", "notlar",               "TEXT"),
            ]:
                try:
                    c.execute(f"ALTER TABLE {tablo} ADD COLUMN {kolon} {tip}")
                except Exception:
                    pass

    def _load_sample_data(self):
        with self.get_connection() as conn:
            c=conn.cursor()
            c.execute("SELECT COUNT(*) FROM sistem_kullanicilar")
            if c.fetchone()[0]>0: return
            c.execute("INSERT INTO sistem_kullanicilar(kullanici_adi,sifre_hash,rol,ad,soyad) VALUES(?,?,?,?,?)",
                      ("admin",sha256("admin123"),"admin","Admin","Kullanıcı"))
            for ad,uz,em,bio,rat in [
                ("Ahmet Yılmaz","Python & Web","ahmet@example.com","10 yıllık deneyim.",4.8),
                ("Zeynep Demir","UI/UX Design","zeynep@example.com","Google'da tasarımcı.",4.9),
                ("Caner Gezer","Data Science","caner@example.com","Doktora, Veri Bilimi.",4.7)]:
                c.execute("INSERT INTO egitmenler(ad,uzmanlik,email,biyografi,rating) VALUES(?,?,?,?,?)",(ad,uz,em,bio,rat))
            for row in [
                ("PY101","Python Başlangıç","Sıfırdan Python öğren",1,"free",0,30,20,"Programlama"),
                ("WEB101","HTML & CSS","Web tasarımı temelleri",1,"free",0,30,15,"Web"),
                ("UI101","UI Design Basics","Arayüz tasarımı",2,"free",0,30,12,"Tasarım"),
                ("PY201","Python İleri","Django, FastAPI, REST",1,"pro2",499,25,40,"Programlama"),
                ("WEB201","Modern Web Dev","React, Node.js",1,"pro2",499,25,50,"Web"),
                ("UI201","Advanced UI/UX","Figma, Prototyping",2,"pro2",499,25,45,"Tasarım"),
                ("DS301","Data Science İleri","TensorFlow, PyTorch",3,"pro3",999,15,80,"Veri"),
                ("ML101","Machine Learning","Uygulamalı ML",3,"pro3",999,15,60,"Veri")]:
                c.execute("INSERT INTO kurslar(kurs_kodu,ad,aciklama,egitmen_id,tier,fiyat,kontenjan,sure_saat,kategori) VALUES(?,?,?,?,?,?,?,?,?)",row)
            for row in [
                ("ali_kaya",sha256("1234"),"Ali","Kaya","ali@mail.com","free",250,3),
                ("ayse_celik",sha256("1234"),"Ayşe","Çelik","ayse@mail.com","pro2",780,5),
                ("caner_ozkan",sha256("1234"),"Caner","Özkan","caner@mail.com","pro3",1350,8),
                ("dilan_yurt",sha256("1234"),"Dilan","Yurt","dilan@mail.com","pro2",420,4),
                ("emre_sahin",sha256("1234"),"Emre","Şahin","emre@mail.com","free",80,1),
                ("fatma_koc",sha256("1234"),"Fatma","Koç","fatma@mail.com","pro2",560,4),
                ("gokhan_ay",sha256("1234"),"Gökhan","Ay","gokhan@mail.com","free",130,2)]:
                c.execute("INSERT INTO ogrenciler(kullanici_adi,sifre_hash,ad,soyad,email,paket,xp,level) VALUES(?,?,?,?,?,?,?,?)",row)
            for oid,kid,d,il,xp in [
                (1,1,"devam_ediyor",30,50),(1,2,"tamamlandi",100,100),
                (2,4,"devam_ediyor",45,80),(2,5,"tamamlandi",100,150),
                (2,1,"tamamlandi",100,100),(3,7,"devam_ediyor",20,60),
                (3,4,"tamamlandi",100,150),(3,5,"tamamlandi",100,150),
                (4,4,"devam_ediyor",70,80),(4,1,"tamamlandi",100,100),
                (5,1,"devam_ediyor",10,10),(6,4,"devam_ediyor",55,80),
                (6,5,"devam_ediyor",80,100),(7,1,"devam_ediyor",40,50)]:
                c.execute("INSERT OR IGNORE INTO kayitlar(ogrenci_id,kurs_id,durum,ilerleme,xp_kazanilan) VALUES(?,?,?,?,?)",(oid,kid,d,il,xp))
            for oid,kid,ek,fiyat in [(2,4,"assignment",80),(2,5,"feedback",60),(3,7,"mentoring",200),(3,7,"certificate",100)]:
                c.execute("INSERT INTO enhancement_kayitlar(ogrenci_id,kurs_id,enhancement_kodu,fiyat) VALUES(?,?,?,?)",(oid,kid,ek,fiyat))
            for oid,tur,xp,ac in [
                (1,"video_izle",10,"Python - Ders 1"),(1,"quiz_tamamla",25,"Quiz"),
                (1,"kurs_tamamla",100,"HTML tamamlandı"),
                (2,"kurs_tamamla",100,"Web Dev tamamlandı"),(2,"proje_gonder",50,"Final"),
                (3,"kurs_tamamla",100,"DS301"),(3,"enhancement_al",15,"Mentoring")]:
                c.execute("INSERT INTO xp_hareketleri(ogrenci_id,islem_turu,xp_miktari,aciklama) VALUES(?,?,?,?)",(oid,tur,xp,ac))
            for oid,bk in [
                (1,"ilk_adim"),(1,"xp_100"),(1,"ilk_tamamlama"),
                (2,"ilk_adim"),(2,"uc_kurs"),(2,"xp_100"),(2,"xp_500"),
                (2,"ilk_tamamlama"),(2,"uc_tamamlama"),(2,"pro_uye"),
                (3,"ilk_adim"),(3,"uc_kurs"),(3,"bes_kurs"),(3,"xp_100"),
                (3,"xp_500"),(3,"xp_1000"),(3,"ilk_tamamlama"),(3,"uc_tamamlama"),
                (3,"pro_uye"),(3,"ilk_enhancement"),
                (4,"ilk_adim"),(4,"xp_100"),(4,"pro_uye"),(5,"ilk_adim")]:
                c.execute("INSERT OR IGNORE INTO badge_kayitlar(ogrenci_id,badge_kodu) VALUES(?,?)",(oid,bk))

    def tum_kurslar_getir(self,tier_filtre=None):
        with self.get_connection() as conn:
            c=conn.cursor()
            q="SELECT k.*,e.ad as egitmen_adi FROM kurslar k LEFT JOIN egitmenler e ON k.egitmen_id=e.id"
            if tier_filtre and tier_filtre!="hepsi":
                c.execute(q+" WHERE k.tier=? ORDER BY k.tier,k.ad",(tier_filtre,))
            else: c.execute(q+" ORDER BY k.tier,k.ad")
            return c.fetchall()

    def tum_egitmenler_getir(self):
        with self.get_connection() as conn:
            c=conn.cursor(); c.execute("SELECT * FROM egitmenler WHERE durum='aktif' ORDER BY ad"); return c.fetchall()

    def tum_ogrenciler_getir(self):
        with self.get_connection() as conn:
            c=conn.cursor()
            c.execute("""SELECT o.*,COUNT(k.id) as kurs_sayisi FROM ogrenciler o
                LEFT JOIN kayitlar k ON o.id=k.ogrenci_id
                WHERE o.durum='aktif' GROUP BY o.id ORDER BY o.xp DESC""")
            return c.fetchall()

    def ogrenci_kayitlar_getir(self,ogrenci_id):
        with self.get_connection() as conn:
            c=conn.cursor()
            c.execute("""SELECT k.*,kr.ad as kurs_adi,kr.tier,kr.fiyat
                FROM kayitlar k JOIN kurslar kr ON k.kurs_id=kr.id
                WHERE k.ogrenci_id=? ORDER BY k.kayit_tarihi DESC""",(ogrenci_id,))
            return c.fetchall()

    def ogrenci_enhancements_getir(self,ogrenci_id,kurs_id):
        with self.get_connection() as conn:
            c=conn.cursor()
            c.execute("SELECT enhancement_kodu FROM enhancement_kayitlar WHERE ogrenci_id=? AND kurs_id=?",(ogrenci_id,kurs_id))
            return [r[0] for r in c.fetchall()]

    def dashboard_istatistik(self):
        with self.get_connection() as conn:
            c=conn.cursor()
            return {
                "toplam_ogrenci":c.execute("SELECT COUNT(*) FROM ogrenciler WHERE durum='aktif'").fetchone()[0],
                "toplam_kurs":c.execute("SELECT COUNT(*) FROM kurslar").fetchone()[0],
                "toplam_egitmen":c.execute("SELECT COUNT(*) FROM egitmenler WHERE durum='aktif'").fetchone()[0],
                "toplam_kayit":c.execute("SELECT COUNT(*) FROM kayitlar").fetchone()[0],
                "pro_ogrenci":c.execute("SELECT COUNT(*) FROM ogrenciler WHERE paket!='free'").fetchone()[0],
                "tamamlanan":c.execute("SELECT COUNT(*) FROM kayitlar WHERE durum='tamamlandi'").fetchone()[0],
                "toplam_xp":c.execute("SELECT COALESCE(SUM(xp),0) FROM ogrenciler").fetchone()[0],
                "toplam_badge":c.execute("SELECT COUNT(*) FROM badge_kayitlar").fetchone()[0],
            }

    def xp_ekle(self,ogrenci_id,islem_turu,aciklama=""):
        miktar=XP_TABLOSU.get(islem_turu,0)
        if miktar<=0 and islem_turu!="badge_kazanimi": return 0,[]
        with self.get_connection() as conn:
            c=conn.cursor()
            if miktar>0:
                c.execute("INSERT INTO xp_hareketleri(ogrenci_id,islem_turu,xp_miktari,aciklama) VALUES(?,?,?,?)",(ogrenci_id,islem_turu,miktar,aciklama))
                c.execute("UPDATE ogrenciler SET xp=xp+? WHERE id=?",(miktar,ogrenci_id))
            yeni_xp=c.execute("SELECT xp FROM ogrenciler WHERE id=?",(ogrenci_id,)).fetchone()[0]
            c.execute("UPDATE ogrenciler SET level=? WHERE id=?",(level_hesapla(yeni_xp),ogrenci_id))
            c.execute("INSERT INTO aktivite_logu(ogrenci_id,islem_turu) VALUES(?,?)",(ogrenci_id,islem_turu))
        return miktar,self._badge_kontrol(ogrenci_id)

    def xp_hareketleri_getir(self,ogrenci_id,limit=20):
        with self.get_connection() as conn:
            c=conn.cursor()
            c.execute("SELECT * FROM xp_hareketleri WHERE ogrenci_id=? ORDER BY tarih DESC LIMIT ?",(ogrenci_id,limit))
            return c.fetchall()

    def ogrenci_xp_getir(self,ogrenci_id):
        with self.get_connection() as conn:
            c=conn.cursor()
            row=c.execute("SELECT xp,level FROM ogrenciler WHERE id=?",(ogrenci_id,)).fetchone()
            return (row["xp"],row["level"]) if row else (0,1)

    def _badge_kontrol(self,ogrenci_id):
        with self.get_connection() as conn:
            c=conn.cursor()
            xp=c.execute("SELECT xp FROM ogrenciler WHERE id=?",(ogrenci_id,)).fetchone()[0]
            lv=level_hesapla(xp)
            kayit=c.execute("SELECT COUNT(*) FROM kayitlar WHERE ogrenci_id=?",(ogrenci_id,)).fetchone()[0]
            tamam=c.execute("SELECT COUNT(*) FROM kayitlar WHERE ogrenci_id=? AND durum='tamamlandi'",(ogrenci_id,)).fetchone()[0]
            enh=c.execute("SELECT COUNT(*) FROM enhancement_kayitlar WHERE ogrenci_id=?",(ogrenci_id,)).fetchone()[0]
            paket=c.execute("SELECT paket FROM ogrenciler WHERE id=?",(ogrenci_id,)).fetchone()[0]
            mevcut={r[0] for r in c.execute("SELECT badge_kodu FROM badge_kayitlar WHERE ogrenci_id=?",(ogrenci_id,)).fetchall()}
            kontroller={"ilk_adim":kayit>=1,"uc_kurs":kayit>=3,"bes_kurs":kayit>=5,
                "ilk_tamamlama":tamam>=1,"uc_tamamlama":tamam>=3,"tam_tamamlama":tamam>=5,
                "xp_100":xp>=100,"xp_500":xp>=500,"xp_1000":xp>=1000,
                "level_5":lv>=5,"level_10":lv>=10,"ilk_enhancement":enh>=1,
                "pro_uye":paket in("pro2","pro3")}
            kazanildi=[]
            for bk,kosul in kontroller.items():
                if kosul and bk not in mevcut:
                    c.execute("INSERT OR IGNORE INTO badge_kayitlar(ogrenci_id,badge_kodu) VALUES(?,?)",(ogrenci_id,bk))
                    bonus=BADGES[bk].get("xp",0)
                    if bonus>0:
                        c.execute("INSERT INTO xp_hareketleri(ogrenci_id,islem_turu,xp_miktari,aciklama) VALUES(?,?,?,?)",
                                  (ogrenci_id,"badge_kazanimi",bonus,f"Badge: {BADGES[bk]['ad']}"))
                        c.execute("UPDATE ogrenciler SET xp=xp+? WHERE id=?",(bonus,ogrenci_id))
                    kazanildi.append(bk)
        return kazanildi

    def ogrenci_badge_getir(self,ogrenci_id):
        with self.get_connection() as conn:
            c=conn.cursor()
            c.execute("SELECT badge_kodu,kazanilma_tarihi FROM badge_kayitlar WHERE ogrenci_id=? ORDER BY kazanilma_tarihi",(ogrenci_id,))
            return c.fetchall()

    def leaderboard_getir(self,limit=10):
        with self.get_connection() as conn:
            c=conn.cursor()
            c.execute("""SELECT o.id,o.ad,o.soyad,o.paket,o.xp,o.level,
                COUNT(DISTINCT b.badge_kodu) as badge_sayisi,
                COUNT(DISTINCT k.id) as tamamlanan
                FROM ogrenciler o
                LEFT JOIN badge_kayitlar b ON o.id=b.ogrenci_id
                LEFT JOIN kayitlar k ON o.id=k.ogrenci_id AND k.durum='tamamlandi'
                WHERE o.durum='aktif' GROUP BY o.id ORDER BY o.xp DESC LIMIT ?""",(limit,))
            return c.fetchall()

    def kurs_ekle(self,kod,ad,aciklama,egitmen_id,tier,fiyat,kontenjan,sure,kategori):
        with self.get_connection() as conn:
            try:
                conn.execute("INSERT INTO kurslar(kurs_kodu,ad,aciklama,egitmen_id,tier,fiyat,kontenjan,sure_saat,kategori) VALUES(?,?,?,?,?,?,?,?,?)",
                             (kod,ad,aciklama,egitmen_id,tier,fiyat,kontenjan,sure,kategori))
                return True,"Kurs başarıyla eklendi."
            except sqlite3.IntegrityError: return False,"Bu kurs kodu zaten kullanılıyor!"

    def kurs_guncelle(self,kurs_id,ad,aciklama,egitmen_id,tier,fiyat,kontenjan,sure,kategori):
        with self.get_connection() as conn:
            conn.execute("UPDATE kurslar SET ad=?,aciklama=?,egitmen_id=?,tier=?,fiyat=?,kontenjan=?,sure_saat=?,kategori=? WHERE id=?",
                         (ad,aciklama,egitmen_id,tier,fiyat,kontenjan,sure,kategori,kurs_id))
            return True,"Kurs güncellendi."

    def kurs_sil(self,kurs_id):
        with self.get_connection() as conn:
            conn.execute("DELETE FROM kayitlar WHERE kurs_id=?",(kurs_id,))
            conn.execute("DELETE FROM kurslar WHERE id=?",(kurs_id,))
            return True,"Kurs silindi."

    def egitmen_ekle(self,ad,uzmanlik,email,biyografi):
        with self.get_connection() as conn:
            try:
                conn.execute("INSERT INTO egitmenler(ad,uzmanlik,email,biyografi) VALUES(?,?,?,?)",(ad,uzmanlik,email,biyografi))
                return True,"Eğitmen eklendi."
            except sqlite3.IntegrityError: return False,"Bu eğitmen zaten mevcut."

    def egitmen_guncelle(self,eid,ad,uzmanlik,email,biyografi):
        with self.get_connection() as conn:
            conn.execute("UPDATE egitmenler SET ad=?,uzmanlik=?,email=?,biyografi=? WHERE id=?",(ad,uzmanlik,email,biyografi,eid))
            return True,"Eğitmen güncellendi."

    def egitmen_sil(self,eid):
        with self.get_connection() as conn:
            conn.execute("UPDATE egitmenler SET durum='pasif' WHERE id=?",(eid,))
            return True,"Eğitmen pasife alındı."

    def ogrenci_ekle(self,kullanici_adi,sifre,ad,soyad,email,paket="free"):
        with self.get_connection() as conn:
            try:
                conn.execute("INSERT INTO ogrenciler(kullanici_adi,sifre_hash,ad,soyad,email,paket) VALUES(?,?,?,?,?,?)",
                             (kullanici_adi,sha256(sifre),ad,soyad,email,paket))
                return True,"Öğrenci eklendi."
            except sqlite3.IntegrityError: return False,"Bu kullanıcı adı alınmış."

    def ogrenci_guncelle(self,oid,ad,soyad,email,paket):
        with self.get_connection() as conn:
            conn.execute("UPDATE ogrenciler SET ad=?,soyad=?,email=?,paket=? WHERE id=?",(ad,soyad,email,paket,oid))
            return True,"Öğrenci güncellendi."

    def ogrenci_sil(self,oid):
        with self.get_connection() as conn:
            conn.execute("UPDATE ogrenciler SET durum='pasif' WHERE id=?",(oid,)); return True,"Öğrenci pasife alındı."

    def kursa_kaydet(self,ogrenci_id,kurs_id):
        with self.get_connection() as conn:
            try:
                row=conn.execute("SELECT k.kontenjan,COUNT(ka.id) as dolu FROM kurslar k LEFT JOIN kayitlar ka ON k.id=ka.kurs_id WHERE k.id=? GROUP BY k.id",(kurs_id,)).fetchone()
                if row and row["dolu"]>=row["kontenjan"]: return False,"Kontenjan dolu!"
                conn.execute("INSERT INTO kayitlar(ogrenci_id,kurs_id) VALUES(?,?)",(ogrenci_id,kurs_id))
                return True,"Öğrenci kursa kaydedildi."
            except sqlite3.IntegrityError: return False,"Öğrenci bu kursa zaten kayıtlı."

    def ilerleme_guncelle(self,ogrenci_id,kurs_id,yeni_ilerleme):
        with self.get_connection() as conn:
            c=conn.cursor()
            if not c.execute("SELECT id FROM kayitlar WHERE ogrenci_id=? AND kurs_id=?",(ogrenci_id,kurs_id)).fetchone():
                return False,"Kayıt bulunamadı."
            durum="tamamlandi" if yeni_ilerleme>=100 else "devam_ediyor"
            tamam=datetime.now().isoformat() if yeni_ilerleme>=100 else None
            c.execute("UPDATE kayitlar SET ilerleme=?,durum=?,tamamlanma_tarihi=? WHERE ogrenci_id=? AND kurs_id=?",
                      (min(100,yeni_ilerleme),durum,tamam,ogrenci_id,kurs_id))
            return True,durum

    def enhancement_satin_al(self,ogrenci_id,kurs_id,enhancement_kodu):
        with self.get_connection() as conn:
            if conn.execute("SELECT id FROM enhancement_kayitlar WHERE ogrenci_id=? AND kurs_id=? AND enhancement_kodu=?",
                            (ogrenci_id,kurs_id,enhancement_kodu)).fetchone():
                return False,"Bu enhancement zaten aktif."
            fiyat=ENHANCEMENTS.get(enhancement_kodu,{}).get("fiyat",0)
            conn.execute("INSERT INTO enhancement_kayitlar(ogrenci_id,kurs_id,enhancement_kodu,fiyat) VALUES(?,?,?,?)",(ogrenci_id,kurs_id,enhancement_kodu,fiyat))
            conn.execute("INSERT INTO odemeler(ogrenci_id,kurs_id,enhancement_kodu,miktar,aciklama) VALUES(?,?,?,?,?)",
                         (ogrenci_id,kurs_id,enhancement_kodu,fiyat,f"{ENHANCEMENTS[enhancement_kodu]['ad']} satın alındı"))
            return True,f"{ENHANCEMENTS[enhancement_kodu]['ad']} aktive edildi."

    # ── MENTOR ────────────────────────────────────────────────────────────
    def mentor_iliskileri_getir(self,filtre=None):
        with self.get_connection() as conn:
            c=conn.cursor()
            q="""SELECT mi.*,o.ad as ogr_ad,o.soyad as ogr_soyad,o.paket,
                e.ad as eg_ad,e.uzmanlik
                FROM mentor_iliskileri mi
                JOIN ogrenciler o ON mi.ogrenci_id=o.id
                JOIN egitmenler e ON mi.egitmen_id=e.id"""
            if filtre and filtre!="hepsi":
                c.execute(q+" WHERE mi.durum=? ORDER BY mi.baslangic_tarihi DESC",(filtre,))
            else:
                c.execute(q+" ORDER BY mi.baslangic_tarihi DESC")
            return c.fetchall()

    def mentor_ata(self,ogrenci_id,egitmen_id,notlar=""):
        with self.get_connection() as conn:
            try:
                conn.execute("INSERT INTO mentor_iliskileri(ogrenci_id,egitmen_id,durum,notlar) VALUES(?,?,'aktif',?)",
                             (ogrenci_id,egitmen_id,notlar))
                return True,"Mentor başarıyla atandı."
            except sqlite3.IntegrityError:
                return False,"Bu öğrenciye bu eğitmen zaten atanmış."

    def mentor_durum_guncelle(self,ilişki_id,yeni_durum):
        with self.get_connection() as conn:
            conn.execute("UPDATE mentor_iliskileri SET durum=? WHERE id=?",(yeni_durum,ilişki_id))
            return True,f"Durum güncellendi: {yeni_durum}"

    def mentor_sil(self,iliski_id):
        with self.get_connection() as conn:
            conn.execute("DELETE FROM mentor_iliskileri WHERE id=?",(iliski_id,))
            return True,"Mentor ilişkisi silindi."

    # ── SERTİFİKA ─────────────────────────────────────────────────────────
    def sertifika_talep_getir(self,filtre=None):
        with self.get_connection() as conn:
            c=conn.cursor()
            q="""SELECT s.*,o.ad as ogr_ad,o.soyad as ogr_soyad,
                k.ad as kurs_adi,k.tier,
                e.ad as eg_ad
                FROM sertifikalar s
                JOIN ogrenciler o ON s.ogrenci_id=o.id
                JOIN kurslar k ON s.kurs_id=k.id
                LEFT JOIN egitmenler e ON s.onaylayan_egitmen_id=e.id"""
            if filtre and filtre!="hepsi":
                c.execute(q+" WHERE s.durum=? ORDER BY s.talep_tarihi DESC",(filtre,))
            else:
                c.execute(q+" ORDER BY s.talep_tarihi DESC")
            return c.fetchall()

    def sertifika_talep_olustur(self,ogrenci_id,kurs_id):
        with self.get_connection() as conn:
            # Kurs tamamlandı mı kontrol et
            row=conn.execute("SELECT durum FROM kayitlar WHERE ogrenci_id=? AND kurs_id=?",(ogrenci_id,kurs_id)).fetchone()
            if not row or row["durum"]!="tamamlandi":
                return False,"Kurs henüz tamamlanmamış!"
            # Zaten talep var mı
            mevcut=conn.execute("SELECT id FROM sertifikalar WHERE ogrenci_id=? AND kurs_id=?",(ogrenci_id,kurs_id)).fetchone()
            if mevcut:
                return False,"Bu kurs için zaten sertifika talebi var."
            conn.execute("INSERT INTO sertifikalar(ogrenci_id,kurs_id) VALUES(?,?)",(ogrenci_id,kurs_id))
            return True,"Sertifika talebi oluşturuldu."

    def sertifika_onayla(self,sertifika_id,egitmen_id):
        with self.get_connection() as conn:
            import random,string
            sno="CERT-"+"".join(random.choices(string.ascii_uppercase+string.digits,k=8))
            conn.execute("""UPDATE sertifikalar SET durum='onaylandi',onay_tarihi=CURRENT_TIMESTAMP,
                onaylayan_egitmen_id=?,sertifika_no=? WHERE id=?""",(egitmen_id,sno,sertifika_id))
            return True,f"Sertifika onaylandı! No: {sno}"

    def sertifika_reddet(self,sertifika_id):
        with self.get_connection() as conn:
            conn.execute("UPDATE sertifikalar SET durum='reddedildi' WHERE id=?",(sertifika_id,))
            return True,"Sertifika talebi reddedildi."

    def sertifika_bilgi_getir(self,sertifika_id):
        with self.get_connection() as conn:
            c=conn.cursor()
            return c.execute("""SELECT s.*,o.ad as ogr_ad,o.soyad as ogr_soyad,o.email as ogr_email,
                k.ad as kurs_adi,k.tier,k.sure_saat,
                e.ad as eg_ad,e.uzmanlik,
                (SELECT COUNT(*) FROM badge_kayitlar WHERE ogrenci_id=s.ogrenci_id) as badge_sayisi,
                (SELECT xp FROM ogrenciler WHERE id=s.ogrenci_id) as toplam_xp
                FROM sertifikalar s
                JOIN ogrenciler o ON s.ogrenci_id=o.id
                JOIN kurslar k ON s.kurs_id=k.id
                LEFT JOIN egitmenler e ON s.onaylayan_egitmen_id=e.id
                WHERE s.id=?""",(sertifika_id,)).fetchone()

    # ── İSTATİSTİK ────────────────────────────────────────────────────────
    def grafik_verileri_getir(self):
        with self.get_connection() as conn:
            c=conn.cursor()
            # Tier dağılımı
            tier_data={}
            for row in c.execute("SELECT paket as tier,COUNT(*) as cnt FROM ogrenciler GROUP BY paket").fetchall():
                tier_data[row["tier"]]=row["cnt"]
            # Kurs tamamlanma oranları
            kurs_tamamlanma=c.execute("""SELECT k.ad,COUNT(ka.id) as toplam,
                SUM(CASE WHEN ka.durum='tamamlandi' THEN 1 ELSE 0 END) as tamam
                FROM kurslar k LEFT JOIN kayitlar ka ON k.id=ka.kurs_id
                GROUP BY k.id ORDER BY toplam DESC LIMIT 6""").fetchall()
            # XP dağılımı (top öğrenciler)
            xp_data=c.execute("SELECT ad,xp FROM ogrenciler WHERE durum='aktif' ORDER BY xp DESC LIMIT 7").fetchall()
            # Aylık kayıt trendi (son 6 ay)
            kayit_trend=c.execute("""SELECT strftime('%m/%Y',kayit_tarihi) as ay,COUNT(*) as cnt
                FROM kayitlar WHERE kayit_tarihi >= date('now','-6 months')
                GROUP BY ay ORDER BY ay""").fetchall()
            return {"tier_data":tier_data,"kurs_tamamlanma":kurs_tamamlanma,
                    "xp_data":xp_data,"kayit_trend":kayit_trend}

    def excel_rapor_olustur(self,dosya_yolu):
        if not OPENPYXL_OK: return False,"openpyxl kurulu değil."
        wb=openpyxl.Workbook()
        # Öğrenciler sayfası
        ws=wb.active; ws.title="Öğrenciler"
        basliklar=["ID","Ad","Soyad","Kullanıcı","Paket","XP","Level","Durum"]
        for col,b in enumerate(basliklar,1):
            cell=ws.cell(row=1,column=col,value=b)
            cell.font=Font(bold=True,color="FFFFFF")
            cell.fill=PatternFill("solid",fgColor="0077B6")
            cell.alignment=Alignment(horizontal="center")
        with self.get_connection() as conn:
            c=conn.cursor()
            for row,ogr in enumerate(c.execute("SELECT id,ad,soyad,kullanici_adi,paket,xp,level,durum FROM ogrenciler").fetchall(),2):
                for col,val in enumerate(list(ogr),1):
                    ws.cell(row=row,column=col,value=val)
        # Kurslar sayfası
        ws2=wb.create_sheet("Kurslar")
        basliklar2=["ID","Kod","Ad","Tier","Fiyat","Kontenjan","Süre"]
        for col,b in enumerate(basliklar2,1):
            cell=ws2.cell(row=1,column=col,value=b)
            cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="059669")
        with self.get_connection() as conn:
            c=conn.cursor()
            for row,k in enumerate(c.execute("SELECT id,kurs_kodu,ad,tier,fiyat,kontenjan,sure_saat FROM kurslar").fetchall(),2):
                for col,val in enumerate(list(k),1):
                    ws2.cell(row=row,column=col,value=val)
        # Sertifikalar sayfası
        ws3=wb.create_sheet("Sertifikalar")
        basliklar3=["Sertifika No","Öğrenci","Kurs","Durum","Talep Tarihi","Onay Tarihi"]
        for col,b in enumerate(basliklar3,1):
            cell=ws3.cell(row=1,column=col,value=b)
            cell.font=Font(bold=True,color="FFFFFF"); cell.fill=PatternFill("solid",fgColor="D97706")
        with self.get_connection() as conn:
            c=conn.cursor()
            for row,s in enumerate(c.execute("""SELECT s.sertifika_no,o.ad||' '||COALESCE(o.soyad,''),
                k.ad,s.durum,s.talep_tarihi,s.onay_tarihi FROM sertifikalar s
                JOIN ogrenciler o ON s.ogrenci_id=o.id JOIN kurslar k ON s.kurs_id=k.id""").fetchall(),2):
                for col,val in enumerate(list(s),1):
                    ws3.cell(row=row,column=col,value=val)
        wb.save(dosya_yolu)
        return True,dosya_yolu

    # ── TIER 5: Bildirimler ──────────────────────────────────────────────
    def bildirim_ekle(self,baslik,mesaj,tur="bilgi",hedef_id=None):
        with self.get_connection() as conn:
            conn.execute("""CREATE TABLE IF NOT EXISTS bildirimler(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                baslik TEXT NOT NULL,mesaj TEXT,
                tur TEXT DEFAULT 'bilgi',
                hedef_ogrenci_id INTEGER,
                okundu INTEGER DEFAULT 0,
                tarih TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
            conn.execute("INSERT INTO bildirimler(baslik,mesaj,tur,hedef_ogrenci_id) VALUES(?,?,?,?)",
                         (baslik,mesaj,tur,hedef_id))

    def bildirimler_getir(self,okunmamis_only=False):
        with self.get_connection() as conn:
            conn.execute("""CREATE TABLE IF NOT EXISTS bildirimler(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                baslik TEXT NOT NULL,mesaj TEXT,
                tur TEXT DEFAULT 'bilgi',
                hedef_ogrenci_id INTEGER,
                okundu INTEGER DEFAULT 0,
                tarih TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
            c=conn.cursor()
            q="SELECT * FROM bildirimler"
            if okunmamis_only: q+=" WHERE okundu=0"
            c.execute(q+" ORDER BY tarih DESC LIMIT 50")
            return c.fetchall()

    def bildirim_oku(self,bildirim_id=None):
        with self.get_connection() as conn:
            if bildirim_id:
                conn.execute("UPDATE bildirimler SET okundu=1 WHERE id=?",(bildirim_id,))
            else:
                conn.execute("UPDATE bildirimler SET okundu=1")

    def okunmamis_bildirim_sayisi(self):
        with self.get_connection() as conn:
            try:
                return conn.execute("SELECT COUNT(*) FROM bildirimler WHERE okundu=0").fetchone()[0]
            except: return 0

    # ── TIER 5: Ayarlar ──────────────────────────────────────────────────
    def ayar_getir(self,anahtar,varsayilan=""):
        with self.get_connection() as conn:
            conn.execute("""CREATE TABLE IF NOT EXISTS ayarlar(
                anahtar TEXT PRIMARY KEY,deger TEXT,
                guncelleme_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
            row=conn.execute("SELECT deger FROM ayarlar WHERE anahtar=?",(anahtar,)).fetchone()
            return row["deger"] if row else varsayilan

    def ayar_kaydet(self,anahtar,deger):
        with self.get_connection() as conn:
            conn.execute("""CREATE TABLE IF NOT EXISTS ayarlar(
                anahtar TEXT PRIMARY KEY,deger TEXT,
                guncelleme_tarihi TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
            conn.execute("INSERT OR REPLACE INTO ayarlar(anahtar,deger) VALUES(?,?)",(anahtar,deger))

    # ── TIER 5: Global arama ─────────────────────────────────────────────
    def global_ara(self,kelime):
        if not kelime or len(kelime)<2: return {}
        k=f"%{kelime}%"
        with self.get_connection() as conn:
            c=conn.cursor()
            ogr=c.execute("SELECT id,ad,soyad,paket,xp FROM ogrenciler WHERE (ad LIKE ? OR soyad LIKE ? OR kullanici_adi LIKE ?) AND durum='aktif'",(k,k,k)).fetchall()
            kurs=c.execute("SELECT id,kurs_kodu,ad,tier FROM kurslar WHERE ad LIKE ? OR kurs_kodu LIKE ? OR aciklama LIKE ?",(k,k,k)).fetchall()
            egit=c.execute("SELECT id,ad,uzmanlik FROM egitmenler WHERE ad LIKE ? OR uzmanlik LIKE ?",(k,k)).fetchall()
            return {"ogrenciler":ogr,"kurslar":kurs,"egitmenler":egit}

    # ── TIER 5: Kurs detay istatistikleri ────────────────────────────────
    def kurs_detay_getir(self,kurs_id):
        with self.get_connection() as conn:
            c=conn.cursor()
            kurs=c.execute("SELECT k.*,e.ad as eg_ad,e.uzmanlik,e.rating FROM kurslar k LEFT JOIN egitmenler e ON k.egitmen_id=e.id WHERE k.id=?",(kurs_id,)).fetchone()
            kayit_sayisi=c.execute("SELECT COUNT(*) FROM kayitlar WHERE kurs_id=?",(kurs_id,)).fetchone()[0]
            tamam_sayisi=c.execute("SELECT COUNT(*) FROM kayitlar WHERE kurs_id=? AND durum='tamamlandi'",(kurs_id,)).fetchone()[0]
            ort_ilerleme=c.execute("SELECT AVG(ilerleme) FROM kayitlar WHERE kurs_id=?",(kurs_id,)).fetchone()[0] or 0
            toplam_xp=c.execute("SELECT COALESCE(SUM(xp_kazanilan),0) FROM kayitlar WHERE kurs_id=?",(kurs_id,)).fetchone()[0]
            enh_sayisi=c.execute("SELECT COUNT(*) FROM enhancement_kayitlar WHERE kurs_id=?",(kurs_id,)).fetchone()[0]
            return {"kurs":kurs,"kayit_sayisi":kayit_sayisi,"tamam_sayisi":tamam_sayisi,
                    "ort_ilerleme":round(ort_ilerleme,1),"toplam_xp":toplam_xp,"enh_sayisi":enh_sayisi}

    # ── TIER 6: Profil & İstatistik ────────────────────────────────────────
    def ogrenci_profil_getir(self, ogrenci_id):
        with self.get_connection() as conn:
            c = conn.cursor()
            ogr = c.execute("SELECT * FROM ogrenciler WHERE id=?", (ogrenci_id,)).fetchone()
            if not ogr: return None
            kayit_sayisi   = c.execute("SELECT COUNT(*) FROM kayitlar WHERE ogrenci_id=?", (ogrenci_id,)).fetchone()[0]
            tamam_sayisi   = c.execute("SELECT COUNT(*) FROM kayitlar WHERE ogrenci_id=? AND durum='tamamlandi'", (ogrenci_id,)).fetchone()[0]
            badge_sayisi   = c.execute("SELECT COUNT(*) FROM badge_kayitlar WHERE ogrenci_id=?", (ogrenci_id,)).fetchone()[0]
            enh_sayisi     = c.execute("SELECT COUNT(*) FROM enhancement_kayitlar WHERE ogrenci_id=?", (ogrenci_id,)).fetchone()[0]
            mentor_var     = c.execute("SELECT COUNT(*) FROM mentor_iliskileri WHERE ogrenci_id=? AND durum='aktif'", (ogrenci_id,)).fetchone()[0]
            sertifika_sayisi = c.execute("SELECT COUNT(*) FROM sertifikalar WHERE ogrenci_id=? AND durum='onaylandi'", (ogrenci_id,)).fetchone()[0]
            son_hareketler = c.execute(
                "SELECT islem_turu,xp_miktari,aciklama,tarih FROM xp_hareketleri WHERE ogrenci_id=? ORDER BY tarih DESC LIMIT 5",
                (ogrenci_id,)).fetchall()
            return {"ogr": ogr, "kayit_sayisi": kayit_sayisi, "tamam_sayisi": tamam_sayisi,
                    "badge_sayisi": badge_sayisi, "enh_sayisi": enh_sayisi,
                    "mentor_var": mentor_var, "sertifika_sayisi": sertifika_sayisi,
                    "son_hareketler": son_hareketler}

    def platform_canli_istatistik(self):
        with self.get_connection() as conn:
            c = conn.cursor()
            bugun = datetime.now().strftime("%Y-%m-%d")
            return {
                "bugun_kayit":  c.execute("SELECT COUNT(*) FROM kayitlar WHERE DATE(kayit_tarihi)=?", (bugun,)).fetchone()[0],
                "bugun_tamam":  c.execute("SELECT COUNT(*) FROM kayitlar WHERE DATE(tamamlanma_tarihi)=?", (bugun,)).fetchone()[0],
                "bugun_xp":     c.execute("SELECT COALESCE(SUM(xp_miktari),0) FROM xp_hareketleri WHERE DATE(tarih)=?", (bugun,)).fetchone()[0],
                "aktif_mentor": c.execute("SELECT COUNT(*) FROM mentor_iliskileri WHERE durum='aktif'").fetchone()[0],
                "bekleyen_sert":c.execute("SELECT COUNT(*) FROM sertifikalar WHERE durum='beklemede'").fetchone()[0],
                "okunmamis_b":  self.okunmamis_bildirim_sayisi(),
            }

    def tablo_csv_export(self, tablo_adi, dosya_yolu):
        import csv
        with self.get_connection() as conn:
            c = conn.cursor()
            c.execute(f"SELECT * FROM {tablo_adi}")
            rows = c.fetchall()
            if not rows: return False, "Kayıt bulunamadı."
            with open(dosya_yolu, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f)
                writer.writerow([d[0] for d in c.description])
                writer.writerows(rows)
        return True, dosya_yolu

    def dogrula_admin(self,kullanici_adi,sifre):
        with self.get_connection() as conn:
            return conn.execute("SELECT id FROM sistem_kullanicilar WHERE kullanici_adi=? AND sifre_hash=?",
                                (kullanici_adi,sha256(sifre))).fetchone() is not None

class ThemeManager:
    @staticmethod
    def get_stylesheet():
        return (
            f"QMainWindow,QDialog,QWidget{{background-color:{COLORS['bg_dark']};color:{COLORS['text_main']};font-family:'Segoe UI',sans-serif;}}"
            f"QLabel{{color:{COLORS['text_main']};}}"
            f"QLineEdit,QTextEdit,QComboBox,QSpinBox,QDoubleSpinBox{{background-color:{COLORS['bg_surface']};color:{COLORS['text_main']};border:1px solid {COLORS['border']};border-radius:6px;padding:7px;font-size:12px;}}"
            f"QLineEdit:focus,QTextEdit:focus,QComboBox:focus,QSpinBox:focus{{border:2px solid {COLORS['primary']};}}"
            f"QComboBox::drop-down{{border:none;width:20px;}}"
            f"QComboBox QAbstractItemView{{background-color:{COLORS['bg_card']};color:{COLORS['text_main']};selection-background-color:{COLORS['primary']};border:1px solid {COLORS['border']};}}"
            f"QTabWidget::pane{{border:1px solid {COLORS['border']};background-color:{COLORS['bg_dark']};}}"
            f"QTabBar::tab{{background-color:{COLORS['bg_surface']};color:{COLORS['text_sec']};padding:10px 22px;border:none;font-size:12px;}}"
            f"QTabBar::tab:selected{{background-color:{COLORS['primary']};color:white;font-weight:bold;}}"
            f"QTabBar::tab:hover:!selected{{background-color:{COLORS['bg_card']};color:{COLORS['text_main']};}}"
            f"QGroupBox{{background-color:{COLORS['bg_card']};border:1px solid {COLORS['border']};border-radius:10px;margin-top:12px;padding-top:10px;font-weight:bold;color:{COLORS['text_main']};}}"
            f"QGroupBox::title{{subcontrol-origin:margin;left:12px;padding:0 6px;color:{COLORS['secondary']};}}"
            f"QCheckBox{{color:{COLORS['text_main']};spacing:8px;}}"
            f"QCheckBox::indicator{{width:16px;height:16px;border:2px solid {COLORS['border']};border-radius:4px;background:{COLORS['bg_surface']};}}"
            f"QCheckBox::indicator:checked{{background-color:{COLORS['primary']};border-color:{COLORS['primary']};}}"
            f"QProgressBar{{background-color:{COLORS['bg_surface']};border:none;border-radius:6px;height:12px;color:white;font-size:10px;}}"
            f"QProgressBar::chunk{{background-color:{COLORS['primary']};border-radius:6px;}}"
            f"QScrollArea{{border:none;background:transparent;}}"
            f"QSplitter::handle{{background-color:{COLORS['border']};width:2px;}}"
            f"QListWidget{{background-color:{COLORS['bg_card']};color:{COLORS['text_main']};border:none;font-size:12px;}}"
            f"QListWidget::item{{padding:6px;border-bottom:1px solid {COLORS['border']};}}"
            f"QListWidget::item:selected{{background-color:{COLORS['primary']};color:white;}}"
        )

class AdminLuxBtn(QPushButton):
    def __init__(self,text,color=None,parent=None):
        super().__init__(text,parent); c=color or COLORS["primary"]
        self.setStyleSheet(f"QPushButton{{background-color:{c};color:white;border:none;border-radius:8px;padding:9px 16px;font-weight:bold;font-size:12px;}}"
                           f"QPushButton:hover{{background-color:#005a91;}}QPushButton:pressed{{background-color:#003f5c;}}"
                           f"QPushButton:disabled{{background-color:{COLORS['border']};color:{COLORS['text_sec']};}}")

class KpiKart(QFrame):
    def __init__(self,ikon,baslik,deger,renk,parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"""QFrame{{
            background:qlineargradient(x1:0,y1:0,x2:1,y2:1,
                stop:0 {COLORS['bg_card']},stop:1 {COLORS['bg_surface']});
            border-radius:18px;
            border-left:5px solid {renk};
            border-top:1px solid rgba(255,255,255,0.06);
            border-right:1px solid rgba(255,255,255,0.03);
            border-bottom:1px solid rgba(0,0,0,0.2);
        }}""")
        self.setMinimumSize(200,140)
        lay=QVBoxLayout(self); lay.setContentsMargins(22,20,22,20); lay.setSpacing(10)
        # Üst: başlık + ikon
        top=QHBoxLayout()
        lb=QLabel(baslik)
        lb.setStyleSheet(f"color:{COLORS['text_sec']};font-size:13px;font-weight:600;background:transparent;border:none;letter-spacing:0.5px;")
        top.addWidget(lb); top.addStretch()
        li=QLabel(ikon)
        li.setStyleSheet(f"font-size:22px;background:rgba(255,255,255,0.07);border-radius:10px;padding:5px 9px;border:none;")
        top.addWidget(li); lay.addLayout(top)
        lay.addSpacing(6)
        # Değer — büyük ve belirgin
        ld=QLabel(str(deger))
        f=QFont("Segoe UI",32,QFont.Bold); ld.setFont(f)
        ld.setStyleSheet(f"color:{renk};background:transparent;border:none;letter-spacing:-1px;")
        lay.addWidget(ld)

class XpProgressBar(QFrame):
    def __init__(self,xp,parent=None):
        super().__init__(parent)
        self.setStyleSheet(f"QFrame{{background-color:{COLORS['bg_card']};border-radius:12px;border:1px solid {COLORS['border']};}}")
        lay=QVBoxLayout(self); lay.setContentsMargins(14,10,14,10); lay.setSpacing(6)
        lv=level_hesapla(xp); prog=level_progress(xp)
        row=QHBoxLayout()
        lbl_lv=QLabel(f"⚡ Level {lv}"); lbl_lv.setStyleSheet(f"color:{COLORS['level_color']};font-weight:bold;font-size:14px;background:transparent;border:none;"); row.addWidget(lbl_lv); row.addStretch()
        lbl_xp=QLabel(f"{xp} XP"); lbl_xp.setStyleSheet(f"color:{COLORS['gold']};font-weight:bold;font-size:13px;background:transparent;border:none;"); row.addWidget(lbl_xp); lay.addLayout(row)
        pb=QProgressBar(); pb.setRange(0,100); pb.setValue(prog)
        pb.setStyleSheet(f"QProgressBar{{background:{COLORS['bg_surface']};border:none;border-radius:6px;height:14px;color:transparent;}}"
                         f"QProgressBar::chunk{{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 {COLORS['xp_color']},stop:1 {COLORS['level_color']});border-radius:6px;}}")
        pb.setFormat(f"%{prog}  —  Sonraki Level: {sonraki_level_xp(xp)} XP"); lay.addWidget(pb)

class BadgeKart(QFrame):
    def __init__(self,badge_kodu,kazanildi=True,tarih=None,parent=None):
        super().__init__(parent)
        meta=BADGES.get(badge_kodu,{}); renk=meta.get("renk",COLORS["border"]) if kazanildi else COLORS["border"]
        self.setFixedSize(130,100)
        self.setStyleSheet(f"QFrame{{background-color:{COLORS['bg_card']};border:{'2px' if kazanildi else '1px'} solid {renk};border-radius:10px;}}")
        lay=QVBoxLayout(self); lay.setContentsMargins(8,8,8,8); lay.setSpacing(4); lay.setAlignment(Qt.AlignCenter)
        li=QLabel(meta.get("icon","❓") if kazanildi else "🔒"); li.setAlignment(Qt.AlignCenter)
        li.setStyleSheet("font-size:26px;background:transparent;border:none;"); lay.addWidget(li)
        la=QLabel(meta.get("ad","?") if kazanildi else "???"); la.setAlignment(Qt.AlignCenter); la.setWordWrap(True)
        la.setStyleSheet(f"font-size:10px;font-weight:bold;color:{'white' if kazanildi else COLORS['text_sec']};background:transparent;border:none;"); lay.addWidget(la)
        if kazanildi and tarih:
            lt=QLabel(str(tarih)[:10]); lt.setAlignment(Qt.AlignCenter)
            lt.setStyleSheet(f"font-size:9px;color:{COLORS['text_sec']};background:transparent;border:none;"); lay.addWidget(lt)
        self.setToolTip(meta.get("aciklama","") if kazanildi else "Henüz kazanılmadı")

class LoginDialog(QDialog):
    def __init__(self,db):
        super().__init__(); self.db=db; self.setWindowTitle("🎓 Giriş")
        self.setFixedSize(420,340); self.setStyleSheet(ThemeManager.get_stylesheet()); self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(36,36,36,36); lay.setSpacing(14)
        for txt,sz,bold,style in [("🎓",40,False,"font-size:40px;"),
            ("Online Kurs Platformu",16,True,None),("Admin Paneline Giriş",12,False,f"color:{COLORS['text_sec']};font-size:12px;")]:
            l=QLabel(txt); l.setAlignment(Qt.AlignCenter)
            if sz>30: l.setStyleSheet(style)
            elif bold: f=QFont(); f.setPointSize(sz); f.setBold(True); l.setFont(f)
            elif style: l.setStyleSheet(style)
            lay.addWidget(l)
        lay.addSpacing(8)
        self.inp_user=QLineEdit(); self.inp_user.setPlaceholderText("Kullanıcı Adı"); self.inp_user.setText("admin"); lay.addWidget(self.inp_user)
        self.inp_pass=QLineEdit(); self.inp_pass.setPlaceholderText("Şifre"); self.inp_pass.setEchoMode(QLineEdit.Password)
        self.inp_pass.setText("admin123"); self.inp_pass.returnPressed.connect(self._giris); lay.addWidget(self.inp_pass)
        btn=AdminLuxBtn("Giriş Yap →"); btn.setMinimumHeight(42); btn.clicked.connect(self._giris); lay.addWidget(btn)
    def _giris(self):
        if self.db.dogrula_admin(self.inp_user.text().strip(),self.inp_pass.text()): self.accept()
        else: msg_error(self,"Hata","Kullanıcı adı veya şifre hatalı!")

class XpVerDialog(QDialog):
    def __init__(self,db,ogrenci,parent=None):
        super().__init__(parent); self.db=db; self.ogrenci=ogrenci
        self.setWindowTitle(f"XP Ver — {ogrenci['ad']}"); self.setFixedSize(420,360)
        self.setStyleSheet(ThemeManager.get_stylesheet()); self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(24,24,24,24); lay.setSpacing(14)
        b=QLabel(f"⚡ XP Ver: {self.ogrenci['ad']} {self.ogrenci['soyad'] or ''}")
        f=QFont(); f.setPointSize(13); f.setBold(True); b.setFont(f); lay.addWidget(b)
        xp,_=self.db.ogrenci_xp_getir(self.ogrenci["id"]); lay.addWidget(XpProgressBar(xp))
        form=QFormLayout(); form.setSpacing(10)
        self.cmb=QComboBox()
        for kod,m in XP_TABLOSU.items():
            if m>0: self.cmb.addItem(f"{kod.replace('_',' ').title()} (+{m} XP)",userData=kod)
        form.addRow("İşlem Türü:",self.cmb)
        self.inp_ac=QLineEdit(); self.inp_ac.setPlaceholderText("Açıklama (opsiyonel)")
        form.addRow("Açıklama:",self.inp_ac); lay.addLayout(form)
        self.lbl_on=QLabel(); self.lbl_on.setStyleSheet(f"color:{COLORS['gold']};font-weight:bold;font-size:13px;")
        self.cmb.currentIndexChanged.connect(self._onizle); lay.addWidget(self.lbl_on); self._onizle()
        row=QHBoxLayout()
        bi=AdminLuxBtn("İptal",COLORS["danger"]); bi.clicked.connect(self.reject); row.addWidget(bi); row.addStretch()
        bv=AdminLuxBtn("⚡ XP Ver",COLORS["xp_color"]); bv.clicked.connect(self._ver); row.addWidget(bv); lay.addLayout(row)
    def _onizle(self): self.lbl_on.setText(f"+ {XP_TABLOSU.get(self.cmb.currentData(),0)} XP kazanacak")
    def _ver(self):
        kod=self.cmb.currentData(); ac=self.inp_ac.text().strip() or kod.replace("_"," ").title()
        miktar,yeni=self.db.xp_ekle(self.ogrenci["id"],kod,ac)
        sonuc=f"{miktar} XP verildi!"
        if yeni: sonuc+=f"\n\n🏅 Yeni Badge:\n"+"\n".join(BADGES[b]["ad"] for b in yeni)
        msg_info(self,"Başarılı",sonuc); self.accept()

class LeaderboardDialog(QDialog):
    def __init__(self,db,parent=None):
        super().__init__(parent); self.db=db; self.setWindowTitle("🏆 Leaderboard")
        self.setMinimumSize(700,540); self.setStyleSheet(ThemeManager.get_stylesheet()); self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(24,24,24,24); lay.setSpacing(12)
        b=QLabel("🏆 XP Leaderboard"); f=QFont(); f.setPointSize(16); f.setBold(True); b.setFont(f); b.setAlignment(Qt.AlignCenter); lay.addWidget(b)
        medals={1:COLORS["gold"],2:COLORS["silver"],3:COLORS["bronze"]}
        for sira,ogr in enumerate(self.db.leaderboard_getir(10),1):
            kart=QFrame(); br=medals.get(sira,COLORS["border"])
            kart.setStyleSheet(f"QFrame{{background-color:{COLORS['bg_card']};border:{'3px' if sira<=3 else '1px'} solid {br};border-radius:10px;}}")
            kl=QHBoxLayout(kart); kl.setContentsMargins(16,10,16,10)
            icon={1:"🥇",2:"🥈",3:"🥉"}.get(sira,str(sira))
            ls=QLabel(icon); ls.setStyleSheet(f"font-size:{'22px' if sira<=3 else '14px'};background:transparent;border:none;min-width:40px;"); ls.setAlignment(Qt.AlignCenter); kl.addWidget(ls)
            v=QVBoxLayout(); v.setSpacing(2); meta=TIER_META.get(ogr["paket"],{})
            la=QLabel(f"{ogr['ad']} {ogr['soyad'] or ''}"); la.setStyleSheet(f"font-weight:bold;font-size:13px;background:transparent;border:none;color:{'white' if sira<=3 else COLORS['text_main']};"); v.addWidget(la)
            lp=QLabel(f"{meta.get('icon','')} {ogr['paket'].upper()} | {ogr['tamamlanan']} kurs tamamlandı"); lp.setStyleSheet(f"font-size:11px;color:{meta.get('renk',COLORS['text_sec'])};background:transparent;border:none;"); v.addWidget(lp)
            kl.addLayout(v); kl.addStretch()
            lb2=QLabel(f"🏅 {ogr['badge_sayisi']}"); lb2.setStyleSheet(f"font-size:12px;color:{COLORS['gold']};background:transparent;border:none;min-width:50px;"); lb2.setAlignment(Qt.AlignCenter); kl.addWidget(lb2)
            ll=QLabel(f"Lv.{ogr['level']}"); ll.setStyleSheet(f"font-size:12px;color:{COLORS['level_color']};font-weight:bold;background:transparent;border:none;min-width:45px;"); ll.setAlignment(Qt.AlignCenter); kl.addWidget(ll)
            lx=QLabel(f"{ogr['xp']} XP"); lx.setStyleSheet(f"font-size:14px;font-weight:bold;color:{br};background:transparent;border:none;min-width:80px;"); lx.setAlignment(Qt.AlignRight|Qt.AlignVCenter); kl.addWidget(lx)
            lay.addWidget(kart)
        lay.addStretch()
        bk=AdminLuxBtn("Kapat",COLORS["primary"]); bk.clicked.connect(self.accept); lay.addWidget(bk,alignment=Qt.AlignRight)

class KursDialog(QDialog):
    def __init__(self,db,kurs=None,parent=None):
        super().__init__(parent); self.db=db; self.kurs=kurs
        self.setWindowTitle("Kurs Ekle" if not kurs else "Kurs Düzenle")
        self.setMinimumSize(520,560); self.setStyleSheet(ThemeManager.get_stylesheet()); self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(24,24,24,24); lay.setSpacing(14)
        b=QLabel("📚 "+("Yeni Kurs" if not self.kurs else "Kursu Düzenle")); f=QFont(); f.setPointSize(13); f.setBold(True); b.setFont(f); lay.addWidget(b)
        frm=QFormLayout(); frm.setSpacing(10)
        self.i_kod=QLineEdit(); self.i_ad=QLineEdit(); self.i_ac=QTextEdit(); self.i_ac.setMaximumHeight(70)
        self.c_e=QComboBox()
        for e in self.db.tum_egitmenler_getir(): self.c_e.addItem(f"{e['ad']} ({e['uzmanlik']})",userData=e["id"])
        self.c_t=QComboBox()
        for k,m in TIER_META.items(): self.c_t.addItem(f"{m['icon']} {m['label']} — ₺{m['fiyat']}",userData=k)
        self.c_t.currentIndexChanged.connect(lambda: self.i_f.setValue(TIER_META.get(self.c_t.currentData(),{}).get("fiyat",0)))
        self.i_f=QDoubleSpinBox(); self.i_f.setRange(0,9999); self.i_f.setSuffix(" ₺")
        self.i_k=QSpinBox(); self.i_k.setRange(1,500); self.i_k.setValue(30)
        self.i_s=QSpinBox(); self.i_s.setRange(1,500); self.i_s.setValue(20); self.i_s.setSuffix(" saat")
        self.c_kat=QComboBox()
        for kat in ["Programlama","Web","Tasarım","Veri","İş Dünyası","Dil","Genel"]: self.c_kat.addItem(kat)
        frm.addRow("Kurs Kodu *:",self.i_kod); frm.addRow("Kurs Adı *:",self.i_ad)
        frm.addRow("Açıklama:",self.i_ac); frm.addRow("Eğitmen *:",self.c_e)
        frm.addRow("Tier *:",self.c_t); frm.addRow("Fiyat:",self.i_f)
        frm.addRow("Kontenjan:",self.i_k); frm.addRow("Süre:",self.i_s)
        frm.addRow("Kategori:",self.c_kat); lay.addLayout(frm)
        if self.kurs:
            self.i_kod.setText(self.kurs["kurs_kodu"]); self.i_kod.setEnabled(False)
            self.i_ad.setText(self.kurs["ad"]); self.i_ac.setText(self.kurs["aciklama"] or "")
            for combo,key in [(self.c_e,"egitmen_id"),(self.c_t,"tier")]:
                i=combo.findData(self.kurs[key]);
                if i>=0: combo.setCurrentIndex(i)
            self.i_f.setValue(self.kurs["fiyat"] or 0); self.i_k.setValue(self.kurs["kontenjan"] or 30)
            self.i_s.setValue(self.kurs["sure_saat"] or 20)
            i=self.c_kat.findText(self.kurs["kategori"] or "Genel");
            if i>=0: self.c_kat.setCurrentIndex(i)
        row=QHBoxLayout(); bi=AdminLuxBtn("İptal",COLORS["danger"]); bi.clicked.connect(self.reject); row.addWidget(bi); row.addStretch()
        bk=AdminLuxBtn("💾 Kaydet",COLORS["success"]); bk.clicked.connect(self._kaydet); row.addWidget(bk); lay.addLayout(row)
    def _kaydet(self):
        kod=self.i_kod.text().strip(); ad=self.i_ad.text().strip()
        if not kod or not ad: return msg_warn(self,"Eksik","Kurs kodu ve adı zorunludur!")
        args=(ad,self.i_ac.toPlainText().strip(),self.c_e.currentData(),self.c_t.currentData(),self.i_f.value(),self.i_k.value(),self.i_s.value(),self.c_kat.currentText())
        ok,msg=(self.db.kurs_guncelle(self.kurs["id"],*args) if self.kurs else self.db.kurs_ekle(kod,*args))
        if ok: msg_info(self,"Başarılı",msg); self.accept()
        else: msg_error(self,"Hata",msg)

class EgitmenDialog(QDialog):
    def __init__(self,db,egitmen=None,parent=None):
        super().__init__(parent); self.db=db; self.egitmen=egitmen
        self.setWindowTitle("Eğitmen Ekle" if not egitmen else "Eğitmen Düzenle")
        self.setMinimumSize(460,360); self.setStyleSheet(ThemeManager.get_stylesheet()); self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(24,24,24,24); lay.setSpacing(14)
        b=QLabel("👨‍🏫 "+("Yeni Eğitmen" if not self.egitmen else "Eğitmeni Düzenle")); f=QFont(); f.setPointSize(13); f.setBold(True); b.setFont(f); lay.addWidget(b)
        frm=QFormLayout(); frm.setSpacing(10)
        self.i_ad=QLineEdit(); self.i_uz=QLineEdit(); self.i_em=QLineEdit(); self.i_bio=QTextEdit(); self.i_bio.setMaximumHeight(80)
        frm.addRow("Ad Soyad *:",self.i_ad); frm.addRow("Uzmanlık *:",self.i_uz)
        frm.addRow("E-posta:",self.i_em); frm.addRow("Biyografi:",self.i_bio); lay.addLayout(frm)
        if self.egitmen:
            self.i_ad.setText(self.egitmen["ad"]); self.i_uz.setText(self.egitmen["uzmanlik"] or "")
            self.i_em.setText(self.egitmen["email"] or ""); self.i_bio.setText(self.egitmen["biyografi"] or "")
        row=QHBoxLayout(); bi=AdminLuxBtn("İptal",COLORS["danger"]); bi.clicked.connect(self.reject); row.addWidget(bi); row.addStretch()
        bk=AdminLuxBtn("💾 Kaydet",COLORS["success"]); bk.clicked.connect(self._kaydet); row.addWidget(bk); lay.addLayout(row)
    def _kaydet(self):
        ad=self.i_ad.text().strip(); uz=self.i_uz.text().strip()
        if not ad or not uz: return msg_warn(self,"Eksik","Ad ve uzmanlık zorunludur!")
        ok,msg=(self.db.egitmen_guncelle(self.egitmen["id"],ad,uz,self.i_em.text().strip(),self.i_bio.toPlainText().strip()) if self.egitmen
                else self.db.egitmen_ekle(ad,uz,self.i_em.text().strip(),self.i_bio.toPlainText().strip()))
        if ok: msg_info(self,"Başarılı",msg); self.accept()
        else: msg_error(self,"Hata",msg)

class OgrenciDialog(QDialog):
    def __init__(self,db,ogrenci=None,parent=None):
        super().__init__(parent); self.db=db; self.ogrenci=ogrenci
        self.setWindowTitle("Öğrenci Ekle" if not ogrenci else "Öğrenci Düzenle")
        self.setMinimumSize(460,400); self.setStyleSheet(ThemeManager.get_stylesheet()); self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(24,24,24,24); lay.setSpacing(14)
        b=QLabel("👥 "+("Yeni Öğrenci" if not self.ogrenci else "Öğrenciyi Düzenle")); f=QFont(); f.setPointSize(13); f.setBold(True); b.setFont(f); lay.addWidget(b)
        frm=QFormLayout(); frm.setSpacing(10)
        self.i_k=QLineEdit(); self.i_s=QLineEdit(); self.i_s.setPlaceholderText("(boş bırakılırsa değişmez)")
        self.i_ad=QLineEdit(); self.i_soy=QLineEdit(); self.i_em=QLineEdit()
        self.c_p=QComboBox()
        for k,m in TIER_META.items(): self.c_p.addItem(f"{m['icon']} {m['label']}",userData=k)
        frm.addRow("Kullanıcı Adı *:",self.i_k); frm.addRow("Şifre:",self.i_s)
        frm.addRow("Ad *:",self.i_ad); frm.addRow("Soyad:",self.i_soy)
        frm.addRow("E-posta:",self.i_em); frm.addRow("Paket:",self.c_p); lay.addLayout(frm)
        if self.ogrenci:
            self.i_k.setText(self.ogrenci["kullanici_adi"]); self.i_k.setEnabled(False)
            self.i_ad.setText(self.ogrenci["ad"] or ""); self.i_soy.setText(self.ogrenci["soyad"] or ""); self.i_em.setText(self.ogrenci["email"] or "")
            i=self.c_p.findData(self.ogrenci["paket"]);
            if i>=0: self.c_p.setCurrentIndex(i)
        row=QHBoxLayout(); bi=AdminLuxBtn("İptal",COLORS["danger"]); bi.clicked.connect(self.reject); row.addWidget(bi); row.addStretch()
        bk=AdminLuxBtn("💾 Kaydet",COLORS["success"]); bk.clicked.connect(self._kaydet); row.addWidget(bk); lay.addLayout(row)
    def _kaydet(self):
        k=self.i_k.text().strip(); ad=self.i_ad.text().strip()
        if not k or not ad: return msg_warn(self,"Eksik","Kullanıcı adı ve ad zorunludur!")
        p=self.c_p.currentData()
        if self.ogrenci: ok,msg=self.db.ogrenci_guncelle(self.ogrenci["id"],ad,self.i_soy.text().strip(),self.i_em.text().strip(),p)
        else:
            s=self.i_s.text()
            if not s: return msg_warn(self,"Eksik","Şifre gerekli!")
            ok,msg=self.db.ogrenci_ekle(k,s,ad,self.i_soy.text().strip(),self.i_em.text().strip(),p)
        if ok: msg_info(self,"Başarılı",msg); self.accept()
        else: msg_error(self,"Hata",msg)

class EnhancementShopDialog(QDialog):
    def __init__(self,db,ogrenci,kurs,parent=None):
        super().__init__(parent); self.db=db; self.ogrenci=ogrenci; self.kurs=kurs
        self.setWindowTitle(f"Enhancement Shop — {kurs['ad']}"); self.setMinimumSize(560,500)
        self.setStyleSheet(ThemeManager.get_stylesheet()); self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(24,24,24,24); lay.setSpacing(16)
        b=QLabel("🛒 Enhancement Shop"); f=QFont(); f.setPointSize(14); f.setBold(True); b.setFont(f); lay.addWidget(b)
        lay.addWidget(QLabel(f"Öğrenci: {self.ogrenci['ad']}  |  Kurs: {self.kurs['ad']}"))
        aktif=self.db.ogrenci_enhancements_getir(self.ogrenci["id"],self.kurs["id"])
        scroll=QScrollArea(); scroll.setWidgetResizable(True); ctn=QWidget(); cl=QVBoxLayout(ctn); cl.setSpacing(10)
        self.cb={}
        for kod,meta in ENHANCEMENTS.items():
            kart=QFrame(); kart.setStyleSheet(f"QFrame{{background:{COLORS['bg_card']};border:1px solid {'#059669' if kod in aktif else COLORS['border']};border-radius:10px;}}")
            kl=QHBoxLayout(kart); kl.setContentsMargins(14,12,14,12)
            li=QLabel(meta["icon"]); li.setStyleSheet("font-size:22px;background:transparent;border:none;"); kl.addWidget(li)
            v=QVBoxLayout()
            la=QLabel(meta["ad"]); la.setStyleSheet("font-weight:bold;font-size:13px;background:transparent;border:none;"); v.addWidget(la)
            tr=TIER_META.get(meta["tier"],{}).get("renk",COLORS["primary"])
            lt=QLabel(f"Tier: {meta['tier'].upper()}"); lt.setStyleSheet(f"color:{tr};font-size:11px;background:transparent;border:none;"); v.addWidget(lt)
            kl.addLayout(v); kl.addStretch()
            lf=QLabel(f"₺{meta['fiyat']}"); lf.setStyleSheet(f"color:{COLORS['gold']};font-weight:bold;font-size:14px;background:transparent;border:none;"); kl.addWidget(lf)
            if kod in aktif:
                la2=QLabel("✅ Aktif"); la2.setStyleSheet(f"color:{COLORS['success']};font-weight:bold;background:transparent;border:none;"); kl.addWidget(la2)
            else:
                c=QCheckBox(); self.cb[kod]=c; kl.addWidget(c)
            cl.addWidget(kart)
        cl.addStretch(); scroll.setWidget(ctn); lay.addWidget(scroll)
        self.lbl_top=QLabel("Toplam: ₺0"); self.lbl_top.setStyleSheet(f"font-size:14px;font-weight:bold;color:{COLORS['gold']};"); lay.addWidget(self.lbl_top)
        for c in self.cb.values(): c.stateChanged.connect(self._top)
        row=QHBoxLayout(); bi=AdminLuxBtn("İptal",COLORS["danger"]); bi.clicked.connect(self.reject); row.addWidget(bi); row.addStretch()
        bs=AdminLuxBtn("🛒 Satın Al",COLORS["success"]); bs.clicked.connect(self._satin); row.addWidget(bs); lay.addLayout(row)
    def _top(self): self.lbl_top.setText(f"Toplam: ₺{sum(ENHANCEMENTS[k]['fiyat'] for k,c in self.cb.items() if c.isChecked())}")
    def _satin(self):
        sec=[k for k,c in self.cb.items() if c.isChecked()]
        if not sec: return msg_warn(self,"Uyarı","Hiç enhancement seçilmedi.")
        t=sum(ENHANCEMENTS[k]["fiyat"] for k in sec)
        if not msg_question(self,"Onay",f"{len(sec)} enhancement, toplam ₺{t}\nOnaylıyor musunuz?"): return
        hatalar=[]
        for kod in sec:
            ok,msg=self.db.enhancement_satin_al(self.ogrenci["id"],self.kurs["id"],kod)
            if not ok: hatalar.append(msg)
            else: self.db.xp_ekle(self.ogrenci["id"],"enhancement_al",f"Enhancement: {ENHANCEMENTS[kod]['ad']}")
        if hatalar: msg_warn(self,"Kısmi Başarı","\n".join(hatalar))
        else: msg_info(self,"Başarılı",f"{len(sec)} enhancement aktive edildi!")
        self.accept()


class KurslarSekmesi(QWidget):
    def __init__(self,db): super().__init__(); self.db=db; self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(16,16,16,16); lay.setSpacing(10)
        tb=QHBoxLayout(); l=QLabel("📚 Kurs Yönetimi"); f=QFont(); f.setPointSize(13); f.setBold(True); l.setFont(f); tb.addWidget(l); tb.addStretch()
        self.cf=QComboBox(); self.cf.addItem("Tüm Tier'lar",userData="hepsi")
        for k,m in TIER_META.items(): self.cf.addItem(f"{m['icon']} {m['label']}",userData=k)
        self.cf.currentIndexChanged.connect(self._yenile); tb.addWidget(self.cf)
        be=AdminLuxBtn("➕ Kurs Ekle",COLORS["success"]); be.clicked.connect(self._ekle); tb.addWidget(be); lay.addLayout(tb)
        self.tablo=QTableWidget(); self.tablo.setColumnCount(8)
        self.tablo.setHorizontalHeaderLabels(["ID","Kod","Kurs Adı","Eğitmen","Tier","Fiyat","Kontenjan","Süre"])
        self.tablo.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tablo.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tablo.setAlternatingRowColors(True)
        self.tablo.horizontalHeader().setSectionResizeMode(2,QHeaderView.Stretch)
        self.tablo.setStyleSheet(tablo_stili()); self.tablo.verticalHeader().setVisible(False)
        self.tablo.doubleClicked.connect(self._duzenle); lay.addWidget(self.tablo)
        alt=QHBoxLayout()
        for t,r,fn in [("📖 Detay",COLORS["accent"],self._detay),("✏️ Düzenle",COLORS["warning"],self._duzenle),("🗑 Sil",COLORS["danger"],self._sil)]:
            b=AdminLuxBtn(t,r); b.clicked.connect(fn); alt.addWidget(b)
        alt.addStretch()
        br=AdminLuxBtn("🔄 Yenile",COLORS["primary"]); br.clicked.connect(self._yenile); alt.addWidget(br)
        lay.addLayout(alt); self._yenile()
    def _yenile(self):
        tier=self.cf.currentData() if self.cf.count() else "hepsi"
        ks=self.db.tum_kurslar_getir(tier); self.tablo.setRowCount(len(ks))
        for r,k in enumerate(ks):
            meta=TIER_META.get(k["tier"],{}); renk=QColor(meta.get("renk",COLORS["primary"]))
            for col,val in enumerate([str(k["id"]),k["kurs_kodu"],k["ad"],k["egitmen_adi"] or "—",
                                       f"{meta.get('icon','')} {k['tier'].upper()}",
                                       f"₺{k['fiyat']:.0f}",str(k["kontenjan"]),f"{k['sure_saat']}s"]):
                item=QTableWidgetItem(val); item.setTextAlignment(Qt.AlignCenter)
                if col==4: item.setForeground(renk)
                self.tablo.setItem(r,col,item)
        self.tablo.resizeColumnsToContents(); self.tablo.setColumnWidth(2,240)
    def _sid(self):
        r=self.tablo.currentRow()
        if r<0: msg_warn(self,"Uyarı","Lütfen bir kurs seçin!"); return None
        return int(self.tablo.item(r,0).text())
    def _ekle(self):
        dlg=KursDialog(self.db,parent=self)
        if dlg.exec_()==QDialog.Accepted: self._yenile()
    def _detay(self):
        kid=self._sid()
        if not kid: return
        AdminKursDetay(self.db,kid,self).exec_()
    def _duzenle(self):
        kid=self._sid()
        if not kid: return
        k=next((x for x in self.db.tum_kurslar_getir() if x["id"]==kid),None)
        if k:
            dlg=KursDialog(self.db,kurs=k,parent=self)
            if dlg.exec_()==QDialog.Accepted: self._yenile()
    def _sil(self):
        kid=self._sid()
        if not kid: return
        if msg_question(self,"Sil","Bu kurs silinecek. Onaylıyor musunuz?"):
            ok,msg=self.db.kurs_sil(kid)
            msg_info(self,"Bilgi",msg) if ok else msg_error(self,"Hata",msg)
            self._yenile()


class OgrencilerSekmesi(QWidget):
    def __init__(self,db): super().__init__(); self.db=db; self.sec=None; self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(16,16,16,16); lay.setSpacing(10)
        tb=QHBoxLayout(); l=QLabel("👥 Öğrenci Yönetimi"); f=QFont(); f.setPointSize(13); f.setBold(True); l.setFont(f); tb.addWidget(l); tb.addStretch()
        self.ara=QLineEdit(); self.ara.setPlaceholderText("🔍 Öğrenci ara..."); self.ara.setMaximumWidth(220)
        self.ara.textChanged.connect(self._filtrele); tb.addWidget(self.ara)
        be=AdminLuxBtn("➕ Öğrenci Ekle",COLORS["success"]); be.clicked.connect(self._ekle); tb.addWidget(be); lay.addLayout(tb)
        spl=QSplitter(Qt.Horizontal)
        # Sol: tablo
        sol=QWidget(); sl=QVBoxLayout(sol); sl.setContentsMargins(0,0,0,0)
        self.tablo=QTableWidget(); self.tablo.setColumnCount(6)
        self.tablo.setHorizontalHeaderLabels(["ID","Ad Soyad","Kullanıcı","Paket","XP","Kurs"])
        self.tablo.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tablo.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tablo.setAlternatingRowColors(True); self.tablo.setStyleSheet(tablo_stili())
        self.tablo.verticalHeader().setVisible(False)
        self.tablo.horizontalHeader().setSectionResizeMode(1,QHeaderView.Stretch)
        self.tablo.clicked.connect(self._sec_ogr); sl.addWidget(self.tablo)
        alt=QHBoxLayout()
        for t,r,fn in [("👤 Profil",COLORS["accent"],self._profil),("✏️ Düzenle",COLORS["warning"],self._duzenle),("🗑 Pasife Al",COLORS["danger"],self._sil)]:
            b=AdminLuxBtn(t,r); b.clicked.connect(fn); alt.addWidget(b)
        alt.addStretch(); sl.addLayout(alt); spl.addWidget(sol)
        # Sağ: detay
        sag=QWidget(); sg=QVBoxLayout(sag); sg.setContentsMargins(10,0,0,0); sg.setSpacing(10)
        sg.addWidget(QLabel("Öğrenci Detayı"))
        grp=QGroupBox("Kayıtlı Kurslar"); gl=QVBoxLayout(grp)
        self.tk=QTableWidget(); self.tk.setColumnCount(4)
        self.tk.setHorizontalHeaderLabels(["Kurs","Tier","İlerleme","Durum"])
        self.tk.setEditTriggers(QAbstractItemView.NoEditTriggers); self.tk.setStyleSheet(tablo_stili())
        self.tk.verticalHeader().setVisible(False)
        self.tk.horizontalHeader().setSectionResizeMode(0,QHeaderView.Stretch)
        gl.addWidget(self.tk)
        bs=AdminLuxBtn("🛒 Enhancement Shop",COLORS["gold"]); bs.clicked.connect(self._shop); gl.addWidget(bs)
        kr=QHBoxLayout(); self.cmb_kr=QComboBox()
        for k in self.db.tum_kurslar_getir(): self.cmb_kr.addItem(f"{k['ad']} ({k['tier'].upper()})",userData=k["id"])
        kr.addWidget(self.cmb_kr)
        bkr=AdminLuxBtn("Kursa Kaydet",COLORS["primary"]); bkr.clicked.connect(self._kursa); kr.addWidget(bkr)
        gl.addLayout(kr); sg.addWidget(grp); sg.addStretch()
        spl.addWidget(sag); spl.setSizes([520,380]); lay.addWidget(spl); self._yenile()
    def _yenile(self): self._tum=self.db.tum_ogrenciler_getir(); self._doldur(self._tum)
    def _doldur(self,ogr):
        self.tablo.setRowCount(len(ogr))
        for r,o in enumerate(ogr):
            meta=TIER_META.get(o["paket"],{})
            for col,val in enumerate([str(o["id"]),f"{o['ad']} {o['soyad'] or ''}",o["kullanici_adi"],
                                       f"{meta.get('icon','')} {o['paket'].upper()}",str(o["xp"]),str(o["kurs_sayisi"])]):
                item=QTableWidgetItem(val); item.setTextAlignment(Qt.AlignCenter)
                if col==3: item.setForeground(QColor(meta.get("renk",COLORS["primary"])))
                if col==4: item.setForeground(QColor(COLORS["gold"]))
                self.tablo.setItem(r,col,item)
        self.tablo.resizeColumnsToContents(); self.tablo.setColumnWidth(1,180)
    def _filtrele(self,t):
        if not t: self._doldur(self._tum); return
        tl=t.lower()
        self._doldur([o for o in self._tum if tl in (o["ad"] or "").lower() or tl in (o["kullanici_adi"] or "").lower()])
    def _sec_ogr(self):
        row=self.tablo.currentRow()
        if row<0: return
        oid=int(self.tablo.item(row,0).text())
        self.sec=next((o for o in self._tum if o["id"]==oid),None)
        if not self.sec: return
        ks=self.db.ogrenci_kayitlar_getir(oid); self.tk.setRowCount(len(ks))
        for r,k in enumerate(ks):
            meta=TIER_META.get(k["tier"],{})
            for col,val in enumerate([k["kurs_adi"],f"{meta.get('icon','')} {k['tier'].upper()}",f"%{k['ilerleme']}",k["durum"]]):
                item=QTableWidgetItem(val); item.setTextAlignment(Qt.AlignCenter); self.tk.setItem(r,col,item)
        self.tk.resizeColumnsToContents(); self.tk.setColumnWidth(0,180)
    def _oid(self):
        r=self.tablo.currentRow()
        if r<0: msg_warn(self,"Uyarı","Lütfen bir öğrenci seçin!"); return None
        return int(self.tablo.item(r,0).text())
    def _profil(self):
        oid=self._oid()
        if not oid: return
        OgrenciProfilDialog(self.db,oid,self).exec_()
    def _ekle(self):
        dlg=OgrenciDialog(self.db,parent=self)
        if dlg.exec_()==QDialog.Accepted: self._yenile()
    def _duzenle(self):
        oid=self._oid()
        if not oid: return
        o=next((x for x in self._tum if x["id"]==oid),None)
        if o:
            dlg=OgrenciDialog(self.db,ogrenci=o,parent=self)
            if dlg.exec_()==QDialog.Accepted: self._yenile()
    def _sil(self):
        oid=self._oid()
        if not oid: return
        if msg_question(self,"Pasife Al","Öğrenci pasife alınacak. Onaylıyor musunuz?"):
            ok,msg=self.db.ogrenci_sil(oid)
            msg_info(self,"Bilgi",msg) if ok else msg_error(self,"Hata",msg); self._yenile()
    def _kursa(self):
        oid=self._oid()
        if not oid: return
        kid=self.cmb_kr.currentData()
        if not kid: return msg_warn(self,"Uyarı","Kurs seçin!")
        ok,msg=self.db.kursa_kaydet(oid,kid)
        msg_info(self,"Bilgi",msg) if ok else msg_error(self,"Hata",msg); self._sec_ogr()
    def _shop(self):
        if not self.sec: return msg_warn(self,"Uyarı","Önce bir öğrenci seçin!")
        row=self.tk.currentRow()
        if row<0: return msg_warn(self,"Uyarı","Öğrenci detayında bir kurs seçin!")
        ka=self.tk.item(row,0).text()
        k=next((x for x in self.db.tum_kurslar_getir() if x["ad"]==ka),None)
        if not k: return msg_error(self,"Hata","Kurs bulunamadı.")
        EnhancementShopDialog(self.db,self.sec,k,self).exec_()


class EgitmenlerSekmesi(QWidget):
    def __init__(self,db): super().__init__(); self.db=db; self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(16,16,16,16); lay.setSpacing(10)
        tb=QHBoxLayout(); l=QLabel("👨\u200d🏫 Eğitmen Yönetimi"); f=QFont(); f.setPointSize(13); f.setBold(True); l.setFont(f); tb.addWidget(l); tb.addStretch()
        be=AdminLuxBtn("➕ Eğitmen Ekle",COLORS["success"]); be.clicked.connect(self._ekle); tb.addWidget(be); lay.addLayout(tb)
        self.tablo=QTableWidget(); self.tablo.setColumnCount(6)
        self.tablo.setHorizontalHeaderLabels(["ID","Ad","Uzmanlık","Email","Rating","Durum"])
        self.tablo.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tablo.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tablo.setAlternatingRowColors(True); self.tablo.setStyleSheet(tablo_stili())
        self.tablo.verticalHeader().setVisible(False)
        self.tablo.horizontalHeader().setSectionResizeMode(1,QHeaderView.Stretch)
        self.tablo.doubleClicked.connect(self._duzenle); lay.addWidget(self.tablo)
        alt=QHBoxLayout()
        for t,r,fn in [("✏️ Düzenle",COLORS["warning"],self._duzenle),("🗑 Pasife Al",COLORS["danger"],self._sil)]:
            b=AdminLuxBtn(t,r); b.clicked.connect(fn); alt.addWidget(b)
        alt.addStretch(); lay.addLayout(alt); self._yenile()
    def _yenile(self):
        eg=self.db.tum_egitmenler_getir(); self.tablo.setRowCount(len(eg))
        for r,e in enumerate(eg):
            for col,val in enumerate([str(e["id"]),e["ad"],e["uzmanlik"] or "—",
                                       e["email"] or "—",f"⭐ {e['rating']:.1f}",e["durum"]]):
                item=QTableWidgetItem(val); item.setTextAlignment(Qt.AlignCenter)
                if col==4: item.setForeground(QColor(COLORS["gold"]))
                self.tablo.setItem(r,col,item)
        self.tablo.resizeColumnsToContents(); self.tablo.setColumnWidth(1,200)
    def _sid(self):
        r=self.tablo.currentRow()
        if r<0: msg_warn(self,"Uyarı","Lütfen bir eğitmen seçin!"); return None
        return int(self.tablo.item(r,0).text())
    def _ekle(self):
        dlg=EgitmenDialog(self.db,parent=self)
        if dlg.exec_()==QDialog.Accepted: self._yenile()
    def _duzenle(self):
        eid=self._sid()
        if not eid: return
        e=next((x for x in self.db.tum_egitmenler_getir() if x["id"]==eid),None)
        if e:
            dlg=EgitmenDialog(self.db,egitmen=e,parent=self)
            if dlg.exec_()==QDialog.Accepted: self._yenile()
    def _sil(self):
        eid=self._sid()
        if not eid: return
        if msg_question(self,"Pasife Al","Eğitmen pasife alınacak. Onaylıyor musunuz?"):
            ok,msg=self.db.egitmen_sil(eid)
            msg_info(self,"Bilgi",msg) if ok else msg_error(self,"Hata",msg); self._yenile()


class GamificationSekmesi(QWidget):
    def __init__(self,db): super().__init__(); self.db=db; self.secili=None; self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(16,16,16,16); lay.setSpacing(12)
        tb=QHBoxLayout(); l=QLabel("🎮 Gamification Merkezi"); f=QFont(); f.setPointSize(13); f.setBold(True); l.setFont(f)
        tb.addWidget(l); tb.addStretch()
        btn_lb=AdminLuxBtn("🏆 Leaderboard",COLORS["gold"])
        btn_lb.setStyleSheet(btn_lb.styleSheet().replace("color:white",f"color:{COLORS['bg_dark']}"))
        btn_lb.clicked.connect(lambda: LeaderboardDialog(self.db,self).exec_()); tb.addWidget(btn_lb)
        lay.addLayout(tb)
        spl=QSplitter(Qt.Horizontal)
        # Sol: öğrenci listesi
        sol=QWidget(); sl=QVBoxLayout(sol); sl.setContentsMargins(0,0,0,0); sl.setSpacing(8)
        ll=QLabel("Öğrenciler"); ll.setStyleSheet(f"color:{COLORS['text_sec']};font-size:11px;font-weight:bold;"); sl.addWidget(ll)
        self.liste=QListWidget(); self.liste.itemClicked.connect(self._sec); sl.addWidget(self.liste)
        btn_xp=AdminLuxBtn("⚡ XP Ver",COLORS["xp_color"]); btn_xp.clicked.connect(self._xp_ver); sl.addWidget(btn_xp)
        spl.addWidget(sol)
        # Sağ: detay
        sag=QWidget(); sg=QVBoxLayout(sag); sg.setContentsMargins(10,0,0,0); sg.setSpacing(10)
        self.lbl_ad=QLabel("← Öğrenci seçin"); f2=QFont(); f2.setPointSize(13); f2.setBold(True); self.lbl_ad.setFont(f2); sg.addWidget(self.lbl_ad)
        self.xp_ctn=QWidget(); self.xp_lay=QVBoxLayout(self.xp_ctn); self.xp_lay.setContentsMargins(0,0,0,0); sg.addWidget(self.xp_ctn)
        self.dtabs=QTabWidget()
        # Badge tab
        self.badge_scroll=QScrollArea(); self.badge_scroll.setWidgetResizable(True)
        self.dtabs.addTab(self.badge_scroll,"🏅 Rozetler")
        # XP geçmişi
        self.xp_liste=QListWidget(); self.dtabs.addTab(self.xp_liste,"⚡ XP Geçmişi")
        # İlerleme
        self.il_ctn=QWidget(); self.il_lay=QVBoxLayout(self.il_ctn); self.il_lay.setContentsMargins(0,0,0,0); self.il_lay.setSpacing(8)
        il_scroll=QScrollArea(); il_scroll.setWidgetResizable(True); il_scroll.setWidget(self.il_ctn)
        self.dtabs.addTab(il_scroll,"📈 İlerleme")
        sg.addWidget(self.dtabs)
        spl.addWidget(sag); spl.setSizes([280,600]); lay.addWidget(spl)
        self._liste_doldur()
    def _liste_doldur(self):
        self.liste.clear(); self._cache=self.db.tum_ogrenciler_getir()
        for o in self._cache:
            meta=TIER_META.get(o["paket"],{})
            item=QListWidgetItem(f"{meta.get('icon','')} {o['ad']} {o['soyad'] or ''}  ·  Lv.{o['level']}  ·  {o['xp']} XP")
            item.setData(Qt.UserRole,o["id"]); self.liste.addItem(item)
    def _sec(self,item):
        oid=item.data(Qt.UserRole)
        self.secili=next((o for o in self._cache if o["id"]==oid),None)
        if self.secili: self._detay()
    def _detay(self):
        o=self.secili
        self.lbl_ad.setText(f"{o['ad']} {o['soyad'] or ''}  —  {o['paket'].upper()}")
        # XP bar
        while self.xp_lay.count():
            c=self.xp_lay.takeAt(0)
            if c.widget(): c.widget().deleteLater()
        xp,_=self.db.ogrenci_xp_getir(o["id"]); self.xp_lay.addWidget(XpProgressBar(xp))
        # Badge grid
        bkayit=self.db.ogrenci_badge_getir(o["id"])
        kazanilan={r["badge_kodu"]:r["kazanilma_tarihi"] for r in bkayit}
        bctn=QWidget(); grid=QGridLayout(bctn); grid.setSpacing(10); grid.setContentsMargins(10,10,10,10)
        for i,bk in enumerate(BADGES.keys()):
            grid.addWidget(BadgeKart(bk,bk in kazanilan,kazanilan.get(bk)),i//4,i%4)
        pct=int(len(kazanilan)/len(BADGES)*100)
        ozet=QLabel(f"🏅 {len(kazanilan)}/{len(BADGES)} rozet kazanıldı (%{pct})")
        ozet.setStyleSheet(f"color:{COLORS['gold']};font-weight:bold;font-size:12px;"); ozet.setAlignment(Qt.AlignCenter)
        wrap=QWidget(); wl=QVBoxLayout(wrap); wl.addWidget(ozet); wl.addWidget(bctn); wl.addStretch()
        self.badge_scroll.setWidget(wrap)
        # XP geçmişi
        self.xp_liste.clear()
        for h in self.db.xp_hareketleri_getir(o["id"]):
            item=QListWidgetItem(f"⚡ +{h['xp_miktari']} XP  ·  {h['islem_turu'].replace('_',' ').title()}  ·  {h['aciklama'] or ''}  ·  {str(h['tarih'])[:16]}")
            item.setForeground(QColor(COLORS["gold"])); self.xp_liste.addItem(item)
        # İlerleme paneli
        while self.il_lay.count():
            c=self.il_lay.takeAt(0)
            if c.widget(): c.widget().deleteLater()
        for k in self.db.ogrenci_kayitlar_getir(o["id"]):
            kart=QFrame()
            kart.setStyleSheet(f"QFrame{{background:{COLORS['bg_card']};border-radius:8px;border:1px solid {COLORS['border']};}}")
            kl=QVBoxLayout(kart); kl.setContentsMargins(12,10,12,10); kl.setSpacing(6)
            row=QHBoxLayout(); meta=TIER_META.get(k["tier"],{})
            lk=QLabel(f"{meta.get('icon','')} {k['kurs_adi']}")
            lk.setStyleSheet("font-weight:bold;font-size:12px;background:transparent;border:none;"); row.addWidget(lk); row.addStretch()
            dr=COLORS["success"] if k["durum"]=="tamamlandi" else COLORS["warning"]
            ld=QLabel("✅ Tamamlandı" if k["durum"]=="tamamlandi" else f"📖 %{k['ilerleme']}")
            ld.setStyleSheet(f"color:{dr};font-weight:bold;font-size:11px;background:transparent;border:none;"); row.addWidget(ld); kl.addLayout(row)
            pb=QProgressBar(); pb.setRange(0,100); pb.setValue(k["ilerleme"])
            pbr=COLORS["success"] if k["ilerleme"]==100 else COLORS["primary"]
            pb.setStyleSheet(f"QProgressBar{{background:{COLORS['bg_surface']};border:none;border-radius:5px;height:10px;color:transparent;}}QProgressBar::chunk{{background:{pbr};border-radius:5px;}}"); kl.addWidget(pb)
            ir=QHBoxLayout()
            spin=QSpinBox(); spin.setRange(0,100); spin.setValue(k["ilerleme"]); spin.setSuffix("%"); spin.setMaximumWidth(90)
            spin.setStyleSheet(f"background:{COLORS['bg_surface']};color:{COLORS['text_main']};border:1px solid {COLORS['border']};border-radius:4px;padding:3px;font-size:11px;")
            btn_g=AdminLuxBtn("Güncelle",COLORS["accent"]); btn_g.setMaximumWidth(90)
            btn_g.setStyleSheet(btn_g.styleSheet()+"padding:4px 8px;font-size:11px;")
            kid_ref=k["kurs_id"]; oid_ref=o["id"]
            def _guncelle(ch=False,kid=kid_ref,oid=oid_ref,s=spin,pb_ref=pb,lbl=ld):
                ok,durum=self.db.ilerleme_guncelle(oid,kid,s.value())
                if ok:
                    pb_ref.setValue(s.value())
                    if durum=="tamamlandi":
                        lbl.setText("✅ Tamamlandı"); lbl.setStyleSheet(f"color:{COLORS['success']};font-weight:bold;font-size:11px;background:transparent;border:none;")
                        miktar,yeni=self.db.xp_ekle(oid,"kurs_tamamla","Kurs tamamlandı")
                        if yeni: msg_info(self,"🏅 Yeni Badge!","Tebrikler!\n"+"\n".join(BADGES[b]["ad"] for b in yeni))
                    self._liste_doldur()
            btn_g.clicked.connect(_guncelle)
            ir.addWidget(QLabel("İlerleme:")); ir.addWidget(spin); ir.addWidget(btn_g); ir.addStretch()
            lxp=QLabel(f"XP: +{k['xp_kazanilan']}"); lxp.setStyleSheet(f"color:{COLORS['gold']};font-size:11px;background:transparent;border:none;"); ir.addWidget(lxp)
            kl.addLayout(ir); self.il_lay.addWidget(kart)
        self.il_lay.addStretch()
    def _xp_ver(self):
        if not self.secili: return msg_warn(self,"Uyarı","Önce bir öğrenci seçin!")
        dlg=XpVerDialog(self.db,self.secili,self)
        if dlg.exec_()==QDialog.Accepted:
            self._cache=self.db.tum_ogrenciler_getir()
            self.secili=next((o for o in self._cache if o["id"]==self.secili["id"]),None)
            self._liste_doldur()
            if self.secili: self._detay()


class AdminDashboard(QWidget):
    def __init__(self,db): super().__init__(); self.db=db; self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(20)
        lbl=QLabel("📊 Genel Durum")
        f=QFont("Segoe UI",15,QFont.Bold); lbl.setFont(f)
        lbl.setStyleSheet(f"color:{COLORS['text_main']};background:transparent;border:none;")
        lay.addWidget(lbl)
        stats=self.db.dashboard_istatistik()
        grid=QGridLayout(); grid.setSpacing(16)
        grid.setRowMinimumHeight(0, 170)
        grid.setRowMinimumHeight(1, 170)
        for i,(ik,b,d,r) in enumerate([
            ("👥","Toplam Öğrenci",stats["toplam_ogrenci"],COLORS["primary"]),
            ("📚","Toplam Kurs",   stats["toplam_kurs"],   COLORS["accent"]),
            ("👨\u200d🏫","Eğitmen",stats["toplam_egitmen"],COLORS["secondary"]),
            ("📝","Toplam Kayıt", stats["toplam_kayit"],  COLORS["success"]),
            ("⭐","Pro Öğrenci",  stats["pro_ogrenci"],   COLORS["gold"]),
            ("✅","Tamamlanan",   stats["tamamlanan"],    COLORS["tier_pro2"]),
            ("⚡","Toplam XP",   stats["toplam_xp"],     COLORS["xp_color"]),
            ("🏅","Toplam Badge",stats["toplam_badge"],  COLORS["warning"]),
        ]):
            kart = KpiKart(ik,b,d,r)
            kart.setMinimumHeight(170)
            grid.addWidget(kart, i//4, i%4)
        lay.addLayout(grid)
        # Top 3 leaderboard
        lb_frame=QGroupBox("🏆 XP Sıralaması — Top 3"); lb_lay=QHBoxLayout(lb_frame)
        medals={1:COLORS["gold"],2:COLORS["silver"],3:COLORS["bronze"]}
        for i,ogr in enumerate(self.db.leaderboard_getir(3),1):
            mr=medals[i]; icon={1:"🥇",2:"🥈",3:"🥉"}[i]
            kutu=QFrame(); kutu.setStyleSheet(f"QFrame{{background:{COLORS['bg_surface']};border-radius:10px;border:2px solid {mr};}}")
            kl=QVBoxLayout(kutu); kl.setContentsMargins(12,10,12,10); kl.setAlignment(Qt.AlignCenter)
            for txt,st in [(icon,"font-size:24px;background:transparent;border:none;"),
                           (f"{ogr['ad']} {ogr['soyad'] or ''}",f"font-weight:bold;color:{mr};background:transparent;border:none;"),
                           (f"Lv.{ogr['level']}  ·  {ogr['xp']} XP",f"font-size:11px;color:{COLORS['text_sec']};background:transparent;border:none;")]:
                l=QLabel(txt); l.setAlignment(Qt.AlignCenter); l.setStyleSheet(st); kl.addWidget(l)
            lb_lay.addWidget(kutu)
        lay.addWidget(lb_frame)
        # Tier dağılımı
        tf=QGroupBox("Kurs Tier Dağılımı"); tl=QHBoxLayout(tf)
        for tier_kodu,meta in TIER_META.items():
            kutu=QFrame(); kutu.setStyleSheet(f"QFrame{{background:{COLORS['bg_surface']};border-radius:10px;border-left:5px solid {meta['renk']};}}")
            kl=QVBoxLayout(kutu); kl.setContentsMargins(12,10,12,10)
            lt=QLabel(f"{meta['icon']} {meta['label']}"); lt.setStyleSheet(f"color:{meta['renk']};font-weight:bold;font-size:13px;background:transparent;border:none;"); kl.addWidget(lt)
            lf=QLabel("Ücretsiz" if meta["fiyat"]==0 else f"₺{meta['fiyat']}"); lf.setStyleSheet(f"color:{COLORS['text_sec']};font-size:11px;background:transparent;border:none;"); kl.addWidget(lf)
            tl.addWidget(kutu)
        lay.addWidget(tf); lay.addStretch()



# ═══════════════════════════════════════════════════════════════════════════════
# 🎓 MENTOR ATAMADialog
# ═══════════════════════════════════════════════════════════════════════════════

class MentorAtaDialog(QDialog):
    def __init__(self,db,ogrenci,parent=None):
        super().__init__(parent); self.db=db; self.ogrenci=ogrenci
        self.setWindowTitle(f"Mentor Ata — {ogrenci['ad']}"); self.setFixedSize(440,360)
        self.setStyleSheet(ThemeManager.get_stylesheet()); self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(24,24,24,24); lay.setSpacing(14)
        b=QLabel(f"🧑‍🏫 Mentor Ata: {self.ogrenci['ad']} {self.ogrenci['soyad'] or ''}")
        f=QFont(); f.setPointSize(13); f.setBold(True); b.setFont(f); lay.addWidget(b)
        frm=QFormLayout(); frm.setSpacing(10)
        self.c_eg=QComboBox()
        for e in self.db.tum_egitmenler_getir():
            self.c_eg.addItem(f"{e['ad']} — {e['uzmanlik']}",userData=e["id"])
        frm.addRow("Eğitmen:",self.c_eg)
        self.i_not=QTextEdit(); self.i_not.setPlaceholderText("Notlar, hedefler..."); self.i_not.setMaximumHeight(100)
        frm.addRow("Notlar:",self.i_not); lay.addLayout(frm)
        row=QHBoxLayout(); bi=AdminLuxBtn("İptal",COLORS["danger"]); bi.clicked.connect(self.reject); row.addWidget(bi)
        row.addStretch(); ba=AdminLuxBtn("✅ Ata",COLORS["success"]); ba.clicked.connect(self._ata); row.addWidget(ba)
        lay.addLayout(row)
    def _ata(self):
        eid=self.c_eg.currentData()
        if not eid: return msg_warn(self,"Uyarı","Eğitmen seçin!")
        ok,msg=self.db.mentor_ata(self.ogrenci["id"],eid,self.i_not.toPlainText().strip())
        if ok:
            self.db.xp_ekle(self.ogrenci["id"],"profil_doldur","Mentor atandı")
            msg_info(self,"Başarılı",msg); self.accept()
        else: msg_error(self,"Hata",msg)


# ═══════════════════════════════════════════════════════════════════════════════
# 🏆 SERTİFİKA ONAY DIALOG
# ═══════════════════════════════════════════════════════════════════════════════

class SertifikaOnayDialog(QDialog):
    def __init__(self,db,sertifika,parent=None):
        super().__init__(parent); self.db=db; self.sert=sertifika
        self.setWindowTitle(f"Sertifika — {sertifika['ogr_ad']}"); self.setFixedSize(480,400)
        self.setStyleSheet(ThemeManager.get_stylesheet()); self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(24,24,24,24); lay.setSpacing(14)
        b=QLabel("📜 Sertifika Talebi"); f=QFont(); f.setPointSize(14); f.setBold(True); b.setFont(f); lay.addWidget(b)
        # Bilgi kartı
        kart=QFrame(); kart.setStyleSheet(f"QFrame{{background:{COLORS['bg_card']};border-radius:10px;border:1px solid {COLORS['border']};}}")
        kl=QVBoxLayout(kart); kl.setContentsMargins(16,14,16,14); kl.setSpacing(8)
        tier_renk=TIER_META.get(self.sert["tier"],{}).get("renk",COLORS["primary"])
        for lbl,val,renk in [
            ("👤 Öğrenci",f"{self.sert['ogr_ad']} {self.sert['soyad'] or ''}",COLORS["text_main"]),
            ("📚 Kurs",self.sert["kurs_adi"],tier_renk),
            ("📅 Talep",str(self.sert["talep_tarihi"])[:16],COLORS["text_sec"]),
            ("📊 Durum",self.sert["durum"].upper(),COLORS["warning"]),
        ]:
            row=QHBoxLayout()
            ll=QLabel(lbl); ll.setStyleSheet(f"color:{COLORS['text_sec']};font-size:11px;min-width:100px;background:transparent;border:none;")
            lv=QLabel(str(val or "—")); lv.setStyleSheet(f"color:{renk};font-weight:bold;font-size:12px;background:transparent;border:none;")
            row.addWidget(ll); row.addWidget(lv); row.addStretch(); kl.addLayout(row)
        lay.addWidget(kart)
        # Eğitmen seç (onay için)
        frm=QFormLayout(); frm.setSpacing(10)
        self.c_eg=QComboBox()
        for e in self.db.tum_egitmenler_getir(): self.c_eg.addItem(e["ad"],userData=e["id"])
        frm.addRow("Onaylayan Eğitmen:",self.c_eg); lay.addLayout(frm)
        lay.addStretch()
        row=QHBoxLayout()
        br=AdminLuxBtn("❌ Reddet",COLORS["danger"]); br.clicked.connect(self._reddet); row.addWidget(br)
        row.addStretch()
        bo=AdminLuxBtn("✅ Onayla & PDF",COLORS["success"]); bo.clicked.connect(self._onayla); row.addWidget(bo)
        lay.addLayout(row)
    def _onayla(self):
        eid=self.c_eg.currentData()
        ok,msg=self.db.sertifika_onayla(self.sert["id"],eid)
        if ok:
            bilgi=self.db.sertifika_bilgi_getir(self.sert["id"])
            self._pdf_olustur(bilgi)
            msg_info(self,"Başarılı",msg+"\nPDF kaydedildi!"); self.accept()
        else: msg_error(self,"Hata",msg)
    def _reddet(self):
        if msg_question(self,"Reddet","Talep reddedilecek. Emin misiniz?"):
            ok,msg=self.db.sertifika_reddet(self.sert["id"])
            msg_info(self,"Bilgi",msg); self.accept()
    def _pdf_olustur(self,bilgi):
        if not MATPLOTLIB_OK: return
        try:
            fig=plt.figure(figsize=(11.7,8.3),facecolor="#0f0f1a")
            ax=fig.add_axes([0,0,1,1]); ax.set_xlim(0,1); ax.set_ylim(0,1); ax.axis("off")
            ax.set_facecolor("#0f0f1a")
            # Dış çerçeve
            from matplotlib.patches import FancyBboxPatch
            outer=FancyBboxPatch((0.03,0.03),0.94,0.94,boxstyle="round,pad=0.01",
                                  linewidth=3,edgecolor="#ffd700",facecolor="#1a1a2e")
            ax.add_patch(outer)
            inner=FancyBboxPatch((0.05,0.05),0.90,0.90,boxstyle="round,pad=0.01",
                                  linewidth=1,edgecolor="#0077b6",facecolor="none")
            ax.add_patch(inner)
            # İçerik
            tier_renk=TIER_META.get(bilgi["tier"] or "free",{}).get("renk","#0077b6")
            ax.text(0.5,0.88,"🎓 BAŞARI SERTİFİKASI",ha="center",va="center",fontsize=24,
                    fontweight="bold",color="#ffd700",transform=ax.transAxes)
            ax.text(0.5,0.80,"Online Kurs Platformu",ha="center",va="center",fontsize=14,
                    color="#a1a1aa",transform=ax.transAxes)
            ax.axhline(y=0.76,xmin=0.1,xmax=0.9,color="#0077b6",linewidth=1.5)
            ax.text(0.5,0.70,"Bu sertifika aşağıdaki kişiye takdim edilmiştir:",
                    ha="center",va="center",fontsize=12,color="#a1a1aa")
            ogr_adi=f"{bilgi['ogr_ad']} {bilgi['soyad'] or ''}".strip() if 'soyad' in bilgi.keys() else bilgi['ogr_ad']
            ax.text(0.5,0.62,ogr_adi,ha="center",va="center",fontsize=22,
                    fontweight="bold",color="#f4f4f5")
            ax.text(0.5,0.53,f"«{bilgi['kurs_adi']}»",ha="center",va="center",fontsize=16,
                    color=tier_renk,style="italic")
            ax.text(0.5,0.45,f"kursunu başarıyla tamamlamıştır.",ha="center",va="center",
                    fontsize=13,color="#a1a1aa")
            # Alt bilgiler
            ax.text(0.25,0.32,f"Sertifika No: {bilgi['sertifika_no']}",ha="center",va="center",
                    fontsize=10,color="#ffd700")
            ax.text(0.5,0.32,f"Süre: {bilgi['sure_saat']} saat",ha="center",va="center",
                    fontsize=10,color="#a1a1aa")
            ax.text(0.75,0.32,f"XP: {bilgi['toplam_xp']}  |  Badge: {bilgi['badge_sayisi']}",
                    ha="center",va="center",fontsize=10,color="#a1a1aa")
            ax.axhline(y=0.27,xmin=0.1,xmax=0.9,color="#0077b6",linewidth=0.8)
            ax.text(0.3,0.20,f"Onaylayan: {bilgi['eg_ad']}",ha="center",va="center",
                    fontsize=11,color="#f4f4f5")
            onay_tarihi=str(bilgi["onay_tarihi"] or "")[:10]
            ax.text(0.7,0.20,f"Tarih: {onay_tarihi}",ha="center",va="center",
                    fontsize=11,color="#f4f4f5")
            # Alt logo
            ax.text(0.5,0.10,"★  Online Kurs Platformu — TIER 4  ★",ha="center",va="center",
                    fontsize=10,color="#3f3f46")
            dosya=f"sertifika_{bilgi['sertifika_no']}.pdf"
            plt.savefig(dosya,format="pdf",bbox_inches="tight",dpi=150,facecolor=fig.get_facecolor())
            plt.close(fig)
        except Exception as e:
            print(f"PDF hatası: {e}")


# ═══════════════════════════════════════════════════════════════════════════════
# 📊 GRAFİK SEKMESİ (Matplotlib)
# ═══════════════════════════════════════════════════════════════════════════════


class QtBarChart(QWidget):
    """PyQt5 ile çizilen bar chart — matplotlib yok"""
    def __init__(self, data, colors, title, xlabel="", ylabel="", horizontal=False, parent=None):
        super().__init__(parent)
        self.data = data; self.colors = colors; self.title = title
        self.xlabel = xlabel; self.ylabel = ylabel; self.horizontal = horizontal
        self.setMinimumHeight(300)

    def paintEvent(self, e):
        from PyQt5.QtGui import QPainter, QColor, QFont, QPen, QBrush
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        W, H = self.width(), self.height()
        # Arka plan
        p.fillRect(0, 0, W, H, QColor("#1a1a2e"))
        if not self.data: p.end(); return

        pad_l=60; pad_r=24; pad_t=44; pad_b=54
        chart_w = W - pad_l - pad_r
        chart_h = H - pad_t - pad_b

        # Başlık
        p.setPen(QColor("#f4f4f5"))
        tf = QFont("Segoe UI", 11, QFont.Bold); p.setFont(tf)
        p.drawText(0, 8, W, 28, 0x0084, self.title)  # AlignHCenter|AlignVCenter

        max_val = max((v for _,v in self.data), default=1) or 1

        if not self.horizontal:
            n = len(self.data)
            bar_w = max(8, int(chart_w / (n*1.5+0.5)))
            gap   = max(4, (chart_w - n*bar_w) // max(n+1,1))
            lf = QFont("Segoe UI", 8); p.setFont(lf)
            for i,(lbl,val) in enumerate(self.data):
                bh = int(chart_h * val / max_val)
                x  = pad_l + gap + i*(bar_w+gap)
                y  = pad_t + chart_h - bh
                clr = self.colors[i % len(self.colors)]
                p.setBrush(QBrush(QColor(clr))); p.setPen(Qt.NoPen)
                p.drawRoundedRect(x, y, bar_w, bh, 4, 4)
                # Değer
                p.setPen(QColor("#f4f4f5")); p.setFont(lf)
                p.drawText(x-8, y-18, bar_w+16, 16, 0x0084, str(val))
                # Label
                p.setPen(QColor("#a1a1aa"))
                p.drawText(x-20, H-pad_b+4, bar_w+40, 20, 0x0084,
                           lbl[:10]+"…" if len(lbl)>10 else lbl)
            # Y ekseni çizgisi
            p.setPen(QPen(QColor("#3f3f46"), 1))
            p.drawLine(pad_l, pad_t, pad_l, pad_t+chart_h)
            p.drawLine(pad_l, pad_t+chart_h, pad_l+chart_w, pad_t+chart_h)
        else:
            n = len(self.data)
            bar_h = max(8, int(chart_h / (n*1.5+0.5)))
            gap   = max(4, (chart_h - n*bar_h) // max(n+1,1))
            lf = QFont("Segoe UI", 8); p.setFont(lf)
            for i,(lbl,val) in enumerate(self.data):
                bw = int(chart_w * val / max_val)
                x  = pad_l
                y  = pad_t + gap + i*(bar_h+gap)
                clr = self.colors[i % len(self.colors)]
                p.setBrush(QBrush(QColor(clr))); p.setPen(Qt.NoPen)
                p.drawRoundedRect(x, y, bw, bar_h, 4, 4)
                # Değer
                p.setPen(QColor("#f4f4f5")); p.setFont(lf)
                p.drawText(x+bw+4, y, 60, bar_h, 0x0021, f"{val} XP")
                # Label
                p.setPen(QColor("#a1a1aa"))
                p.drawText(0, y, pad_l-4, bar_h, 0x0022,
                           lbl[:10]+"…" if len(lbl)>10 else lbl)
            p.setPen(QPen(QColor("#3f3f46"), 1))
            p.drawLine(pad_l, pad_t, pad_l, pad_t+chart_h)
            p.drawLine(pad_l, pad_t+chart_h, pad_l+chart_w, pad_t+chart_h)
        p.end()


class QtPieChart(QWidget):
    """PyQt5 ile çizilen pie chart"""
    def __init__(self, data, colors, title, parent=None):
        super().__init__(parent)
        self.data=data; self.colors=colors; self.title=title
        self.setMinimumHeight(300)

    def paintEvent(self, e):
        from PyQt5.QtGui import QPainter, QColor, QPen, QBrush, QFont
        from PyQt5.QtCore import QRect
        import math
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        W, H = self.width(), self.height()
        p.fillRect(0, 0, W, H, QColor("#1a1a2e"))
        if not self.data: p.end(); return

        p.setPen(QColor("#f4f4f5"))
        tf = QFont("Segoe UI", 11, QFont.Bold); p.setFont(tf)
        p.drawText(0, 8, W, 28, 0x0084, self.title)

        total = sum(v for _,v in self.data) or 1
        cx, cy = W//2, H//2 + 10
        r = min(W, H-60) // 2 - 20
        angle = -90.0

        for i,(lbl,val) in enumerate(self.data):
            span = 360.0 * val / total
            clr  = self.colors[i % len(self.colors)]
            p.setBrush(QBrush(QColor(clr))); p.setPen(Qt.NoPen)
            p.drawPie(cx-r, cy-r, 2*r, 2*r,
                      int(angle*16), int(span*16))
            # Yüzde etiketi
            mid_a = math.radians(angle + span/2)
            tx = cx + int(r*0.65*math.cos(mid_a))
            ty = cy + int(r*0.65*math.sin(mid_a))
            p.setPen(QColor("#1a1a2e"))
            f2 = QFont("Segoe UI", 8, QFont.Bold); p.setFont(f2)
            pct = f"{val/total*100:.0f}%"
            p.drawText(tx-20, ty-8, 40, 16, 0x0084, pct)
            angle += span

        # Legend
        lf = QFont("Segoe UI", 8); p.setFont(lf)
        lx = 12; ly = H - 20 - len(self.data)*16
        for i,(lbl,val) in enumerate(self.data):
            clr = self.colors[i % len(self.colors)]
            p.fillRect(lx, ly+i*16, 10, 10, QColor(clr))
            p.setPen(QColor("#a1a1aa"))
            p.drawText(lx+14, ly+i*16, 120, 14, 0x0021, f"{lbl} ({val})")
        p.end()


class QtLineChart(QWidget):
    """PyQt5 ile çizilen line chart"""
    def __init__(self, labels, values, title, color="#0077b6", parent=None):
        super().__init__(parent)
        self.labels=labels; self.values=values
        self.title=title; self.color=color
        self.setMinimumHeight(300)

    def paintEvent(self, e):
        from PyQt5.QtGui import QPainter, QColor, QPen, QBrush, QFont, QPolygonF
        from PyQt5.QtCore import QPointF
        p = QPainter(self); p.setRenderHint(QPainter.Antialiasing)
        W, H = self.width(), self.height()
        p.fillRect(0, 0, W, H, QColor("#1a1a2e"))
        if not self.values: p.end(); return

        p.setPen(QColor("#f4f4f5"))
        tf = QFont("Segoe UI", 11, QFont.Bold); p.setFont(tf)
        p.drawText(0, 8, W, 28, 0x0084, self.title)

        pad_l=60; pad_r=24; pad_t=44; pad_b=54
        cw = W-pad_l-pad_r; ch = H-pad_t-pad_b
        max_v = max(self.values) or 1
        n = len(self.values)

        # Grid çizgileri
        p.setPen(QPen(QColor("#2d2d3a"), 1))
        for i in range(4):
            y = pad_t + ch*i//3
            p.drawLine(pad_l, y, pad_l+cw, y)

        # Alan dolgusu
        if n > 1:
            pts = []
            for i,v in enumerate(self.values):
                x = pad_l + i*cw//(n-1) if n>1 else pad_l+cw//2
                y = pad_t + ch - int(ch*v/max_v)
                pts.append(QPointF(x,y))
            # Polygon
            poly = QPolygonF([QPointF(pts[0].x(),pad_t+ch)] +
                             pts + [QPointF(pts[-1].x(),pad_t+ch)])
            clr = QColor(self.color); clr.setAlpha(40)
            p.setBrush(QBrush(clr)); p.setPen(Qt.NoPen)
            p.drawPolygon(poly)
            # Çizgi
            p.setPen(QPen(QColor(self.color), 2))
            p.setBrush(Qt.NoBrush)
            for i in range(len(pts)-1):
                p.drawLine(pts[i], pts[i+1])
            # Noktalar
            p.setBrush(QBrush(QColor("#ffd700")))
            p.setPen(Qt.NoPen)
            for pt in pts:
                p.drawEllipse(pt, 5, 5)

        # X etiketleri
        lf = QFont("Segoe UI", 7); p.setFont(lf)
        p.setPen(QColor("#a1a1aa"))
        for i,lbl in enumerate(self.labels):
            x = pad_l + i*cw//(n-1) if n>1 else pad_l+cw//2
            p.drawText(x-25, H-pad_b+6, 50, 16, 0x0084,
                       lbl[:8]+"…" if len(lbl)>8 else lbl)

        p.setPen(QPen(QColor("#3f3f46"), 1))
        p.drawLine(pad_l, pad_t, pad_l, pad_t+ch)
        p.drawLine(pad_l, pad_t+ch, pad_l+cw, pad_t+ch)
        p.end()


class GrafikSekmesi(QWidget):
    def __init__(self, db):
        super().__init__(); self.db = db
        self._chart_widget = None
        self._build()

    def _build(self):
        self._lay = QVBoxLayout(self)
        self._lay.setContentsMargins(20,20,20,20); self._lay.setSpacing(14)

        # Başlık
        lbl = QLabel("Istatistik Grafikleri")
        f = QFont("Segoe UI", 14, QFont.Bold); lbl.setFont(f)
        lbl.setStyleSheet(f"color:{COLORS['accent']};background:transparent;border:none;")
        self._lay.addWidget(lbl)

        # Kontrol barı
        tb = QHBoxLayout()
        lbl2 = QLabel("Grafik türü:")
        lbl2.setStyleSheet("background:transparent;border:none;font-size:13px;")
        tb.addWidget(lbl2)
        self.cmb = QComboBox()
        self.cmb.setFixedWidth(220)
        for g in ["Paket Dagilimi","Kurs Tamamlanma","XP Siralaması","Kayit Trendi"]:
            self.cmb.addItem(g)
        self.cmb.currentIndexChanged.connect(self._goster)
        tb.addWidget(self.cmb); tb.addStretch()
        btn = AdminLuxBtn("Goster", COLORS["primary"])
        btn.clicked.connect(self._goster); tb.addWidget(btn)
        self._lay.addLayout(tb)

        # Grafik alanı — placeholder
        self._placeholder = QLabel("Grafik yükleniyor...")
        self._placeholder.setAlignment(Qt.AlignCenter)
        self._placeholder.setMinimumHeight(340)
        self._placeholder.setStyleSheet(
            f"color:{COLORS['text_sec']};font-size:14px;background:{COLORS['bg_card']};"
            f"border-radius:14px;border:1px solid {COLORS['border']};")
        self._lay.addWidget(self._placeholder, stretch=1)

    def showEvent(self, event):
        super().showEvent(event)
        QTimer.singleShot(80, self._goster)

    def _goster(self):
        # Eski grafik widget'ını kaldır
        if self._chart_widget is not None:
            self._lay.removeWidget(self._chart_widget)
            self._chart_widget.hide()
            self._chart_widget.deleteLater()
            self._chart_widget = None

        self._placeholder.show()

        try:
            idx = self.cmb.currentIndex()
            v   = self.db.grafik_verileri_getir()

            w = None
            if idx == 0:  # Paket dağılımı
                td = v["tier_data"]
                data   = [(TIER_META.get(k,{}).get("label",k), cnt) for k,cnt in td.items()]
                colors = [TIER_META.get(k,{}).get("renk","#0077b6") for k in td]
                w = QtPieChart(data, colors, "Ogrenci Paket Dagilimi")

            elif idx == 1:  # Kurs tamamlanma
                kt = v["kurs_tamamlanma"]
                data   = [(r["ad"][:12], r["toplam"] or 0) for r in kt]
                colors = [COLORS["primary"]]*len(data)
                w = QtBarChart(data, colors, "Kurs Tamamlanma", ylabel="Ogrenci")

            elif idx == 2:  # XP sıralaması
                xd = v["xp_data"]
                data   = [(r["ad"], r["xp"]) for r in xd]
                colors = ["#ffd700","#c0c0c0","#cd7f32"] + [COLORS["primary"]]*(len(data)-3)
                w = QtBarChart(data, colors, "XP Siralaması", horizontal=True)

            elif idx == 3:  # Kayıt trendi
                kt = v["kayit_trend"]
                if kt:
                    labels = [r["ay"]  for r in kt]
                    values = [r["cnt"] for r in kt]
                else:
                    labels, values = ["Veri yok"], [0]
                w = QtLineChart(labels, values, "Aylik Kayit Trendi")

            if w:
                w.setStyleSheet(f"background:{COLORS['bg_card']};border-radius:14px;border:1px solid {COLORS['border']};")
                w.setMinimumHeight(340)
                self._placeholder.hide()
                self._lay.addWidget(w, stretch=1)
                self._chart_widget = w
                w.update()

        except Exception as e:
            import traceback; traceback.print_exc()
            self._placeholder.setText(f"Grafik hatası: {e}")
            self._placeholder.show()

class MentorSertifikaSekmesi(QWidget):
    def __init__(self,db):
        super().__init__(); self.db=db; self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(16,16,16,16); lay.setSpacing(12)
        lbl=QLabel("🧑‍🏫 Mentor & Sertifika Yönetimi"); f=QFont(); f.setPointSize(13); f.setBold(True); lbl.setFont(f)
        lay.addWidget(lbl)
        tabs=QTabWidget(); lay.addWidget(tabs)
        # ── Mentor alt sekmesi
        mentor_w=QWidget(); ml=QVBoxLayout(mentor_w); ml.setContentsMargins(12,12,12,12); ml.setSpacing(10)
        tb=QHBoxLayout()
        self.cmb_m_filtre=QComboBox()
        for f_,l_ in [("hepsi","Tümü"),("aktif","Aktif"),("beklemede","Beklemede"),("tamamlandi","Tamamlandı")]:
            self.cmb_m_filtre.addItem(l_,userData=f_)
        self.cmb_m_filtre.currentIndexChanged.connect(self._mentor_yenile); tb.addWidget(QLabel("Filtre:")); tb.addWidget(self.cmb_m_filtre)
        tb.addStretch()
        btn_ata=AdminLuxBtn("➕ Mentor Ata",COLORS["success"]); btn_ata.clicked.connect(self._mentor_ata); tb.addWidget(btn_ata)
        btn_sil=AdminLuxBtn("🗑 Sil",COLORS["danger"]); btn_sil.clicked.connect(self._mentor_sil); tb.addWidget(btn_sil)
        ml.addLayout(tb)
        self.mentor_tablo=QTableWidget(); self.mentor_tablo.setColumnCount(6)
        self.mentor_tablo.setHorizontalHeaderLabels(["ID","Öğrenci","Paket","Eğitmen","Uzmanlık","Durum"])
        self.mentor_tablo.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.mentor_tablo.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.mentor_tablo.setAlternatingRowColors(True); self.mentor_tablo.setStyleSheet(tablo_stili())
        self.mentor_tablo.verticalHeader().setVisible(False)
        self.mentor_tablo.horizontalHeader().setSectionResizeMode(1,QHeaderView.Stretch)
        ml.addWidget(self.mentor_tablo)
        tabs.addTab(mentor_w,"🧑‍🏫 Mentor İlişkileri")
        # ── Sertifika alt sekmesi
        sert_w=QWidget(); sl=QVBoxLayout(sert_w); sl.setContentsMargins(12,12,12,12); sl.setSpacing(10)
        tb2=QHBoxLayout()
        self.cmb_s_filtre=QComboBox()
        for f_,l_ in [("hepsi","Tümü"),("beklemede","Beklemede"),("onaylandi","Onaylandı"),("reddedildi","Reddedildi")]:
            self.cmb_s_filtre.addItem(l_,userData=f_)
        self.cmb_s_filtre.currentIndexChanged.connect(self._sert_yenile); tb2.addWidget(QLabel("Filtre:")); tb2.addWidget(self.cmb_s_filtre)
        tb2.addStretch()
        btn_talep=AdminLuxBtn("📝 Talep Oluştur",COLORS["primary"]); btn_talep.clicked.connect(self._sert_talep); tb2.addWidget(btn_talep)
        btn_incele=AdminLuxBtn("🔍 İncele / Onayla",COLORS["success"]); btn_incele.clicked.connect(self._sert_incele); tb2.addWidget(btn_incele)
        sl.addLayout(tb2)
        self.sert_tablo=QTableWidget(); self.sert_tablo.setColumnCount(7)
        self.sert_tablo.setHorizontalHeaderLabels(["ID","Öğrenci","Kurs","Tier","Durum","Talep Tarihi","Sertifika No"])
        self.sert_tablo.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.sert_tablo.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.sert_tablo.setAlternatingRowColors(True); self.sert_tablo.setStyleSheet(tablo_stili())
        self.sert_tablo.verticalHeader().setVisible(False)
        self.sert_tablo.horizontalHeader().setSectionResizeMode(2,QHeaderView.Stretch)
        sl.addWidget(self.sert_tablo)
        tabs.addTab(sert_w,"📜 Sertifikalar")
        self._mentor_yenile(); self._sert_yenile()

    def _mentor_yenile(self):
        filtre=self.cmb_m_filtre.currentData()
        rows=self.db.mentor_iliskileri_getir(filtre); self.mentor_tablo.setRowCount(len(rows))
        durum_renk={"aktif":COLORS["success"],"beklemede":COLORS["warning"],"tamamlandi":COLORS["text_sec"]}
        for r,row in enumerate(rows):
            meta=TIER_META.get(row["paket"],{})
            for col,val in enumerate([str(row["id"]),
                f"{row['ogr_ad']} {row['ogr_soyad'] or ''}",
                f"{meta.get('icon','')} {row['paket'].upper()}",
                row["eg_ad"],row["uzmanlik"] or "—",row["durum"]]):
                item=QTableWidgetItem(val); item.setTextAlignment(Qt.AlignCenter)
                if col==5: item.setForeground(QColor(durum_renk.get(row["durum"],COLORS["text_main"])))
                if col==2: item.setForeground(QColor(meta.get("renk",COLORS["primary"])))
                self.mentor_tablo.setItem(r,col,item)
        self.mentor_tablo.resizeColumnsToContents(); self.mentor_tablo.setColumnWidth(1,160)

    def _mentor_ata(self):
        ogr_list=self.db.tum_ogrenciler_getir()
        if not ogr_list: return msg_warn(self,"Uyarı","Öğrenci bulunamadı.")
        # Öğrenci seçim combo
        dlg=QDialog(self); dlg.setWindowTitle("Öğrenci Seç"); dlg.setFixedSize(360,160)
        dlg.setStyleSheet(ThemeManager.get_stylesheet())
        dl=QVBoxLayout(dlg); dl.setContentsMargins(20,20,20,20); dl.setSpacing(12)
        dl.addWidget(QLabel("Mentor atanacak öğrenciyi seçin:"))
        cmb=QComboBox()
        for o in ogr_list: cmb.addItem(f"{o['ad']} {o['soyad'] or ''} ({o['paket'].upper()})",userData=dict(o))
        dl.addWidget(cmb)
        btn=AdminLuxBtn("Devam →",COLORS["primary"]); btn.clicked.connect(dlg.accept); dl.addWidget(btn)
        if dlg.exec_()!=QDialog.Accepted: return
        ogr=cmb.currentData()
        mentor_dlg=MentorAtaDialog(self.db,ogr,self)
        if mentor_dlg.exec_()==QDialog.Accepted: self._mentor_yenile()

    def _mentor_sil(self):
        row=self.mentor_tablo.currentRow()
        if row<0: return msg_warn(self,"Uyarı","Bir mentor ilişkisi seçin!")
        mid=int(self.mentor_tablo.item(row,0).text())
        if msg_question(self,"Sil","Bu mentor ilişkisi silinecek. Onaylıyor musunuz?"):
            ok,msg=self.db.mentor_sil(mid)
            msg_info(self,"Bilgi",msg); self._mentor_yenile()

    def _sert_yenile(self):
        filtre=self.cmb_s_filtre.currentData()
        rows=self.db.sertifika_talep_getir(filtre); self.sert_tablo.setRowCount(len(rows))
        durum_renk={"beklemede":COLORS["warning"],"onaylandi":COLORS["success"],"reddedildi":COLORS["danger"]}
        for r,row in enumerate(rows):
            meta=TIER_META.get(row["tier"],{})
            for col,val in enumerate([str(row["id"]),
                f"{row['ogr_ad']} {row['ogr_soyad'] or ''}",
                row["kurs_adi"],f"{meta.get('icon','')} {row['tier'].upper()}",
                row["durum"],str(row["talep_tarihi"])[:10],
                row["sertifika_no"] or "—"]):
                item=QTableWidgetItem(val); item.setTextAlignment(Qt.AlignCenter)
                if col==4: item.setForeground(QColor(durum_renk.get(row["durum"],COLORS["text_main"])))
                if col==3: item.setForeground(QColor(meta.get("renk",COLORS["primary"])))
                if col==6 and row["sertifika_no"]: item.setForeground(QColor(COLORS["gold"]))
                self.sert_tablo.setItem(r,col,item)
        self.sert_tablo.resizeColumnsToContents(); self.sert_tablo.setColumnWidth(2,180)

    def _sert_talep(self):
        # Tamamlanan kurslardan talep oluştur
        ogr_list=self.db.tum_ogrenciler_getir()
        dlg=QDialog(self); dlg.setWindowTitle("Sertifika Talebi"); dlg.setFixedSize(420,200)
        dlg.setStyleSheet(ThemeManager.get_stylesheet())
        dl=QVBoxLayout(dlg); dl.setContentsMargins(20,20,20,20); dl.setSpacing(12)
        dl.addWidget(QLabel("Öğrenci:"))
        cmb_o=QComboBox()
        for o in ogr_list: cmb_o.addItem(f"{o['ad']} {o['soyad'] or ''}",userData=o["id"])
        dl.addWidget(cmb_o)
        dl.addWidget(QLabel("Kurs (yalnızca tamamlananlar gösterilir):"))
        cmb_k=QComboBox(); dl.addWidget(cmb_k)
        def ogr_degisti():
            cmb_k.clear()
            oid=cmb_o.currentData()
            for k in self.db.ogrenci_kayitlar_getir(oid):
                if k["durum"]=="tamamlandi":
                    cmb_k.addItem(k["kurs_adi"],userData=k["kurs_id"])
        cmb_o.currentIndexChanged.connect(ogr_degisti); ogr_degisti()
        btn=AdminLuxBtn("📝 Talep Oluştur",COLORS["primary"]); btn.clicked.connect(dlg.accept); dl.addWidget(btn)
        if dlg.exec_()!=QDialog.Accepted: return
        if cmb_k.count()==0: return msg_warn(self,"Uyarı","Tamamlanmış kurs bulunamadı!")
        ok,msg=self.db.sertifika_talep_olustur(cmb_o.currentData(),cmb_k.currentData())
        msg_info(self,"Bilgi",msg) if ok else msg_error(self,"Hata",msg)
        self._sert_yenile()

    def _sert_incele(self):
        row=self.sert_tablo.currentRow()
        if row<0: return msg_warn(self,"Uyarı","Bir sertifika seçin!")
        sid=int(self.sert_tablo.item(row,0).text())
        serts=self.db.sertifika_talep_getir()
        sert=next((s for s in serts if s["id"]==sid),None)
        if not sert: return
        if sert["durum"]!="beklemede":
            return msg_info(self,"Bilgi",f"Bu sertifika zaten '{sert['durum']}' durumunda.")
        dlg=SertifikaOnayDialog(self.db,sert,self)
        if dlg.exec_()==QDialog.Accepted: self._sert_yenile()


# ═══════════════════════════════════════════════════════════════════════════════
# 📋 RAPORLAR SEKMESİ
# ═══════════════════════════════════════════════════════════════════════════════

class RaporlarSekmesi(QWidget):
    def __init__(self, db):
        super().__init__(); self.db = db; self._build()

    def _build(self):
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea{border:none;background:transparent;}")
        outer = QVBoxLayout(self); outer.setContentsMargins(0,0,0,0)
        outer.addWidget(scroll)

        ctn = QWidget(); lay = QVBoxLayout(ctn)
        lay.setContentsMargins(28,24,28,28); lay.setSpacing(20)

        # ── BAŞLIK ────────────────────────────────────────────────────────
        baslik = QLabel("Raporlar & Export")
        bf = QFont("Segoe UI", 18, QFont.Bold); baslik.setFont(bf)
        baslik.setStyleSheet(f"color:{COLORS['accent']};background:transparent;border:none;")
        lay.addWidget(baslik)

        # ── 1. İSTATİSTİK KARTları ───────────────────────────────────────
        stats = self.db.dashboard_istatistik()
        stat_grid = QGridLayout(); stat_grid.setSpacing(14)

        stat_items = [
            ("Toplam Öğrenci", stats["toplam_ogrenci"],   "👥", COLORS["primary"]),
            ("Pro Öğrenci",    stats["pro_ogrenci"],      "⭐", COLORS["tier_pro2"]),
            ("Toplam Kurs",    stats["toplam_kurs"],      "📚", COLORS["accent"]),
            ("Aktif Eğitmen",  stats["toplam_egitmen"],   "👨‍🏫", COLORS["teal"]),
            ("Toplam Kayıt",   stats["toplam_kayit"],     "📝", COLORS["success"]),
            ("Tamamlanan",     stats["tamamlanan"],       "✅", COLORS["green"]),
            ("Toplam XP",      stats["toplam_xp"],        "⚡", COLORS["xp_color"]),
            ("Toplam Badge",   stats["toplam_badge"],     "🏅", COLORS["gold"]),
        ]

        for i, (lbl_txt, val, ikon, renk) in enumerate(stat_items):
            kart = QFrame()
            kart.setStyleSheet(f"""QFrame{{
                background:qlineargradient(x1:0,y1:0,x2:1,y2:1,
                    stop:0 {COLORS['bg_card']},stop:1 {COLORS['bg_surface']});
                border-radius:16px;
                border-left:5px solid {renk};
                border-top:1px solid rgba(255,255,255,0.06);
                border-right:1px solid rgba(255,255,255,0.03);
                border-bottom:1px solid rgba(0,0,0,0.2);
            }}""")
            kl = QVBoxLayout(kart); kl.setContentsMargins(18,16,18,16); kl.setSpacing(6)

            top_row = QHBoxLayout()
            lbl2 = QLabel(lbl_txt)
            lbl2.setStyleSheet(f"color:{COLORS['text_sec']};font-size:12px;font-weight:600;background:transparent;border:none;")
            top_row.addWidget(lbl2); top_row.addStretch()
            ik_lbl = QLabel(ikon)
            ik_lbl.setStyleSheet(f"font-size:18px;background:rgba(255,255,255,0.05);border-radius:8px;padding:3px 6px;border:none;")
            top_row.addWidget(ik_lbl)
            kl.addLayout(top_row)

            val_lbl = QLabel(f"{val:,}")
            vf = QFont("Segoe UI", 26, QFont.Bold); val_lbl.setFont(vf)
            val_lbl.setStyleSheet(f"color:{renk};background:transparent;border:none;letter-spacing:-0.5px;")
            kl.addWidget(val_lbl)

            stat_grid.addWidget(kart, i//4, i%4)

        lay.addLayout(stat_grid)

        # ── 2. EXCEL EXPORT ────────────────────────────────────────────────
        excel_kart = QFrame()
        excel_kart.setStyleSheet(f"""QFrame{{
            background:qlineargradient(x1:0,y1:0,x2:1,y2:0,
                stop:0 {COLORS['bg_card']},stop:1 {COLORS['bg_surface']});
            border-radius:18px;
            border:1px solid {COLORS['tier_pro2']};
        }}""")
        ek = QHBoxLayout(excel_kart); ek.setContentsMargins(28,22,28,22); ek.setSpacing(20)

        ikon_frame = QFrame()
        ikon_frame.setFixedSize(64,64)
        ikon_frame.setStyleSheet(f"QFrame{{background:{COLORS['tier_pro2']}22;border-radius:14px;border:1px solid {COLORS['tier_pro2']}44;}}")
        il = QVBoxLayout(ikon_frame); il.setAlignment(Qt.AlignCenter)
        ik2 = QLabel("📊"); ik2.setStyleSheet("font-size:28px;background:transparent;border:none;"); il.addWidget(ik2)
        ek.addWidget(ikon_frame)

        tv = QVBoxLayout(); tv.setSpacing(4)
        t1 = QLabel("Excel Raporu Oluştur")
        t1f = QFont("Segoe UI",15,QFont.Bold); t1.setFont(t1f)
        t1.setStyleSheet(f"color:{COLORS['text_main']};background:transparent;border:none;")
        t2 = QLabel("Öğrenciler, Kurslar ve Sertifikaları 3 sayfalı Excel dosyasına aktar")
        t2.setStyleSheet(f"color:{COLORS['text_sec']};font-size:12px;background:transparent;border:none;")
        tv.addWidget(t1); tv.addWidget(t2)
        ek.addLayout(tv); ek.addStretch()

        btn_xl = QPushButton("  Excel İndir")
        btn_xl.setFixedSize(160,46)
        btn_xl.setStyleSheet(f"""QPushButton{{
            background:{COLORS['tier_pro2']};color:white;border:none;
            border-radius:12px;font-weight:bold;font-size:13px;
        }}QPushButton:hover{{background:#047857;}}""")
        btn_xl.clicked.connect(self._excel_export); ek.addWidget(btn_xl)
        lay.addWidget(excel_kart)

        # ── 3. ÖZET RAPOR ─────────────────────────────────────────────────
        ozet_kart = QFrame()
        ozet_kart.setStyleSheet(f"""QFrame{{
            background:{COLORS['bg_card']};border-radius:18px;
            border:1px solid {COLORS['border']};
        }}""")
        ok = QVBoxLayout(ozet_kart); ok.setContentsMargins(28,22,28,22); ok.setSpacing(14)

        ozet_header = QHBoxLayout()
        ozet_title = QLabel("Platform Özet Raporu")
        otf = QFont("Segoe UI",14,QFont.Bold); ozet_title.setFont(otf)
        ozet_title.setStyleSheet(f"color:{COLORS['text_main']};background:transparent;border:none;")
        ozet_header.addWidget(ozet_title); ozet_header.addStretch()
        btn_yenile = QPushButton("  Raporu Yenile")
        btn_yenile.setFixedSize(150,38)
        btn_yenile.setStyleSheet(f"""QPushButton{{
            background:{COLORS['primary']};color:white;border:none;
            border-radius:10px;font-weight:bold;font-size:12px;
        }}QPushButton:hover{{background:{COLORS['primary_h']};}}""")
        btn_yenile.clicked.connect(self._ozet_goster)
        ozet_header.addWidget(btn_yenile)
        ok.addLayout(ozet_header)

        # Özet içerik — grid formatında
        self.ozet_grid_w = QWidget()
        self.ozet_grid_w.setStyleSheet("background:transparent;")
        self.ozet_grid = QGridLayout(self.ozet_grid_w)
        self.ozet_grid.setSpacing(10)
        ok.addWidget(self.ozet_grid_w)

        lay.addWidget(ozet_kart)

        # ── 4. SERTİFİKA ÖZETİ ────────────────────────────────────────────
        sert_kart = QFrame()
        sert_kart.setStyleSheet(f"""QFrame{{
            background:{COLORS['bg_card']};border-radius:18px;
            border:1px solid {COLORS['border']};
        }}""")
        sk = QVBoxLayout(sert_kart); sk.setContentsMargins(28,22,28,22); sk.setSpacing(14)

        sert_title = QLabel("Sertifika Durumu")
        stf = QFont("Segoe UI",14,QFont.Bold); sert_title.setFont(stf)
        sert_title.setStyleSheet(f"color:{COLORS['text_main']};background:transparent;border:none;")
        sk.addWidget(sert_title)

        self.sert_row_w = QHBoxLayout()
        self.sert_row_w.setSpacing(14)
        sk.addLayout(self.sert_row_w)
        lay.addWidget(sert_kart)

        lay.addStretch()
        scroll.setWidget(ctn)

        # İlk yükleme
        self._ozet_goster()

    def _ozet_kart(self, lbl, val, renk, ikon):
        kart = QFrame()
        kart.setStyleSheet(f"""QFrame{{
            background:{COLORS['bg_surface']};border-radius:14px;
            border-top:3px solid {renk};
            border-left:none;border-right:none;border-bottom:none;
        }}""")
        kl = QVBoxLayout(kart); kl.setContentsMargins(16,14,16,14); kl.setSpacing(6); kl.setAlignment(Qt.AlignCenter)
        ik = QLabel(ikon); ik.setAlignment(Qt.AlignCenter)
        ik.setStyleSheet("font-size:24px;background:transparent;border:none;")
        v = QLabel(str(val))
        vf = QFont("Segoe UI",22,QFont.Bold); v.setFont(vf)
        v.setAlignment(Qt.AlignCenter)
        v.setStyleSheet(f"color:{renk};background:transparent;border:none;")
        lb = QLabel(lbl); lb.setAlignment(Qt.AlignCenter)
        lb.setStyleSheet(f"color:{COLORS['text_sec']};font-size:11px;background:transparent;border:none;")
        kl.addWidget(ik); kl.addWidget(v); kl.addWidget(lb)
        return kart

    def _ozet_goster(self):
        stats = self.db.dashboard_istatistik()
        lb = self.db.leaderboard_getir(3)

        # Özet grid temizle
        while self.ozet_grid.count():
            item = self.ozet_grid.takeAt(0)
            if item.widget(): item.widget().deleteLater()

        ozet_items = [
            ("Toplam Öğrenci", stats["toplam_ogrenci"],  COLORS["primary"],   "👥"),
            ("Pro Öğrenci",    stats["pro_ogrenci"],     COLORS["tier_pro2"], "⭐"),
            ("Toplam Kurs",    stats["toplam_kurs"],     COLORS["accent"],    "📚"),
            ("Aktif Eğitmen",  stats["toplam_egitmen"],  COLORS["teal"],      "👨‍🏫"),
            ("Toplam Kayıt",   stats["toplam_kayit"],    COLORS["success"],   "📝"),
            ("Tamamlanan",     stats["tamamlanan"],      COLORS["green"],     "✅"),
            ("Toplam XP",      stats["toplam_xp"],       COLORS["xp_color"],  "⚡"),
            ("Toplam Badge",   stats["toplam_badge"],    COLORS["gold"],      "🏅"),
        ]
        for i,(lbl,val,renk,ikon) in enumerate(ozet_items):
            self.ozet_grid.addWidget(self._ozet_kart(lbl,val,renk,ikon), i//4, i%4)

        # Top 3 leaderboard
        top3_kart = QFrame()
        top3_kart.setStyleSheet(f"QFrame{{background:{COLORS['bg_surface']};border-radius:12px;border:none;}}")
        t3l = QHBoxLayout(top3_kart); t3l.setContentsMargins(14,12,14,12); t3l.setSpacing(8)
        medals = {0:COLORS["gold"],1:COLORS["silver"],2:COLORS["bronze"]}
        icons  = {0:"🥇",1:"🥈",2:"🥉"}
        for i,o in enumerate(lb):
            chip = QFrame()
            chip.setStyleSheet(f"QFrame{{background:{COLORS['bg_card']};border-radius:10px;border:1px solid {medals.get(i,COLORS['border'])};border-left:none;border-right:none;border-top:none;border-bottom-left-radius:0;border-bottom-right-radius:0;}}")
            cl = QVBoxLayout(chip); cl.setContentsMargins(12,10,12,10); cl.setAlignment(Qt.AlignCenter); cl.setSpacing(2)
            cl.addWidget(self._lbl2(icons.get(i,""),"font-size:20px;",Qt.AlignCenter))
            cl.addWidget(self._lbl2(f"{o['ad']} {o['soyad'] or ''}",f"font-weight:bold;font-size:12px;color:{medals.get(i,COLORS['text_main'])};",Qt.AlignCenter))
            cl.addWidget(self._lbl2(f"{o['xp']} XP  ·  Lv.{o['level']}",f"font-size:11px;color:{COLORS['text_sec']};",Qt.AlignCenter))
            t3l.addWidget(chip)
        t3l.addStretch()
        self.ozet_grid.addWidget(top3_kart, 2, 0, 1, 4)

        # Sertifika özeti
        while self.sert_row_w.count():
            item = self.sert_row_w.takeAt(0)
            if item.widget(): item.widget().deleteLater()

        serts = self.db.sertifika_talep_getir()
        bek = sum(1 for s in serts if s["durum"]=="beklemede")
        ona = sum(1 for s in serts if s["durum"]=="onaylandi")
        red = sum(1 for s in serts if s["durum"]=="reddedildi")

        for val,lbl,renk,ikon in [(bek,"Beklemede",COLORS["warning"],"⏳"),
                                    (ona,"Onaylandi", COLORS["success"],"✅"),
                                    (red,"Reddedildi",COLORS["danger"], "❌")]:
            k = QFrame()
            k.setStyleSheet(f"""QFrame{{background:{COLORS['bg_surface']};border-radius:14px;
                border-top:4px solid {renk};border-left:none;border-right:none;border-bottom:none;}}""")
            kl = QVBoxLayout(k); kl.setContentsMargins(20,16,20,16); kl.setSpacing(4); kl.setAlignment(Qt.AlignCenter)
            kl.addWidget(self._lbl2(ikon,"font-size:26px;",Qt.AlignCenter))
            vl = QLabel(str(val)); vf=QFont("Segoe UI",28,QFont.Bold); vl.setFont(vf)
            vl.setAlignment(Qt.AlignCenter); vl.setStyleSheet(f"color:{renk};background:transparent;border:none;")
            kl.addWidget(vl)
            kl.addWidget(self._lbl2(lbl,f"font-size:12px;color:{COLORS['text_sec']};",Qt.AlignCenter))
            self.sert_row_w.addWidget(k)
        self.sert_row_w.addStretch()

    def _lbl2(self, txt, st="", align=None):
        l = QLabel(txt)
        l.setStyleSheet(st+"background:transparent;border:none;")
        if align: l.setAlignment(align)
        return l

    def _lbl(self, txt, sz, bold, renk):
        l = QLabel(txt)
        l.setStyleSheet(f"color:{renk};font-size:{sz}px;{'font-weight:bold;' if bold else ''}background:transparent;border:none;")
        return l

    def _excel_export(self):
        if not OPENPYXL_OK:
            return msg_warn(self,"Uyarı","Excel export için: pip install openpyxl")
        from PyQt5.QtWidgets import QFileDialog
        dosya,_ = QFileDialog.getSaveFileName(self,"Excel Kaydet",
            f"platform_raporu_{__import__('datetime').datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            "Excel (*.xlsx)")
        if not dosya: return
        ok,yol = self.db.excel_rapor_olustur(dosya)
        if ok: msg_info(self,"Başarılı",f"Excel kaydedildi:\n{yol}")
        else:  msg_warn(self,"Hata",str(yol))


class AdminBildirimBtn(QPushButton):
    def __init__(self,db,parent=None):
        super().__init__("🔔",parent); self.db=db
        self.setFixedSize(42,42)
        self.setStyleSheet(f"""QPushButton{{background:{COLORS['bg_surface']};color:white;
            border:1px solid {COLORS['border']};border-radius:21px;font-size:18px;font-weight:bold;}}
            QPushButton:hover{{background:{COLORS['primary']};border-color:{COLORS['primary']};}}""")
        self.timer=QTimer(); self.timer.timeout.connect(self._guncelle); self.timer.start(5000)
        self._guncelle()
    def _guncelle(self):
        n=self.db.okunmamis_bildirim_sayisi()
        self.setText(f"🔔{n}" if n>0 else "🔔")
        stil_renk=COLORS["warning"] if n>0 else COLORS["border"]
        self.setStyleSheet(f"""QPushButton{{background:{COLORS['bg_surface']};color:{'#ffd700' if n>0 else 'white'};
            border:2px solid {stil_renk};border-radius:21px;font-size:{'14px' if n>0 else '18px'};font-weight:bold;}}
            QPushButton:hover{{background:{COLORS['primary']};border-color:{COLORS['primary']};}}""")


# ═══════════════════════════════════════════════════════════════════════════════
# 🔔 BİLDİRİM PANELİ DIALOG
# ═══════════════════════════════════════════════════════════════════════════════

class BildirimPanelDialog(QDialog):
    def __init__(self,db,parent=None):
        super().__init__(parent); self.db=db
        self.setWindowTitle("🔔 Bildirimler"); self.setMinimumSize(520,540)
        self.setStyleSheet(ThemeManager.get_stylesheet()); self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(12)
        # Başlık + Tümünü oku
        tb=QHBoxLayout()
        b=QLabel("🔔 Bildirim Merkezi"); f=QFont(); f.setPointSize(14); f.setBold(True); b.setFont(f); tb.addWidget(b)
        tb.addStretch()
        btn_oku=AdminLuxBtn("✅ Tümünü Oku",COLORS["success"]); btn_oku.clicked.connect(self._tumu_oku); tb.addWidget(btn_oku)
        lay.addLayout(tb)
        # Filtre
        tf=QHBoxLayout()
        self.cmb_filtre=QComboBox()
        for v,l in [("hepsi","Tümü"),("okunmamis","Okunmamış")]:
            self.cmb_filtre.addItem(l,userData=v)
        self.cmb_filtre.currentIndexChanged.connect(self._listele); tf.addWidget(QLabel("Göster:")); tf.addWidget(self.cmb_filtre); tf.addStretch()
        # Yeni bildirim ekle
        btn_yeni=AdminLuxBtn("➕ Bildirim Ekle",COLORS["primary"]); btn_yeni.clicked.connect(self._yeni_ekle); tf.addWidget(btn_yeni)
        lay.addLayout(tf)
        # Liste
        self.scroll=QScrollArea(); self.scroll.setWidgetResizable(True)
        self.liste_w=QWidget(); self.liste_lay=QVBoxLayout(self.liste_w); self.liste_lay.setContentsMargins(0,0,0,0); self.liste_lay.setSpacing(8)
        self.scroll.setWidget(self.liste_w); lay.addWidget(self.scroll)
        btn_kapat=AdminLuxBtn("Kapat",COLORS["primary"]); btn_kapat.clicked.connect(self.accept); lay.addWidget(btn_kapat,alignment=Qt.AlignRight)
        self._listele()
    def _listele(self):
        while self.liste_lay.count():
            c=self.liste_lay.takeAt(0)
            if c.widget(): c.widget().deleteLater()
        filtre=self.cmb_filtre.currentData()
        bildirimler=self.db.bildirimler_getir(okunmamis_only=(filtre=="okunmamis"))
        tur_renk={"bilgi":COLORS["primary"],"basari":COLORS["success"],"uyari":COLORS["warning"],"hata":COLORS["danger"]}
        tur_icon={"bilgi":"ℹ️","basari":"✅","uyari":"⚠️","hata":"❌"}
        if not bildirimler:
            lbl=QLabel("Bildirim yok"); lbl.setAlignment(Qt.AlignCenter)
            lbl.setStyleSheet(f"color:{COLORS['text_sec']};font-size:13px;padding:40px;"); self.liste_lay.addWidget(lbl)
        for b in bildirimler:
            tur=b["tur"] if "tur" in b.keys() else "bilgi"
            renk=tur_renk.get(tur,COLORS["primary"])
            okundu=b["okundu"] if "okundu" in b.keys() else 1
            kart=QFrame()
            kart.setStyleSheet(f"QFrame{{background:{'#1a1a2e' if okundu else '#2d2d3a'};border-left:4px solid {renk};border-radius:8px;border-top:1px solid {COLORS['border']};border-right:1px solid {COLORS['border']};border-bottom:1px solid {COLORS['border']};}}")
            kl=QHBoxLayout(kart); kl.setContentsMargins(14,10,14,10); kl.setSpacing(12)
            icon_lbl=QLabel(tur_icon.get(tur,"ℹ️")); icon_lbl.setStyleSheet("font-size:20px;background:transparent;border:none;"); kl.addWidget(icon_lbl)
            v=QVBoxLayout(); v.setSpacing(2)
            baslik_lbl=QLabel(b["baslik"] if "baslik" in b.keys() else ""); baslik_lbl.setStyleSheet(f"font-weight:bold;font-size:12px;color:{'#f4f4f5' if not okundu else COLORS['text_sec']};background:transparent;border:none;"); v.addWidget(baslik_lbl)
            mesaj_lbl=QLabel(b["mesaj"] if "mesaj" in b.keys() else ""); mesaj_lbl.setStyleSheet(f"font-size:11px;color:{COLORS['text_sec']};background:transparent;border:none;"); v.addWidget(mesaj_lbl)
            tarih_lbl=QLabel(str(b["tarih"] if "tarih" in b.keys() else "")[:16]); tarih_lbl.setStyleSheet(f"font-size:10px;color:{COLORS['border']};background:transparent;border:none;"); v.addWidget(tarih_lbl)
            kl.addLayout(v); kl.addStretch()
            if not okundu:
                bid=b["id"] if "id" in b.keys() else None
                btn_oku=QPushButton("✓"); btn_oku.setFixedSize(28,28)
                btn_oku.setStyleSheet(f"QPushButton{{background:{COLORS['success']};color:white;border:none;border-radius:14px;font-weight:bold;}}QPushButton:hover{{background:{COLORS['success_h']};}}")
                btn_oku.clicked.connect(lambda ch=False,i=bid: (self.db.bildirim_oku(i),self._listele()))
                kl.addWidget(btn_oku)
            self.liste_lay.addWidget(kart)
        self.liste_lay.addStretch()
    def _tumu_oku(self):
        self.db.bildirim_oku(); self._listele()
    def _yeni_ekle(self):
        dlg=QDialog(self); dlg.setWindowTitle("Bildirim Ekle"); dlg.setFixedSize(400,300)
        dlg.setStyleSheet(ThemeManager.get_stylesheet())
        dl=QVBoxLayout(dlg); dl.setContentsMargins(20,20,20,20); dl.setSpacing(12)
        frm=QFormLayout(); frm.setSpacing(10)
        i_b=QLineEdit(); i_b.setPlaceholderText("Bildirim başlığı")
        i_m=QTextEdit(); i_m.setPlaceholderText("Bildirim mesajı"); i_m.setMaximumHeight(80)
        c_t=QComboBox()
        for v,l in [("bilgi","ℹ️ Bilgi"),("basari","✅ Başarı"),("uyari","⚠️ Uyarı"),("hata","❌ Hata")]:
            c_t.addItem(l,userData=v)
        frm.addRow("Başlık *:",i_b); frm.addRow("Mesaj:",i_m); frm.addRow("Tür:",c_t); dl.addLayout(frm)
        btn=AdminLuxBtn("📢 Gönder",COLORS["primary"]); dl.addWidget(btn)
        def _gonder():
            if not i_b.text().strip(): return msg_warn(dlg,"Uyarı","Başlık boş olamaz!")
            self.db.bildirim_ekle(i_b.text().strip(),i_m.toPlainText().strip(),c_t.currentData())
            dlg.accept(); self._listele()
        btn.clicked.connect(_gonder)
        dlg.exec_()


# ═══════════════════════════════════════════════════════════════════════════════
# 🔍 GLOBAL ARAMA DIALOG
# ═══════════════════════════════════════════════════════════════════════════════

class GlobalAramaDialog(QDialog):
    def __init__(self,db,parent=None):
        super().__init__(parent); self.db=db
        self.setWindowTitle("🔍 Global Arama"); self.setMinimumSize(620,520)
        self.setStyleSheet(ThemeManager.get_stylesheet()); self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(20,20,20,20); lay.setSpacing(12)
        b=QLabel("🔍 Platform Geneli Arama"); f=QFont(); f.setPointSize(14); f.setBold(True); b.setFont(f); lay.addWidget(b)
        # Arama kutusu
        ara_row=QHBoxLayout()
        self.inp_ara=QLineEdit(); self.inp_ara.setPlaceholderText("Öğrenci adı, kurs kodu, eğitmen...")
        self.inp_ara.setMinimumHeight(40)
        self.inp_ara.setStyleSheet(f"background:{COLORS['bg_surface']};color:{COLORS['text_main']};border:2px solid {COLORS['primary']};border-radius:8px;padding:8px 14px;font-size:13px;")
        self.inp_ara.textChanged.connect(self._ara); ara_row.addWidget(self.inp_ara)
        btn_temizle=AdminLuxBtn("✕",COLORS["danger"]); btn_temizle.setFixedSize(40,40)
        btn_temizle.clicked.connect(lambda: self.inp_ara.clear()); ara_row.addWidget(btn_temizle)
        lay.addLayout(ara_row)
        # Sonuç alanı (tabs)
        self.res_tabs=QTabWidget(); lay.addWidget(self.res_tabs,stretch=1)
        # Öğrenci sonuçları
        self.ogr_liste=QListWidget(); self.res_tabs.addTab(self.ogr_liste,"👥 Öğrenciler (0)")
        # Kurs sonuçları
        self.kurs_liste=QListWidget(); self.res_tabs.addTab(self.kurs_liste,"📚 Kurslar (0)")
        # Eğitmen sonuçları
        self.eg_liste=QListWidget(); self.res_tabs.addTab(self.eg_liste,"👨‍🏫 Eğitmenler (0)")
        # Boş durum
        self.lbl_bos=QLabel("Aramak için yazmaya başlayın..."); self.lbl_bos.setAlignment(Qt.AlignCenter)
        self.lbl_bos.setStyleSheet(f"color:{COLORS['text_sec']};font-size:13px;padding:40px;"); lay.addWidget(self.lbl_bos)
        btn_kapat=AdminLuxBtn("Kapat",COLORS["primary"]); btn_kapat.clicked.connect(self.accept); lay.addWidget(btn_kapat,alignment=Qt.AlignRight)
    def _ara(self,kelime):
        self.lbl_bos.setVisible(len(kelime)<2)
        self.res_tabs.setVisible(len(kelime)>=2)
        if len(kelime)<2:
            for l in [self.ogr_liste,self.kurs_liste,self.eg_liste]: l.clear()
            return
        sonuclar=self.db.global_ara(kelime)
        # Öğrenciler
        self.ogr_liste.clear()
        for o in sonuclar.get("ogrenciler",[]):
            meta=TIER_META.get(o["paket"],{})
            item=QListWidgetItem(f"  {meta.get('icon','')} {o['ad']} {o['soyad'] or ''}   ·   {o['paket'].upper()}   ·   {o['xp']} XP")
            item.setForeground(QColor(meta.get("renk",COLORS["primary"]))); self.ogr_liste.addItem(item)
        self.res_tabs.setTabText(0,f"👥 Öğrenciler ({len(sonuclar.get('ogrenciler',[]))})")
        # Kurslar
        self.kurs_liste.clear()
        for k in sonuclar.get("kurslar",[]):
            meta=TIER_META.get(k["tier"],{})
            item=QListWidgetItem(f"  {meta.get('icon','')} [{k['kurs_kodu']}] {k['ad']}   ·   {k['tier'].upper()}")
            item.setForeground(QColor(meta.get("renk",COLORS["primary"]))); self.kurs_liste.addItem(item)
        self.res_tabs.setTabText(1,f"📚 Kurslar ({len(sonuclar.get('kurslar',[]))})")
        # Eğitmenler
        self.eg_liste.clear()
        for e in sonuclar.get("egitmenler",[]):
            item=QListWidgetItem(f"  👨‍🏫 {e['ad']}   ·   {e['uzmanlik'] or '—'}")
            item.setForeground(QColor(COLORS["secondary"])); self.eg_liste.addItem(item)
        self.res_tabs.setTabText(2,f"👨‍🏫 Eğitmenler ({len(sonuclar.get('egitmenler',[]))})")


# ═══════════════════════════════════════════════════════════════════════════════
# 📖 KURS DETAY DIALOG (Tier5)
# ═══════════════════════════════════════════════════════════════════════════════

class AdminKursDetay(QDialog):
    def __init__(self,db,kurs_id,parent=None):
        super().__init__(parent); self.db=db; self.kurs_id=kurs_id
        self.setWindowTitle("📖 Kurs Detayı"); self.setMinimumSize(640,560)
        self.setStyleSheet(ThemeManager.get_stylesheet()); self._build()
    def _build(self):
        veri=self.db.kurs_detay_getir(self.kurs_id)
        if not veri or not veri["kurs"]: return
        k=veri["kurs"]
        lay=QVBoxLayout(self); lay.setContentsMargins(24,24,24,24); lay.setSpacing(14)
        meta=TIER_META.get(k["tier"],{})
        # Başlık
        tb=QHBoxLayout()
        lbl_tier=QLabel(f"{meta.get('icon','')} {meta.get('label','')}"); lbl_tier.setStyleSheet(f"background:{meta.get('renk',COLORS['primary'])};color:white;font-weight:bold;border-radius:6px;padding:4px 10px;font-size:11px;")
        tb.addWidget(lbl_tier)
        lbl_kod=QLabel(f"[{k['kurs_kodu']}]"); lbl_kod.setStyleSheet(f"color:{COLORS['text_sec']};font-size:12px;background:transparent;border:none;"); tb.addWidget(lbl_kod)
        tb.addStretch(); lay.addLayout(tb)
        lbl_ad=QLabel(k["ad"]); f=QFont(); f.setPointSize(16); f.setBold(True); lbl_ad.setFont(f); lay.addWidget(lbl_ad)
        # Açıklama
        if k["aciklama"]:
            lbl_ac=QLabel(k["aciklama"]); lbl_ac.setWordWrap(True)
            lbl_ac.setStyleSheet(f"color:{COLORS['text_sec']};font-size:12px;background:transparent;border:none;"); lay.addWidget(lbl_ac)
        # İstatistik kartları
        stat_row=QHBoxLayout(); stat_row.setSpacing(10)
        for ikon,lbl_txt,val,renk in [
            ("👥","Kayıtlı",str(veri["kayit_sayisi"]),COLORS["primary"]),
            ("✅","Tamamlayan",str(veri["tamam_sayisi"]),COLORS["success"]),
            ("📊","Ort. İlerleme",f"%{veri['ort_ilerleme']}",COLORS["warning"]),
            ("⚡","Toplam XP",str(veri["toplam_xp"]),COLORS["xp_color"]),
            ("🛒","Enhancement",str(veri["enh_sayisi"]),COLORS["accent"]),
        ]:
            kutu=QFrame(); kutu.setStyleSheet(f"QFrame{{background:{COLORS['bg_surface']};border-radius:10px;border:1px solid {COLORS['border']};min-width:100px;}}")
            kl=QVBoxLayout(kutu); kl.setContentsMargins(10,10,10,10); kl.setAlignment(Qt.AlignCenter)
            kl.addWidget(self._lbl(ikon,"font-size:22px;")); kl.addWidget(self._lbl(val,f"font-size:18px;font-weight:bold;color:{renk};")); kl.addWidget(self._lbl(lbl_txt,f"font-size:10px;color:{COLORS['text_sec']};"))
            stat_row.addWidget(kutu)
        lay.addLayout(stat_row)
        # Eğitmen bilgisi
        eg_frame=QFrame(); eg_frame.setStyleSheet(f"QFrame{{background:{COLORS['bg_card']};border-radius:10px;border:1px solid {COLORS['border']};}}")
        eg_lay=QHBoxLayout(eg_frame); eg_lay.setContentsMargins(16,12,16,12)
        eg_lay.addWidget(self._lbl("👨‍🏫","font-size:26px;"))
        ev=QVBoxLayout(); ev.setSpacing(2)
        ev.addWidget(self._lbl(k["eg_ad"] or "—",f"font-weight:bold;font-size:13px;color:{COLORS['text_main']};"))
        ev.addWidget(self._lbl(k["uzmanlik"] or "",f"font-size:11px;color:{COLORS['text_sec']};"))
        eg_lay.addLayout(ev); eg_lay.addStretch()
        eg_lay.addWidget(self._lbl(f"⭐ {k['rating']:.1f}",f"color:{COLORS['gold']};font-weight:bold;font-size:14px;"))
        lay.addWidget(eg_frame)
        # İlerleme barı (kontenjan doluluk)
        if veri["kayit_sayisi"]>0:
            dol_pct=min(100,int(veri["kayit_sayisi"]/(k["kontenjan"] or 1)*100))
            pb=QProgressBar(); pb.setRange(0,100); pb.setValue(dol_pct)
            pb.setFormat(f"Kontenjan: {veri['kayit_sayisi']}/{k['kontenjan']}  (%{dol_pct})")
            pb.setStyleSheet(f"QProgressBar{{background:{COLORS['bg_surface']};border:none;border-radius:6px;height:18px;color:white;font-size:11px;text-align:center;}}QProgressBar::chunk{{background:{COLORS['primary']};border-radius:6px;}}"); lay.addWidget(pb)
        lay.addStretch()
        btn_kapat=AdminLuxBtn("Kapat",COLORS["primary"]); btn_kapat.clicked.connect(self.accept); lay.addWidget(btn_kapat,alignment=Qt.AlignRight)
    def _lbl(self,txt,st=""):
        l=QLabel(txt); l.setStyleSheet(st+"background:transparent;border:none;"); return l


# ═══════════════════════════════════════════════════════════════════════════════
# ⚙️ AYARLAR SEKMESİ
# ═══════════════════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════════════
# 🎓 ÖĞRENCİ GÖRÜNÜMÜ SEKMESİ — Admin paneli içinde öğrenci deneyimini göster
# ═══════════════════════════════════════════════════════════════════════════════

class OgrenciGorünumSekmesi(QWidget):
    """
    Admin paneli içinde öğrenci deneyimini ve platform özelliklerini
    görsel olarak sunan bilgi/tanıtım sayfası.
    """
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db; self._build()

    def _build(self):
        # Ana scroll area
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea{border:none;background:transparent;}")
        ana_lay = QVBoxLayout(self); ana_lay.setContentsMargins(0,0,0,0)
        ana_lay.addWidget(scroll)

        ctn = QWidget()
        lay = QVBoxLayout(ctn); lay.setContentsMargins(32,28,32,32); lay.setSpacing(24)

        # ── BAŞLIK ────────────────────────────────────────────────────────────
        hero = QFrame()
        hero.setStyleSheet(f"""QFrame{{
            background:qlineargradient(x1:0,y1:0,x2:1,y2:1,
                stop:0 #0d1117,stop:0.5 #161b22,stop:1 #0d1117);
            border-radius:20px;
            border:1px solid {COLORS['primary']};
        }}""")
        hl = QVBoxLayout(hero); hl.setContentsMargins(40,36,40,36); hl.setSpacing(12)
        hl.setAlignment(Qt.AlignCenter)

        hero_icon = QLabel("🎓"); hero_icon.setAlignment(Qt.AlignCenter)
        hero_icon.setStyleSheet("font-size:52px;background:transparent;border:none;")
        hl.addWidget(hero_icon)

        hero_title = QLabel("Öğrenci Deneyimi")
        f1=QFont("Segoe UI",26,QFont.Bold); hero_title.setFont(f1)
        hero_title.setAlignment(Qt.AlignCenter)
        hero_title.setStyleSheet(f"color:{COLORS['accent']};background:transparent;border:none;letter-spacing:1px;")
        hl.addWidget(hero_title)

        hero_sub = QLabel("Online Kurs Platformu'nda öğrenciler neler yapabilir?")
        hero_sub.setAlignment(Qt.AlignCenter)
        hero_sub.setStyleSheet(f"color:{COLORS['text_sec']};font-size:14px;background:transparent;border:none;")
        hl.addWidget(hero_sub)

        # Anlık istatistikler hero'da
        stats = self.db.dashboard_istatistik()
        stat_row = QHBoxLayout(); stat_row.setSpacing(20); stat_row.setAlignment(Qt.AlignCenter)
        for val, lbl, renk in [
            (stats["toplam_ogrenci"], "Aktif Öğrenci", COLORS["primary"]),
            (stats["toplam_kurs"],    "Kurs",          COLORS["accent"]),
            (stats["tamamlanan"],     "Tamamlanan",    COLORS["success"]),
            (stats["toplam_badge"],   "Badge",         COLORS["gold"]),
        ]:
            kutu = QFrame()
            kutu.setStyleSheet(f"""QFrame{{background:rgba(255,255,255,0.05);
                border-radius:12px;border:1px solid rgba(255,255,255,0.08);}}""")
            kl = QVBoxLayout(kutu); kl.setContentsMargins(20,14,20,14); kl.setAlignment(Qt.AlignCenter)
            v = QLabel(str(val)); vf=QFont("Segoe UI",22,QFont.Bold); v.setFont(vf)
            v.setAlignment(Qt.AlignCenter)
            v.setStyleSheet(f"color:{renk};background:transparent;border:none;")
            l = QLabel(lbl); l.setAlignment(Qt.AlignCenter)
            l.setStyleSheet(f"color:{COLORS['text_sec']};font-size:11px;background:transparent;border:none;")
            kl.addWidget(v); kl.addWidget(l)
            stat_row.addWidget(kutu)
        hl.addSpacing(8); hl.addLayout(stat_row)
        lay.addWidget(hero)

        # ── TİER PAKETLERİ ────────────────────────────────────────────────────
        lay.addWidget(self._bolum_baslik("📦 Öğrenci Paket Seçenekleri",
            "Her öğrenci ihtiyacına göre paket seçebilir"))

        tier_row = QHBoxLayout(); tier_row.setSpacing(16)
        tier_ozellikler = {
            "free": {
                "icon": "🆓", "ad": "FREE",
                "renk": COLORS["tier_free"],
                "fiyat": "Ücretsiz",
                "ozellikler": [
                    "✅ Tüm ücretsiz kurslara erişim",
                    "✅ Video izleme (+10 XP)",
                    "✅ QA Forum (+30 ₺ ek)",
                    "✅ Ekstra kaynaklar (+20 ₺ ek)",
                    "✅ XP & Level sistemi",
                    "✅ Badge koleksiyonu",
                    "✅ Kurs ilerleme takibi",
                    "❌ Pro kurslar",
                    "❌ Mentor atama",
                    "❌ Sertifika",
                ]
            },
            "pro2": {
                "icon": "⭐", "ad": "PRO",
                "renk": COLORS["tier_pro2"],
                "fiyat": "₺499 / ay",
                "ozellikler": [
                    "✅ FREE'nin tüm özellikleri",
                    "✅ Pro kurslar dahil",
                    "✅ Proje ödevi (+80 ₺ ek)",
                    "✅ Birebir feedback (+60 ₺ ek)",
                    "✅ Gelişim analitik (+40 ₺ ek)",
                    "✅ Quiz tamamlama (+25 XP)",
                    "✅ Proje gönderme (+50 XP)",
                    "✅ Leaderboard'da üst sıra",
                    "❌ Premium kurslar",
                    "❌ 1-on-1 mentoring",
                ]
            },
            "pro3": {
                "icon": "👑", "ad": "PREMIUM",
                "renk": COLORS["tier_pro3"],
                "fiyat": "₺999 / ay",
                "ozellikler": [
                    "✅ PRO'nun tüm özellikleri",
                    "✅ Premium kurslar dahil",
                    "✅ 1-on-1 Mentoring (+200 ₺ ek)",
                    "✅ Sertifika talebi (+100 ₺ ek)",
                    "✅ Kurs tamamlama (+100 XP)",
                    "✅ Öncelikli eğitmen erişimi",
                    "✅ PDF sertifika çıktısı",
                    "✅ Tüm Enhancement'lar",
                    "✅ Mentor ilişkisi yönetimi",
                    "✅ Tüm avantajlar aktif",
                ]
            }
        }

        for tier_kodu, meta in tier_ozellikler.items():
            kart = QFrame()
            is_premium = tier_kodu == "pro3"
            kart.setStyleSheet(f"""QFrame{{
                background:qlineargradient(x1:0,y1:0,x2:0,y2:1,
                    stop:0 {COLORS['bg_card']},stop:1 {COLORS['bg_surface']});
                border-radius:18px;
                border:{'2px' if is_premium else '1px'} solid {meta['renk']};
            }}""")
            kl = QVBoxLayout(kart); kl.setContentsMargins(24,24,24,24); kl.setSpacing(14)

            # Rozet (premium'da özel)
            if is_premium:
                badge = QLabel("🔥 En Popüler"); badge.setAlignment(Qt.AlignCenter)
                badge.setStyleSheet(f"background:{meta['renk']};color:white;border-radius:10px;padding:4px 12px;font-size:11px;font-weight:bold;border:none;")
                kl.addWidget(badge)

            # İkon + ad
            ikon_lbl = QLabel(meta["icon"]); ikon_lbl.setAlignment(Qt.AlignCenter)
            ikon_lbl.setStyleSheet(f"font-size:36px;background:transparent;border:none;")
            kl.addWidget(ikon_lbl)

            ad_lbl = QLabel(meta["ad"])
            af = QFont("Segoe UI",18,QFont.Bold); ad_lbl.setFont(af)
            ad_lbl.setAlignment(Qt.AlignCenter)
            ad_lbl.setStyleSheet(f"color:{meta['renk']};background:transparent;border:none;")
            kl.addWidget(ad_lbl)

            fiyat_lbl = QLabel(meta["fiyat"])
            ff = QFont("Segoe UI",14,QFont.Bold); fiyat_lbl.setFont(ff)
            fiyat_lbl.setAlignment(Qt.AlignCenter)
            fiyat_lbl.setStyleSheet(f"color:{COLORS['text_main']};background:transparent;border:none;")
            kl.addWidget(fiyat_lbl)

            # Ayraç
            sep = QFrame(); sep.setFrameShape(QFrame.HLine)
            sep.setStyleSheet(f"background:{meta['renk']};max-height:1px;border:none;margin:4px 0;")
            kl.addWidget(sep)

            # Özellik listesi
            for oz in meta["ozellikler"]:
                aktif = oz.startswith("✅")
                oz_lbl = QLabel(oz)
                oz_lbl.setStyleSheet(f"""color:{'#ffffff' if aktif else COLORS['text_dim']};
                    font-size:12px;background:transparent;border:none;
                    padding:2px 0;""")
                kl.addWidget(oz_lbl)

            tier_row.addWidget(kart)

        lay.addLayout(tier_row)

        # ── XP & GAMİFİCATİON ─────────────────────────────────────────────────
        lay.addWidget(self._bolum_baslik("⚡ XP & Gamification Sistemi",
            "Öğrenciler her aktivitede XP kazanır ve level atlar"))

        xp_grid = QGridLayout(); xp_grid.setSpacing(14)
        xp_aktiviteler = [
            ("🎬","Video İzleme",   "+10 XP",  "Her ders videosunu izleyerek XP kazan",     COLORS["primary"]),
            ("📝","Quiz Çözme",     "+25 XP",  "Quiz'leri tamamla, bilgini test et",         COLORS["accent"]),
            ("📦","Proje Gönder",   "+50 XP",  "Proje ödevini teslim et ve XP kazan",        COLORS["tier_pro2"]),
            ("🎓","Kurs Bitir",     "+100 XP", "Kursu %100 tamamlayınca büyük XP ödülü",     COLORS["success"]),
            ("🛒","Enhancement Al", "+15 XP",  "Kurs için ek özellik satın al",              COLORS["purple"]),
            ("👤","Profil Doldur",  "+20 XP",  "Profil bilgilerini güncelle",                COLORS["teal"]),
            ("🔑","İlk Giriş",     "+5 XP",   "Günlük ilk girişte bonus XP",                COLORS["gold"]),
            ("🏅","Badge Kazan",   "Bonus XP", "Her rozet farklı miktarda bonus XP verir",   COLORS["tier_pro3"]),
        ]
        for i, (ik, ad, xp, acik, renk) in enumerate(xp_aktiviteler):
            kutu = QFrame()
            kutu.setStyleSheet(f"""QFrame{{
                background:{COLORS['bg_card']};border-radius:14px;
                border-left:4px solid {renk};
                border-top:1px solid rgba(255,255,255,0.05);
                border-right:1px solid rgba(255,255,255,0.03);
                border-bottom:1px solid rgba(0,0,0,0.15);
            }}""")
            kl = QHBoxLayout(kutu); kl.setContentsMargins(16,14,16,14); kl.setSpacing(14)

            ik_lbl = QLabel(ik)
            ik_lbl.setFixedSize(44,44); ik_lbl.setAlignment(Qt.AlignCenter)
            ik_lbl.setStyleSheet(f"font-size:22px;background:rgba(255,255,255,0.06);border-radius:10px;border:none;")
            kl.addWidget(ik_lbl)

            txt = QVBoxLayout(); txt.setSpacing(2)
            ad_l = QLabel(ad); af=QFont("Segoe UI",13,QFont.Bold); ad_l.setFont(af)
            ad_l.setStyleSheet(f"color:{COLORS['text_main']};background:transparent;border:none;")
            acik_l = QLabel(acik)
            acik_l.setStyleSheet(f"color:{COLORS['text_sec']};font-size:11px;background:transparent;border:none;")
            txt.addWidget(ad_l); txt.addWidget(acik_l); kl.addLayout(txt); kl.addStretch()

            xp_lbl = QLabel(xp); xf=QFont("Segoe UI",14,QFont.Bold); xp_lbl.setFont(xf)
            xp_lbl.setStyleSheet(f"color:{renk};background:transparent;border:none;")
            kl.addWidget(xp_lbl)

            xp_grid.addWidget(kutu, i//2, i%2)

        lay.addLayout(xp_grid)

        # ── BADGE SİSTEMİ ─────────────────────────────────────────────────────
        lay.addWidget(self._bolum_baslik("🏅 Rozet Sistemi",
            f"Öğrenciler {len(BADGES)} farklı rozet kazanabilir"))

        badge_scroll = QScrollArea(); badge_scroll.setWidgetResizable(True)
        badge_scroll.setFixedHeight(170)
        badge_scroll.setStyleSheet("QScrollArea{border:none;background:transparent;}")
        badge_w = QWidget(); badge_w.setStyleSheet("background:transparent;")
        badge_lay = QHBoxLayout(badge_w); badge_lay.setSpacing(12); badge_lay.setContentsMargins(4,4,4,4)

        # Platform genelindeki badge durumunu çek
        try:
            with self.db.get_connection() as conn:
                kazanim_sayilari = dict(conn.execute(
                    "SELECT badge_kodu,COUNT(*) FROM badge_kayitlar GROUP BY badge_kodu"
                ).fetchall())
        except:
            kazanim_sayilari = {}

        for bk, bmeta in BADGES.items():
            kutu = QFrame()
            kutu.setFixedSize(130, 148)
            kutu.setStyleSheet(f"""QFrame{{
                background:{COLORS['bg_card']};
                border-radius:14px;
                border:1.5px solid {bmeta['renk']};
            }}""")
            kl = QVBoxLayout(kutu); kl.setContentsMargins(10,10,10,10); kl.setSpacing(4); kl.setAlignment(Qt.AlignCenter)

            ik = QLabel(bmeta["icon"]); ik.setAlignment(Qt.AlignCenter)
            ik.setStyleSheet("font-size:28px;background:transparent;border:none;")
            kl.addWidget(ik)

            nm = QLabel(bmeta["ad"]); nm.setAlignment(Qt.AlignCenter); nm.setWordWrap(True)
            nm.setStyleSheet(f"font-size:11px;font-weight:bold;color:white;background:transparent;border:none;")
            kl.addWidget(nm)

            kazanan = kazanim_sayilari.get(bk, 0)
            kz = QLabel(f"{kazanan} öğrenci"); kz.setAlignment(Qt.AlignCenter)
            kz.setStyleSheet(f"font-size:10px;color:{bmeta['renk']};background:transparent;border:none;")
            kl.addWidget(kz)

            badge_lay.addWidget(kutu)

        badge_lay.addStretch()
        badge_scroll.setWidget(badge_w)
        lay.addWidget(badge_scroll)

        # ── ENHANCEMENT SİSTEMİ ───────────────────────────────────────────────
        lay.addWidget(self._bolum_baslik("🛒 Enhancement (Yükseltme) Sistemi",
            "Öğrenciler kurslarına ek özellikler satın alabilir"))

        enh_grid = QGridLayout(); enh_grid.setSpacing(12)
        for i, (ek, emeta) in enumerate(ENHANCEMENTS.items()):
            tier_renk = TIER_META.get(emeta["tier"],{}).get("renk", COLORS["primary"])
            kutu = QFrame()
            kutu.setStyleSheet(f"""QFrame{{
                background:{COLORS['bg_card']};border-radius:14px;
                border:1px solid rgba(255,255,255,0.07);
                border-top:3px solid {tier_renk};
            }}""")
            kl = QHBoxLayout(kutu); kl.setContentsMargins(16,14,16,14); kl.setSpacing(12)

            ik = QLabel(emeta["icon"]); ik.setFixedSize(40,40); ik.setAlignment(Qt.AlignCenter)
            ik.setStyleSheet(f"font-size:22px;background:{tier_renk}22;border-radius:8px;border:none;")
            kl.addWidget(ik)

            v = QVBoxLayout(); v.setSpacing(2)
            nm = QLabel(emeta["ad"]); nf=QFont("Segoe UI",12,QFont.Bold); nm.setFont(nf)
            nm.setStyleSheet(f"color:{COLORS['text_main']};background:transparent;border:none;")
            tr = QLabel(f"{TIER_META.get(emeta['tier'],{}).get('label','')} ve üzeri")
            tr.setStyleSheet(f"color:{tier_renk};font-size:10px;background:transparent;border:none;")
            v.addWidget(nm); v.addWidget(tr); kl.addLayout(v); kl.addStretch()

            fiyat = QLabel(f"₺{emeta['fiyat']}"); ff=QFont("Segoe UI",14,QFont.Bold); fiyat.setFont(ff)
            fiyat.setStyleSheet(f"color:{COLORS['gold']};background:transparent;border:none;")
            kl.addWidget(fiyat)

            enh_grid.addWidget(kutu, i//2, i%2)

        lay.addLayout(enh_grid)

        # ── ÖĞRENCI YOLCULUĞU ─────────────────────────────────────────────────
        lay.addWidget(self._bolum_baslik("🗺️ Öğrenci Yolculuğu",
            "Platformdaki tipik öğrenci deneyimi adım adım"))

        adimlar = [
            ("1","🔑","Kayıt & Giriş",
             "Öğrenci platforma kayıt olur. İlk girişte +5 XP kazanır ve profili oluşturulur.",
             COLORS["primary"]),
            ("2","📚","Kurs Seçimi",
             "Katalogdan FREE, PRO veya PREMIUM kursları inceler. Tier'ına uygun kurslara kaydolur.",
             COLORS["accent"]),
            ("3","📖","Öğrenme Süreci",
             "Video izler (+10 XP), quiz çözer (+25 XP), proje gönderir (+50 XP). Her adımda XP birikir.",
             COLORS["tier_pro2"]),
            ("4","🛒","Enhancement Alımı",
             "İsterse canlı chat, QA forum, birebir feedback gibi ek özellikler satın alır.",
             COLORS["purple"]),
            ("5","🧑‍🏫","Mentor Desteği",
             "Premium öğrenciler eğitmen ile birebir mentorluk ilişkisi kurabilir.",
             COLORS["teal"]),
            ("6","🎓","Sertifika",
             "Kursu %100 tamamlayan öğrenci sertifika talep eder. Eğitmen onaylar, PDF çıktı alınır.",
             COLORS["gold"]),
            ("7","🏆","Leaderboard",
             "En fazla XP kazanan öğrenciler leaderboard'da öne çıkar. Rekabet motivasyonu sağlar.",
             COLORS["tier_pro3"]),
        ]

        for num, ik, ad, acik, renk in adimlar:
            kart = QFrame()
            kart.setStyleSheet(f"""QFrame{{
                background:{COLORS['bg_card']};border-radius:14px;
                border-left:4px solid {renk};
                border-top:1px solid rgba(255,255,255,0.05);
                border-right:none;border-bottom:none;
            }}""")
            kl = QHBoxLayout(kart); kl.setContentsMargins(20,16,20,16); kl.setSpacing(18)

            # Numara
            num_lbl = QLabel(num)
            num_lbl.setFixedSize(40,40); num_lbl.setAlignment(Qt.AlignCenter)
            num_lbl.setStyleSheet(f"""color:white;font-size:16px;font-weight:bold;
                background:{renk};border-radius:20px;border:none;""")
            kl.addWidget(num_lbl)

            # İkon
            ik_lbl = QLabel(ik)
            ik_lbl.setStyleSheet("font-size:24px;background:transparent;border:none;")
            kl.addWidget(ik_lbl)

            # Metin
            txt = QVBoxLayout(); txt.setSpacing(3)
            ad_l = QLabel(ad); af=QFont("Segoe UI",13,QFont.Bold); ad_l.setFont(af)
            ad_l.setStyleSheet(f"color:{renk};background:transparent;border:none;")
            acik_l = QLabel(acik); acik_l.setWordWrap(True)
            acik_l.setStyleSheet(f"color:{COLORS['text_sec']};font-size:12px;background:transparent;border:none;")
            txt.addWidget(ad_l); txt.addWidget(acik_l)
            kl.addLayout(txt)

            lay.addWidget(kart)

        lay.addSpacing(20)
        scroll.setWidget(ctn)

    def _bolum_baslik(self, baslik, alt=""):
        w = QWidget()
        vl = QVBoxLayout(w); vl.setContentsMargins(0,8,0,4); vl.setSpacing(4)
        b = QLabel(baslik); bf=QFont("Segoe UI",16,QFont.Bold); b.setFont(bf)
        b.setStyleSheet(f"color:{COLORS['text_main']};background:transparent;border:none;")
        vl.addWidget(b)
        if alt:
            s = QLabel(alt)
            s.setStyleSheet(f"color:{COLORS['text_sec']};font-size:12px;background:transparent;border:none;")
            vl.addWidget(s)
        # Alt çizgi
        line = QFrame(); line.setFrameShape(QFrame.HLine); line.setFixedHeight(1)
        line.setStyleSheet(f"background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 {COLORS['accent']},stop:1 transparent);border:none;")
        vl.addWidget(line)
        return w

class AyarlarSekmesi(QWidget):
    def __init__(self,db):
        super().__init__(); self.db=db; self._build()
    def _build(self):
        lay=QVBoxLayout(self); lay.setContentsMargins(24,24,24,24); lay.setSpacing(20)
        lbl=QLabel("⚙️ Platform Ayarları"); f=QFont(); f.setPointSize(14); f.setBold(True); lbl.setFont(f); lay.addWidget(lbl)
        # Platform bilgileri
        g1=QGroupBox("🏫 Platform Bilgileri"); g1l=QFormLayout(g1); g1l.setSpacing(12)
        self.i_platform_adi=QLineEdit(self.db.ayar_getir("platform_adi","Online Kurs Platformu"))
        self.i_platform_slogan=QLineEdit(self.db.ayar_getir("platform_slogan","Öğrenmek İçin En İyi Yer"))
        self.i_iletisim_email=QLineEdit(self.db.ayar_getir("iletisim_email","admin@example.com"))
        g1l.addRow("Platform Adı:",self.i_platform_adi); g1l.addRow("Slogan:",self.i_platform_slogan); g1l.addRow("İletişim E-posta:",self.i_iletisim_email)
        lay.addWidget(g1)
        # Paket fiyatları
        g2=QGroupBox("💰 Paket Fiyatları"); g2l=QFormLayout(g2); g2l.setSpacing(12)
        self.i_pro2_fiyat=QDoubleSpinBox(); self.i_pro2_fiyat.setRange(0,9999); self.i_pro2_fiyat.setSuffix(" ₺"); self.i_pro2_fiyat.setValue(float(self.db.ayar_getir("pro2_fiyat","499")))
        self.i_pro3_fiyat=QDoubleSpinBox(); self.i_pro3_fiyat.setRange(0,9999); self.i_pro3_fiyat.setSuffix(" ₺"); self.i_pro3_fiyat.setValue(float(self.db.ayar_getir("pro3_fiyat","999")))
        g2l.addRow("PRO Fiyatı:",self.i_pro2_fiyat); g2l.addRow("PREMIUM Fiyatı:",self.i_pro3_fiyat)
        lay.addWidget(g2)
        # Sistem ayarları
        g3=QGroupBox("🔧 Sistem Ayarları"); g3l=QFormLayout(g3); g3l.setSpacing(12)
        self.cb_bildirim=QCheckBox(); self.cb_bildirim.setChecked(self.db.ayar_getir("bildirimler_aktif","1")=="1")
        self.cb_xp=QCheckBox(); self.cb_xp.setChecked(self.db.ayar_getir("xp_aktif","1")=="1")
        self.i_max_kayit=QSpinBox(); self.i_max_kayit.setRange(1,100); self.i_max_kayit.setValue(int(self.db.ayar_getir("maks_kayit_sayisi","10")))
        g3l.addRow("Bildirimleri Etkinleştir:",self.cb_bildirim); g3l.addRow("XP Sistemini Etkinleştir:",self.cb_xp); g3l.addRow("Maks. Kayıt/Öğrenci:",self.i_max_kayit)
        lay.addWidget(g3)
        # Kaydet butonu
        btn_sav=AdminLuxBtn("💾 Ayarları Kaydet",COLORS["success"]); btn_sav.setMinimumHeight(44); btn_sav.clicked.connect(self._kaydet)
        lay.addWidget(btn_sav)
        # Veritabanı işlemleri
        g4=QGroupBox("🗄️ Veritabanı"); g4l=QHBoxLayout(g4); g4l.setSpacing(12)
        btn_yedek=AdminLuxBtn("📦 Yedek Al",COLORS["primary"]); btn_yedek.clicked.connect(self._yedek); g4l.addWidget(btn_yedek)
        btn_istatistik=AdminLuxBtn("📊 DB İstatistik",COLORS["accent"]); btn_istatistik.clicked.connect(self._db_istatistik); g4l.addWidget(btn_istatistik)
        g4l.addStretch(); lay.addWidget(g4)
        lay.addStretch()
    def _kaydet(self):
        self.db.ayar_kaydet("platform_adi",self.i_platform_adi.text().strip())
        self.db.ayar_kaydet("platform_slogan",self.i_platform_slogan.text().strip())
        self.db.ayar_kaydet("iletisim_email",self.i_iletisim_email.text().strip())
        self.db.ayar_kaydet("pro2_fiyat",str(self.i_pro2_fiyat.value()))
        self.db.ayar_kaydet("pro3_fiyat",str(self.i_pro3_fiyat.value()))
        self.db.ayar_kaydet("bildirimler_aktif","1" if self.cb_bildirim.isChecked() else "0")
        self.db.ayar_kaydet("xp_aktif","1" if self.cb_xp.isChecked() else "0")
        self.db.ayar_kaydet("maks_kayit_sayisi",str(self.i_max_kayit.value()))
        self.db.bildirim_ekle("Ayarlar Kaydedildi","Platform ayarları başarıyla güncellendi.","basari")
        msg_info(self,"Başarılı","Ayarlar kaydedildi!")
    def _yedek(self):
        import shutil
        dosya=f"yedek_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
        try:
            shutil.copy2(self.db.db_name,dosya)
            self.db.bildirim_ekle("Yedek Alındı",f"Veritabanı yedeği: {dosya}","basari")
            msg_info(self,"Başarılı",f"Yedek alındı:\n{dosya}")
        except Exception as e:
            msg_error(self,"Hata",str(e))
    def _db_istatistik(self):
        with self.db.get_connection() as conn:
            c=conn.cursor()
            tablolar=c.execute("SELECT name FROM sqlite_master WHERE type='table' ORDER BY name").fetchall()
            bilgi="Tablo Adı         Kayıt\n"+"─"*30+"\n"
            for t in tablolar:
                try:
                    n=c.execute(f"SELECT COUNT(*) FROM {t[0]}").fetchone()[0]
                    bilgi+=f"{t[0]:<22}{n}\n"
                except: pass
        msg_info(self,"Veritabanı İstatistikleri",bilgi)


# ═══════════════════════════════════════════════════════════════════════════════
# 🎨 TEMA YÖNETİCİSİ (canlı tema değiştirme)
# ═══════════════════════════════════════════════════════════════════════════════

class TemaYoneticisi:
    _aktif_tema = "Okyanus Karanlığı"

    @classmethod
    def tema_uygula(cls, tema_adi, app):
        if tema_adi not in TEMALAR: return
        cls._aktif_tema = tema_adi
        t = TEMALAR[tema_adi]
        for key, val in t.items():
            if key in COLORS:
                COLORS[key] = val
        # Stylesheet yeniden uygula
        app.setStyleSheet(ThemeManager.get_stylesheet())

    @classmethod
    def aktif_tema(cls):
        return cls._aktif_tema


# ═══════════════════════════════════════════════════════════════════════════════
# 💫 SPLASH SCREEN
# ═══════════════════════════════════════════════════════════════════════════════

class AdminSplash(QDialog):
    def __init__(self):
        super().__init__()
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        self.setFixedSize(480, 300)
        self.setStyleSheet(f"""QDialog{{
            background:qlineargradient(x1:0,y1:0,x2:1,y2:1,
                stop:0 {COLORS['bg_dark']},stop:1 {COLORS['bg_card']});
            border:2px solid {COLORS['primary']};border-radius:16px;}}""")
        self._build()
        # Ekran ortası
        from PyQt5.QtWidgets import QDesktopWidget
        geo = QDesktopWidget().screenGeometry()
        self.move((geo.width()-480)//2, (geo.height()-300)//2)

    def _build(self):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(40, 40, 40, 30)
        lay.setSpacing(12)

        lbl_logo = QLabel("🎓")
        lbl_logo.setAlignment(Qt.AlignCenter)
        lbl_logo.setStyleSheet("font-size:52px;background:transparent;border:none;")
        lay.addWidget(lbl_logo)

        lbl_ad = QLabel("Online Kurs Platformu")
        f = QFont(); f.setPointSize(20); f.setBold(True); lbl_ad.setFont(f)
        lbl_ad.setAlignment(Qt.AlignCenter)
        lbl_ad.setStyleSheet(f"color:{COLORS['text_main']};background:transparent;border:none;")
        lay.addWidget(lbl_ad)

        lbl_ver = QLabel("TIER 6  ✨  Final Release")
        lbl_ver.setAlignment(Qt.AlignCenter)
        lbl_ver.setStyleSheet(f"color:{COLORS['gold']};font-size:13px;font-weight:bold;background:transparent;border:none;")
        lay.addWidget(lbl_ver)

        lay.addSpacing(10)

        self.pb = QProgressBar()
        self.pb.setRange(0, 100); self.pb.setValue(0)
        self.pb.setTextVisible(False)
        self.pb.setStyleSheet(f"""QProgressBar{{background:{COLORS['bg_surface']};border:none;
            border-radius:6px;height:8px;}}
            QProgressBar::chunk{{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,
                stop:0 {COLORS['primary']},stop:1 {COLORS['gold']});border-radius:6px;}}""")
        lay.addWidget(self.pb)

        self.lbl_durum = QLabel("Başlatılıyor...")
        self.lbl_durum.setAlignment(Qt.AlignCenter)
        self.lbl_durum.setStyleSheet(f"color:{COLORS['text_sec']};font-size:11px;background:transparent;border:none;")
        lay.addWidget(self.lbl_durum)

        mesajlar = ["Veritabanı hazırlanıyor...","Bileşenler yükleniyor...","Temalar uygulanıyor...","XP sistemi aktif ediliyor...","Hazır! 🚀"]
        self._step = 0; self._msgs = mesajlar
        self._timer = QTimer(); self._timer.timeout.connect(self._adim); self._timer.start(280)

    def _adim(self):
        self._step += 1
        pct = int(self._step / len(self._msgs) * 100)
        self.pb.setValue(pct)
        if self._step <= len(self._msgs):
            self.lbl_durum.setText(self._msgs[self._step-1])
        if self._step >= len(self._msgs):
            self._timer.stop()
            QTimer.singleShot(300, self.accept)


# ═══════════════════════════════════════════════════════════════════════════════
# 👤 ÖĞRENCİ PROFİL KARTI DIALOG
# ═══════════════════════════════════════════════════════════════════════════════

class OgrenciProfilDialog(QDialog):
    def __init__(self, db, ogrenci_id, parent=None):
        super().__init__(parent); self.db = db
        self.setWindowTitle("👤 Öğrenci Profil Kartı")
        self.setMinimumSize(600, 620)
        self.setStyleSheet(ThemeManager.get_stylesheet()); self._build(ogrenci_id)

    def _build(self, ogrenci_id):
        veri = self.db.ogrenci_profil_getir(ogrenci_id)
        if not veri:
            QVBoxLayout(self).addWidget(QLabel("Profil bulunamadı.")); return
        o = veri["ogr"]
        meta = TIER_META.get(o["paket"], {})

        lay = QVBoxLayout(self); lay.setContentsMargins(24, 24, 24, 24); lay.setSpacing(14)

        # Üst kart (avatar + temel bilgi)
        top = QFrame()
        top.setStyleSheet(f"""QFrame{{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,
            stop:0 {COLORS['bg_card']},stop:1 {COLORS['bg_surface']});
            border-radius:14px;border:2px solid {meta.get('renk',COLORS['border'])};}}""")
        tl = QHBoxLayout(top); tl.setContentsMargins(20, 18, 20, 18); tl.setSpacing(20)

        # Avatar dairesi
        av = QLabel(o["ad"][0].upper() if o["ad"] else "?")
        av.setFixedSize(72, 72); av.setAlignment(Qt.AlignCenter)
        av.setStyleSheet(f"""QLabel{{background:{meta.get('renk',COLORS['primary'])};
            color:white;font-size:28px;font-weight:bold;
            border-radius:36px;border:none;}}""")
        tl.addWidget(av)

        # İsim + bilgiler
        v = QVBoxLayout(); v.setSpacing(4)
        lbl_ad = QLabel(f"{o['ad']} {o['soyad'] or ''}")
        fa = QFont(); fa.setPointSize(16); fa.setBold(True); lbl_ad.setFont(fa)
        lbl_ad.setStyleSheet(f"color:{COLORS['text_main']};background:transparent;border:none;")
        v.addWidget(lbl_ad)
        v.addWidget(self._lbl(f"@{o['kullanici_adi']}","color:#a1a1aa;font-size:12px;"))
        v.addWidget(self._lbl(o["email"] or "—","color:#a1a1aa;font-size:11px;"))
        tl.addLayout(v); tl.addStretch()

        # Paket rozeti
        pr = QFrame()
        pr.setStyleSheet(f"QFrame{{background:{meta.get('renk','#0077b6')};border-radius:8px;padding:4px;}}")
        pl = QVBoxLayout(pr); pl.setContentsMargins(12, 8, 12, 8); pl.setSpacing(2)
        pl.addWidget(self._lbl(f"{meta.get('icon','')} {meta.get('label','')}","color:white;font-weight:bold;font-size:13px;"))
        pl.addWidget(self._lbl(f"Lv.{o['level']}  ·  {o['xp']} XP","color:white;font-size:11px;"))
        tl.addWidget(pr)
        lay.addWidget(top)

        # XP progress barı
        lay.addWidget(XpProgressBar(o["xp"]))

        # İstatistik grid (6 kart)
        grid = QGridLayout(); grid.setSpacing(10)
        stats = [
            ("📚","Kayıtlı Kurs",veri["kayit_sayisi"],COLORS["primary"]),
            ("✅","Tamamlanan",veri["tamam_sayisi"],COLORS["success"]),
            ("🏅","Badge Sayısı",veri["badge_sayisi"],COLORS["gold"]),
            ("🛒","Enhancement",veri["enh_sayisi"],COLORS["accent"]),
            ("🧑‍🏫","Aktif Mentor","Var" if veri["mentor_var"] else "Yok",
             COLORS["success"] if veri["mentor_var"] else COLORS["text_sec"]),
            ("📜","Sertifika",veri["sertifika_sayisi"],COLORS["tier_pro3"]),
        ]
        for i,(ik,b,d,r) in enumerate(stats):
            grid.addWidget(KpiKart(ik,b,d,r), i//3, i%3)
        lay.addLayout(grid)

        # Son XP hareketleri
        if veri["son_hareketler"]:
            hgrp = QGroupBox("⚡ Son XP Hareketleri")
            hl = QVBoxLayout(hgrp); hl.setSpacing(4)
            for h in veri["son_hareketler"]:
                row = QHBoxLayout()
                row.addWidget(self._lbl(f"+{h['xp_miktari']} XP",f"color:{COLORS['gold']};font-weight:bold;font-size:12px;min-width:60px;"))
                row.addWidget(self._lbl(h["islem_turu"].replace("_"," ").title(),f"color:{COLORS['text_main']};font-size:11px;"))
                row.addStretch()
                row.addWidget(self._lbl(str(h["tarih"])[:16],f"color:{COLORS['text_sec']};font-size:10px;"))
                frm = QFrame()
                frm.setStyleSheet(f"QFrame{{border-bottom:1px solid {COLORS['border']};background:transparent;}}")
                fl = QVBoxLayout(frm); fl.setContentsMargins(0,2,0,2); fl.addLayout(row)
                hl.addWidget(frm)
            lay.addWidget(hgrp)

        lay.addStretch()
        btn = AdminLuxBtn("Kapat", COLORS["primary"])
        btn.clicked.connect(self.accept); lay.addWidget(btn, alignment=Qt.AlignRight)

    def _lbl(self, txt, st=""):
        l = QLabel(txt); l.setStyleSheet(st+"background:transparent;border:none;")
        return l


# ═══════════════════════════════════════════════════════════════════════════════
# 📊 CANLI DASHBOARD (TIER 6 — otomatik yenileme)
# ═══════════════════════════════════════════════════════════════════════════════

class CanliDashboard(QFrame):
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.setFixedHeight(116)
        self.setStyleSheet(f"""QFrame{{
            background:qlineargradient(x1:0,y1:0,x2:1,y2:0,
                stop:0 {COLORS['bg_card']},stop:0.5 {COLORS['bg_surface']},stop:1 {COLORS['bg_card']});
            border-bottom:1px solid {COLORS['border']};
            border-radius:0;
        }}""")
        lay = QVBoxLayout(self); lay.setContentsMargins(20,8,20,8); lay.setSpacing(6)

        # Üst bar: başlık + zaman
        tb = QHBoxLayout(); tb.setSpacing(0)
        lbl = QLabel("⚡ Bugün — Canlı")
        f = QFont("Segoe UI", 11, QFont.Bold); lbl.setFont(f)
        lbl.setStyleSheet(f"color:{COLORS['accent']};background:transparent;border:none;")
        tb.addWidget(lbl); tb.addStretch()
        self.lbl_son = QLabel("")
        self.lbl_son.setStyleSheet(f"color:{COLORS['text_sec']};font-size:11px;background:transparent;border:none;")
        tb.addWidget(self.lbl_son)
        lay.addLayout(tb)

        # Stat kutular
        self.stat_row = QHBoxLayout(); self.stat_row.setSpacing(10)
        lay.addLayout(self.stat_row)

        self._timer = QTimer()
        self._timer.timeout.connect(self.yenile)
        self._timer.start(10000)
        self.yenile()

    def yenile(self):
        while self.stat_row.count():
            c = self.stat_row.takeAt(0)
            if c.widget(): c.widget().deleteLater()

        st = self.db.platform_canli_istatistik()
        items = [
            ("📝", "Yeni Kayıt",      st["bugun_kayit"],   COLORS["primary"]),
            ("✅", "Tamamlanan",      st["bugun_tamam"],   COLORS["success"]),
            ("⚡", "XP Kazanıldı",   st["bugun_xp"],      COLORS["xp_color"]),
            ("🧑‍🏫", "Aktif Mentor", st["aktif_mentor"],  COLORS["accent"]),
            ("📜", "Bekl. Sertifika", st["bekleyen_sert"], COLORS["warning"]),
            ("🔔", "Bildirim",        st["okunmamis_b"],   COLORS["gold"]),
        ]
        for ik, b, d, r in items:
            kutu = QFrame()
            kutu.setStyleSheet(f"""QFrame{{
                background:{COLORS['bg_card']};
                border-radius:10px;
                border-top:2px solid {r};
                border-left:none;border-right:none;border-bottom:none;
            }}""")
            kutu.setMinimumWidth(140)
            kl = QHBoxLayout(kutu)
            kl.setContentsMargins(14,6,14,6); kl.setSpacing(10)

            # İkon
            li = QLabel(ik)
            li.setFixedSize(34,34); li.setAlignment(Qt.AlignCenter)
            li.setStyleSheet(f"font-size:18px;background:{r}22;border-radius:8px;border:none;")
            kl.addWidget(li)

            # Metin grubu
            txt = QVBoxLayout(); txt.setSpacing(1)
            ld = QLabel(str(d))
            fd = QFont("Segoe UI", 18, QFont.Bold); ld.setFont(fd)
            ld.setStyleSheet(f"color:{r};background:transparent;border:none;")
            lb = QLabel(b)
            lb.setStyleSheet(f"color:{COLORS['text_sec']};font-size:11px;background:transparent;border:none;")
            txt.addWidget(ld); txt.addWidget(lb)
            kl.addLayout(txt)

            self.stat_row.addWidget(kutu)

        self.lbl_son.setText(f"Son güncelleme: {datetime.now().strftime('%H:%M:%S')}")


# ═══════════════════════════════════════════════════════════════════════════════
# 🎨 TEMA SEÇİCİ DIALOG
# ═══════════════════════════════════════════════════════════════════════════════

class TemaSeciciDialog(QDialog):
    tema_degisti = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("🎨 Tema Seç"); self.setFixedSize(500, 360)
        self.setStyleSheet(ThemeManager.get_stylesheet()); self._build()

    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(24, 24, 24, 24); lay.setSpacing(16)
        b = QLabel("🎨 Renk Teması Seç")
        f = QFont(); f.setPointSize(14); f.setBold(True); b.setFont(f); lay.addWidget(b)

        grid = QGridLayout(); grid.setSpacing(12)
        self._tema_btns = {}
        for i, (tema_adi, renkler) in enumerate(TEMALAR.items()):
            kart = QPushButton()
            kart.setFixedHeight(70)
            aktif = tema_adi == TemaYoneticisi.aktif_tema()
            kart.setStyleSheet(f"""QPushButton{{
                background:qlineargradient(x1:0,y1:0,x2:1,y2:1,
                    stop:0 {renkler['bg_card']},stop:1 {renkler['bg_surface']});
                border:{'3px solid #ffd700' if aktif else f"2px solid {renkler['primary']}"};
                border-radius:10px;color:white;font-weight:bold;font-size:13px;}}
                QPushButton:hover{{border:3px solid {renkler['gold']};}}""")
            kart.setText(f"{'✓ ' if aktif else ''}{tema_adi}\n{renkler['primary']}")
            kart.clicked.connect(lambda ch=False, t=tema_adi: self._sec(t))
            self._tema_btns[tema_adi] = kart
            grid.addWidget(kart, i//2, i%2)
        lay.addLayout(grid)

        row = QHBoxLayout()
        bi = AdminLuxBtn("İptal", COLORS["danger"]); bi.clicked.connect(self.reject); row.addWidget(bi)
        row.addStretch()
        lay.addLayout(row)

    def _sec(self, tema_adi):
        self.tema_degisti.emit(tema_adi); self.accept()


# ═══════════════════════════════════════════════════════════════════════════════
# 📤 TABLO EXPORT DIALOG (CSV)
# ═══════════════════════════════════════════════════════════════════════════════

class TabloExportDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent); self.db = db
        self.setWindowTitle("📤 CSV Export"); self.setFixedSize(400, 300)
        self.setStyleSheet(ThemeManager.get_stylesheet()); self._build()

    def _build(self):
        lay = QVBoxLayout(self); lay.setContentsMargins(24, 24, 24, 24); lay.setSpacing(14)
        b = QLabel("📤 Tablo CSV Export")
        f = QFont(); f.setPointSize(13); f.setBold(True); b.setFont(f); lay.addWidget(b)
        lay.addWidget(QLabel("Hangi tabloyu dışa aktarmak istersiniz?"))
        self.cmb = QComboBox()
        tablolar = [
            ("ogrenciler","👥 Öğrenciler"), ("kurslar","📚 Kurslar"),
            ("egitmenler","👨‍🏫 Eğitmenler"), ("kayitlar","📝 Kayıtlar"),
            ("badge_kayitlar","🏅 Badge Kayıtları"), ("xp_hareketleri","⚡ XP Hareketleri"),
            ("sertifikalar","📜 Sertifikalar"), ("mentor_iliskileri","🧑‍🏫 Mentor İlişkileri"),
        ]
        for v,l in tablolar: self.cmb.addItem(l, userData=v)
        lay.addWidget(self.cmb)
        self.lbl_sonuc = QLabel("")
        self.lbl_sonuc.setStyleSheet(f"color:{COLORS['success']};font-size:12px;font-weight:bold;")
        lay.addWidget(self.lbl_sonuc)
        lay.addStretch()
        row = QHBoxLayout()
        bi = AdminLuxBtn("İptal", COLORS["danger"]); bi.clicked.connect(self.reject); row.addWidget(bi)
        row.addStretch()
        be = AdminLuxBtn("⬇️ CSV İndir", COLORS["success"]); be.clicked.connect(self._export); row.addWidget(be)
        lay.addLayout(row)

    def _export(self):
        tablo = self.cmb.currentData()
        dosya = f"{tablo}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
        ok, yol = self.db.tablo_csv_export(tablo, dosya)
        if ok:
            self.lbl_sonuc.setText(f"✅ Kaydedildi: {yol}")
            self.db.bildirim_ekle("CSV Export", f"{tablo} tablosu dışa aktarıldı.", "basari")
        else:
            self.lbl_sonuc.setStyleSheet(f"color:{COLORS['danger']};font-size:12px;")
            self.lbl_sonuc.setText(f"❌ {yol}")

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__(); self.db=AdminDB()
        self.setWindowTitle("🎓 Online Kurs Platformu — Admin Dashboard v6.0  |  TIER 6 ✨ FINAL")
        self.setGeometry(80,80,1400,860); self.setMinimumSize(1150,720)
        self.setStyleSheet(ThemeManager.get_stylesheet()); self._build()

    def _build(self):
        central=QWidget(); lay=QVBoxLayout(central); lay.setContentsMargins(0,0,0,0); lay.setSpacing(0)

        # ── HEADER ──────────────────────────────────────────────────────────
        header=QFrame()
        header.setStyleSheet(f"""QFrame{{background:qlineargradient(x1:0,y1:0,x2:1,y2:0,
            stop:0 {COLORS['bg_card']},stop:1 {COLORS['bg_surface']});
            border-bottom:2px solid {COLORS['primary']};border-radius:0px;}}""")
        header.setFixedHeight(62)
        hl=QHBoxLayout(header); hl.setContentsMargins(20,0,20,0); hl.setSpacing(10)

        lbl_logo=QLabel("🎓"); lbl_logo.setStyleSheet("font-size:24px;background:transparent;border:none;"); hl.addWidget(lbl_logo)
        ll=QLabel("Online Kurs Platformu")
        f=QFont(); f.setPointSize(15); f.setBold(True); ll.setFont(f)
        ll.setStyleSheet(f"color:{COLORS['text_main']};background:transparent;border:none;"); hl.addWidget(ll)
        self.lbl_tema_badge=QLabel(f"✨ {TemaYoneticisi.aktif_tema()}")
        self.lbl_tema_badge.setStyleSheet(f"color:{COLORS['gold']};font-size:10px;background:transparent;border:none;padding-left:8px;")
        hl.addWidget(self.lbl_tema_badge); hl.addStretch()

        # Header butonları
        for txt,fn,renk in [
            ("🎨 Tema",  self._tema_sec,                                      COLORS["accent"]),
            ("🔍 Ara",   lambda: GlobalAramaDialog(self.db,self).exec_(),     COLORS["primary"]),
            ("📤 CSV",   lambda: TabloExportDialog(self.db,self).exec_(),     COLORS["tier_pro2"]),
            ("🚪 Çıkış Yap", self._logout,                                   COLORS["danger"]),
        ]:
            btn=QPushButton(txt); btn.setFixedHeight(36)
            btn.setStyleSheet(f"""QPushButton{{background:{COLORS['bg_surface']};color:{COLORS['text_main']};
                border:1px solid {COLORS['border']};border-radius:8px;padding:0 14px;font-size:12px;font-weight:bold;}}
                QPushButton:hover{{background:{renk};color:white;border-color:{renk};}}""")
            btn.clicked.connect(fn); hl.addWidget(btn)

        self.btn_bildirim=AdminBildirimBtn(self.db)
        self.btn_bildirim.clicked.connect(lambda: (BildirimPanelDialog(self.db,self).exec_(), self.btn_bildirim._guncelle()))
        hl.addWidget(self.btn_bildirim)
        lay.addWidget(header)

        # ── CANLI DASHBOARD ŞERİDİ ──────────────────────────────────────────
        self.canli=CanliDashboard(self.db)
        self.canli.setFixedHeight(100)
        lay.addWidget(self.canli)

        # ── SEKMELER ────────────────────────────────────────────────────────
        self.tabs=QTabWidget()
        self.tabs.addTab(AdminDashboard(self.db),       "📊 Dashboard")
        self.tabs.addTab(KurslarSekmesi(self.db),         "📚 Kurslar")
        self.tabs.addTab(OgrencilerSekmesi(self.db),      "👥 Öğrenciler")
        self.tabs.addTab(EgitmenlerSekmesi(self.db),      "👨\u200d🏫 Eğitmenler")
        self.tabs.addTab(GamificationSekmesi(self.db),    "🎮 Gamification")
        self.tabs.addTab(MentorSertifikaSekmesi(self.db), "🧑\u200d🏫 Mentor & Sertifika")
        self.tabs.addTab(GrafikSekmesi(self.db),          "📈 Grafikler")
        self.tabs.addTab(RaporlarSekmesi(self.db),        "📋 Raporlar")
        self.tabs.addTab(AyarlarSekmesi(self.db),         "⚙️ Ayarlar")
        self.tabs.addTab(OgrenciGorünumSekmesi(self.db),  "🎓 Öğrenci Görünümü")
        lay.addWidget(self.tabs); self.setCentralWidget(central)

        # Kısayollar
        self._kisayol_kur()

        # Başlangıç bildirimi
        self.db.bildirim_ekle("Hoş Geldiniz ✨","Platform TIER 6 Final başarıyla yüklendi!","basari")

    def _kisayol_kur(self):
        from PyQt5.QtWidgets import QShortcut
        from PyQt5.QtGui import QKeySequence
        for tus,fn in [
            ("Ctrl+F", lambda: GlobalAramaDialog(self.db,self).exec_()),
            ("Ctrl+T", self._tema_sec),
            ("Ctrl+E", lambda: TabloExportDialog(self.db,self).exec_()),
            ("Ctrl+B", lambda: (BildirimPanelDialog(self.db,self).exec_(), self.btn_bildirim._guncelle())),
            ("Ctrl+1", lambda: self.tabs.setCurrentIndex(0)),
            ("Ctrl+2", lambda: self.tabs.setCurrentIndex(1)),
            ("Ctrl+3", lambda: self.tabs.setCurrentIndex(2)),
            ("Ctrl+4", lambda: self.tabs.setCurrentIndex(3)),
            ("Ctrl+5", lambda: self.tabs.setCurrentIndex(4)),
            ("Ctrl+6", lambda: self.tabs.setCurrentIndex(5)),
            ("Ctrl+7", lambda: self.tabs.setCurrentIndex(6)),
            ("Ctrl+8", lambda: self.tabs.setCurrentIndex(7)),
            ("Ctrl+9", lambda: self.tabs.setCurrentIndex(8)),
        ]:
            sc=QShortcut(QKeySequence(tus),self); sc.activated.connect(fn)

    def _logout(self):
        cevap = QMessageBox.question(self, "Çıkış Yap",
            "Giriş ekranına dönmek istiyor musunuz?",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if cevap == QMessageBox.Yes:
            giris = getattr(self, "_giris_ekrani_ref", None)
            if giris:
                giris.show()
                giris.raise_()
                giris.activateWindow()
            self.close()

    def _tema_sec(self):
        dlg=TemaSeciciDialog(self)
        dlg.tema_degisti.connect(self._tema_uygula)
        dlg.exec_()

    def _tema_uygula(self,tema_adi):
        TemaYoneticisi.tema_uygula(tema_adi,QApplication.instance())
        self.lbl_tema_badge.setText(f"✨ {tema_adi}")
        self.lbl_tema_badge.setStyleSheet(f"color:{COLORS['gold']};font-size:10px;background:transparent;border:none;padding-left:8px;")
        self.db.ayar_kaydet("aktif_tema",tema_adi)
        self.db.bildirim_ekle("Tema Değiştirildi",f"Yeni tema: {tema_adi}","bilgi")
        self.btn_bildirim._guncelle()




# ═══════════════════════════════════════════════════════════════════════════════
# 🔘 LUXURY BUTTON (Yemek Platformu animasyonlu stili)
# ═══════════════════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════════════
# 🏠 LUXURY KART (Yemek Platformu dashboard kartı)
# ═══════════════════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════════════
# 💫 SPLASH SCREEN (Öğrenci versiyonu)
# ═══════════════════════════════════════════════════════════════════════════════




# ═══════════════════════════════════════════════════════════════════════════════
# 🔐 LOGIN / KAYIT DIALOG (Yemek Platformu tarzı)
# ═══════════════════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════════════════
# ⚡ XP PROGRESS BAR WIDGET
# ═══════════════════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════════════════
# 🏅 BADGE KARTI
# ═══════════════════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════════════════
# 📊 KURS KARTI (grid görünümü için)
# ═══════════════════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════════════
# 📊 DASHBOARD SEKMESİ
# ═══════════════════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════════════════
# 📚 KATALOG SEKMESİ
# ═══════════════════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════════════════
# 📖 KURS DETAY DIALOG (Yemek Platformu TarifDetayDialog tarzı)
# ═══════════════════════════════════════════════════════════════════════════════


# ═══════════════════════════════════════════════════════════════════════════════
# 📋 KURSLARIM SEKMESİ
# ═══════════════════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════════════════
# 🏅 ROZETLER SEKMESİ (Yemek Platformu kart grid stili)
# ═══════════════════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════════════════
# 👤 PROFİL SEKMESİ
# ═══════════════════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════════════════
# 📊 İSTATİSTİK SEKMESİ (Matplotlib grafikler — Yemek Platformu StatisticsWidget stili)
# ═══════════════════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════════════════
# 🔔 BİLDİRİM BUTONU (header için)
# ═══════════════════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════════════════
# 🔔 BİLDİRİM PANEL (Yemek Platformu tarzı kart)
# ═══════════════════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════════════════
# 🏠 ANA PENCERE (Öğrenci Portalı)
# ═══════════════════════════════════════════════════════════════════════════════



# ═══════════════════════════════════════════════════════════════════════════════
# 🚀 MAIN
# ═══════════════════════════════════════════════════════════════════════════════




# ══════════════════════════════════════════════════════════════════════════════
# 🎓 GİRİŞ EKRANI — Admin veya Öğrenci şifreye göre otomatik açılır
# ══════════════════════════════════════════════════════════════════════════════

BG_GIRIS   = "#0f0f1a"
CARD_GIRIS = "#1a1a2e"
SURF_GIRIS = "#2d2d3a"
ACC_GIRIS  = "#f5a623"
GRN_GIRIS  = "#4CAF50"
RED_GIRIS  = "#e63946"

GLOBAL_SS = f"""
QWidget       {{ background-color:{BG_GIRIS}; color:#ffffff; font-family:'Segoe UI',sans-serif; font-size:13px; }}
QFrame        {{ background-color:{CARD_GIRIS}; }}
QLabel        {{ background:transparent; color:#ffffff; }}
QLineEdit     {{ background:{SURF_GIRIS}; color:#ffffff; border:2px solid {ACC_GIRIS};
                 border-radius:10px; padding:11px 14px; font-size:13px; }}
QLineEdit:focus {{ border-color:{GRN_GIRIS}; }}
QPushButton   {{ background:{ACC_GIRIS}; color:#1a1a2e; border:none;
                 border-radius:10px; font-weight:bold; font-size:13px; padding:11px; }}
QPushButton:hover  {{ background:#e09018; }}
QPushButton:pressed{{ background:#c07810; }}
QPushButton:disabled{{ background:{SURF_GIRIS}; color:#666; }}
QTabWidget::pane  {{ border:none; background:{BG_GIRIS}; }}
QTabBar::tab  {{ background:{SURF_GIRIS}; color:#aaaaaa; padding:12px 28px;
                 font-size:13px; font-weight:bold; border:none;
                 border-bottom:3px solid transparent; margin-right:2px; }}
QTabBar::tab:selected  {{ background:{CARD_GIRIS}; color:{ACC_GIRIS}; border-bottom:3px solid {ACC_GIRIS}; }}
QTabBar::tab:hover:!selected {{ color:#ffffff; }}
QTableWidget  {{ background:{CARD_GIRIS}; color:#ffffff; gridline-color:{SURF_GIRIS};
                 border:none; alternate-background-color:{SURF_GIRIS}; font-size:13px; }}
QTableWidget::item {{ padding:9px; }}
QTableWidget::item:selected {{ background:{ACC_GIRIS}; color:#1a1a2e; }}
QHeaderView::section {{ background:{BG_GIRIS}; color:{ACC_GIRIS}; font-weight:bold; padding:11px; border:none; font-size:13px; }}
QGroupBox     {{ background:{CARD_GIRIS}; border:1px solid {SURF_GIRIS}; border-radius:10px;
                 margin-top:14px; padding-top:12px; color:{ACC_GIRIS}; font-weight:bold; font-size:13px; }}
QGroupBox::title {{ subcontrol-origin:margin; left:12px; padding:0 6px; }}
QProgressBar  {{ background:{SURF_GIRIS}; border:none; border-radius:7px; height:14px; color:white; font-size:11px; text-align:center; }}
QProgressBar::chunk {{ background:{ACC_GIRIS}; border-radius:7px; }}
QScrollBar:vertical {{ background:{SURF_GIRIS}; width:9px; border-radius:4px; }}
QScrollBar::handle:vertical {{ background:{ACC_GIRIS}; border-radius:4px; min-height:20px; }}
QScrollBar::add-line:vertical,QScrollBar::sub-line:vertical {{ height:0; }}
QScrollBar:horizontal {{ background:{SURF_GIRIS}; height:9px; border-radius:4px; }}
QScrollBar::handle:horizontal {{ background:{ACC_GIRIS}; border-radius:4px; }}
QListWidget   {{ background:{CARD_GIRIS}; color:#ffffff; border:none; border-radius:8px; font-size:13px; }}
QListWidget::item {{ padding:9px; border-bottom:1px solid {SURF_GIRIS}; }}
QListWidget::item:selected {{ background:{ACC_GIRIS}; color:#1a1a2e; }}
QTextEdit     {{ background:{SURF_GIRIS}; color:#ffffff; border:1px solid {SURF_GIRIS}; border-radius:8px; padding:10px; font-size:13px; }}
QComboBox     {{ background:{SURF_GIRIS}; color:#ffffff; border:1px solid {SURF_GIRIS}; border-radius:6px; padding:8px; font-size:13px; }}
QComboBox:focus {{ border:2px solid {ACC_GIRIS}; }}
QComboBox::drop-down {{ border:none; width:20px; }}
QComboBox QAbstractItemView {{ background:{CARD_GIRIS}; color:#ffffff; selection-background-color:{ACC_GIRIS}; selection-color:#1a1a2e; }}
QCheckBox     {{ color:{ACC_GIRIS}; spacing:8px; font-size:13px; }}
QCheckBox::indicator {{ width:16px; height:16px; border:2px solid {ACC_GIRIS}; border-radius:4px; background:{SURF_GIRIS}; }}
QCheckBox::indicator:checked {{ background:{ACC_GIRIS}; }}
QSpinBox,QDoubleSpinBox {{ background:{SURF_GIRIS}; color:#ffffff; border:1px solid {SURF_GIRIS}; border-radius:6px; padding:7px; font-size:13px; }}
QSpinBox:focus,QDoubleSpinBox:focus {{ border:2px solid {ACC_GIRIS}; }}
QSplitter::handle {{ background:{SURF_GIRIS}; }}
QMenu         {{ background:{CARD_GIRIS}; color:#ffffff; border:1px solid {SURF_GIRIS}; border-radius:6px; font-size:13px; }}
QMenu::item   {{ padding:9px 22px; }}
QMenu::item:selected {{ background:{ACC_GIRIS}; color:#1a1a2e; }}
QScrollArea   {{ border:none; background:transparent; }}
QMessageBox   {{ background:{CARD_GIRIS}; }}
QMessageBox QLabel {{ color:#ffffff; font-size:13px; min-width:260px; }}
QMessageBox QPushButton {{ background:{ACC_GIRIS}; color:#1a1a2e; border:none; border-radius:8px;
    padding:8px 22px; font-weight:bold; font-size:13px; min-width:80px; }}
QDialog       {{ background:{BG_GIRIS}; }}
QToolButton   {{ background:transparent; border:none; color:#ffffff; font-size:13px; }}
"""

class GirisEkrani(QMainWindow):
    def __init__(self):
        super().__init__()
        self.admin_db = AdminDB()   # sample data burada yüklenir
        self.setWindowTitle("🎓 Online Kurs Platformu — Admin Girişi")
        self.setFixedSize(540, 600)
        self._build()

    def _build(self):
        root = QWidget()
        rl = QVBoxLayout(root); rl.setContentsMargins(0,0,0,0); rl.setSpacing(0)
        rl.setAlignment(Qt.AlignCenter)

        # ── Merkez kart ───────────────────────────────────────────────────
        outer = QWidget()
        ol = QVBoxLayout(outer); ol.setAlignment(Qt.AlignCenter)
        ol.setContentsMargins(40,40,40,30); ol.setSpacing(0)

        kart = QFrame()
        kart.setStyleSheet(f"""QFrame{{
            background:{CARD_GIRIS};
            border-radius:24px;
            border:1.5px solid {ACC_GIRIS};
        }}""")
        kl = QVBoxLayout(kart); kl.setContentsMargins(48,40,48,40); kl.setSpacing(0)
        kl.setAlignment(Qt.AlignCenter)

        # Logo + başlık
        logo = QLabel("🎓"); logo.setAlignment(Qt.AlignCenter)
        logo.setStyleSheet("font-size:48px;background:transparent;border:none;")
        kl.addWidget(logo); kl.addSpacing(10)

        t1 = QLabel("ONLİNE KURS PLATFORMU")
        f1 = QFont("Segoe UI", 20, QFont.Bold); t1.setFont(f1)
        t1.setAlignment(Qt.AlignCenter)
        t1.setStyleSheet(f"color:{ACC_GIRIS};background:transparent;border:none;letter-spacing:2px;")
        kl.addWidget(t1); kl.addSpacing(6)

        t2 = QLabel("Admin Girişi")
        t2.setAlignment(Qt.AlignCenter)
        t2.setStyleSheet(f"color:{GRN_GIRIS};font-size:13px;background:transparent;border:none;letter-spacing:1px;")
        kl.addWidget(t2); kl.addSpacing(28)

        # Ayraç
        line = QFrame(); line.setFrameShape(QFrame.HLine); line.setFixedHeight(1)
        line.setStyleSheet(f"background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 transparent,stop:0.5 {ACC_GIRIS},stop:1 transparent);border:none;max-height:1px;")
        kl.addWidget(line); kl.addSpacing(28)

        # Input stili
        INP_SS = f"""QLineEdit{{
            background:{SURF_GIRIS};color:#ffffff;
            border:2px solid {BG_GIRIS};
            border-radius:12px;padding:13px 16px;font-size:14px;
        }}QLineEdit:focus{{border:2px solid {ACC_GIRIS};}}"""

        lbl_u = QLabel("👤  Kullanıcı Adı")
        lbl_u.setStyleSheet(f"color:{COLORS['text_sec']};font-size:12px;font-weight:600;background:transparent;border:none;letter-spacing:0.5px;")
        kl.addWidget(lbl_u); kl.addSpacing(6)
        self.inp_user = QLineEdit()
        self.inp_user.setPlaceholderText("Kullanıcı adınızı girin")
        self.inp_user.setStyleSheet(INP_SS)
        self.inp_user.setFixedHeight(50); kl.addWidget(self.inp_user); kl.addSpacing(16)

        lbl_p = QLabel("🔒  Şifre")
        lbl_p.setStyleSheet(f"color:{COLORS['text_sec']};font-size:12px;font-weight:600;background:transparent;border:none;letter-spacing:0.5px;")
        kl.addWidget(lbl_p); kl.addSpacing(6)
        self.inp_pass = QLineEdit()
        self.inp_pass.setPlaceholderText("Şifrenizi girin")
        self.inp_pass.setEchoMode(QLineEdit.Password)
        self.inp_pass.setStyleSheet(INP_SS)
        self.inp_pass.setFixedHeight(50); kl.addWidget(self.inp_pass)
        self.inp_pass.returnPressed.connect(self._giris)
        kl.addSpacing(10)

        self.lbl_hata = QLabel("")
        self.lbl_hata.setAlignment(Qt.AlignCenter)
        self.lbl_hata.setFixedHeight(22)
        self.lbl_hata.setStyleSheet(f"color:{RED_GIRIS};font-size:12px;background:transparent;border:none;")
        kl.addWidget(self.lbl_hata); kl.addSpacing(14)

        btn = QPushButton("  GİRİŞ YAP  →")
        btn.setFixedHeight(52)
        btn.setStyleSheet(f"""QPushButton{{
            background:{ACC_GIRIS};color:#1a1a2e;
            border:none;border-radius:14px;
            font-weight:bold;font-size:16px;letter-spacing:1px;
        }}QPushButton:hover{{background:#ffb733;}}
        QPushButton:pressed{{background:#e09018;}}""")
        btn.clicked.connect(self._giris); kl.addWidget(btn)
        kl.addSpacing(20)

        # Demo bilgisi — kart içinde
        demo = QLabel("admin  /  admin123")
        demo.setAlignment(Qt.AlignCenter)
        demo.setStyleSheet(f"color:{COLORS['border']};font-size:11px;background:transparent;border:none;")
        kl.addWidget(demo)

        ol.addWidget(kart)
        rl.addWidget(outer)
        self.setCentralWidget(root)

    def closeEvent(self, e):
        # GirisEkrani kapanırsa uygulama biter
        QApplication.quit()
        e.accept()

    def _giris(self):
        user = self.inp_user.text().strip()
        pw   = self.inp_pass.text()
        if not user or not pw:
            self.lbl_hata.setText("Kullanıcı adı ve şifre boş olamaz!"); return

        if self.admin_db.dogrula_admin(user, pw):
            self._admin_ac(); return

        self.lbl_hata.setText("❌  Kullanıcı adı veya şifre yanlış!")
        self.inp_pass.clear(); self.inp_pass.setFocus()

    def _admin_ac(self):
        try:
            splash = AdminSplash(); splash.exec_()
            win = MainWindow()
            win._giris_ekrani_ref = self
            win.show()
            self.hide()
            self.inp_pass.clear()
            self.lbl_hata.setText("")
        except Exception as ex:
            import traceback
            traceback.print_exc()
            self.lbl_hata.setText(f"❌ Hata: {ex}")

    # Öğrenci girişi artık admin paneli içinden yönetiliyor


def main():
    app = QApplication(sys.argv); app.setStyle("Fusion")
    app.setStyleSheet(GLOBAL_SS)
    app.setQuitOnLastWindowClosed(False)  # panel kapanınca uygulama bitmesin
    win = GirisEkrani(); win.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()